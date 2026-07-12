import os
import sys
import shutil
import subprocess
import threading
import time
import hashlib
import stat
import zipfile
import tarfile
import json
import re
import difflib
import fnmatch
import ctypes
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog, font as tkfont

# Optional libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import send2trash
    TRASH_AVAILABLE = True
except ImportError:
    TRASH_AVAILABLE = False

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    WATCHDOG_AVAILABLE = True
except ImportError:
    WATCHDOG_AVAILABLE = False

# ── DPI / Screen scaling ───────────────────────────────────────────────────────
def get_screen_info():
    """Return screen width, height, and a scale factor."""
    try:
        root = tk.Tk()
        root.withdraw()
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        root.destroy()
    except:
        sw, sh = 1920, 1080
    scale = max(0.7, min(1.5, sw / 1440))
    return sw, sh, scale

SCREEN_W, SCREEN_H, SCALE = get_screen_info()

def scaled(value):
    """Scale a pixel value by the screen factor."""
    return max(1, int(value * SCALE))

# ── Modern UI Color Palettes ────────────────────────────────────────────────────
DARK_PALETTE = {
    "BG":         "#0a0e17",
    "PANEL":      "#111827",
    "SIDEBAR":    "#0f172a",
    "ACCENT":     "#6366f1",
    "ACCENT2":    "#8b5cf6",
    "ACCENT3":    "#06b6d4",
    "SUCCESS":    "#10b981",
    "WARNING":    "#f59e0b",
    "DANGER":     "#ef4444",
    "TEXT":       "#f1f5f9",
    "SUBTEXT":    "#94a3b8",
    "BORDER":     "#1e293b",
    "HOVER":      "#1e293b",
    "SEL":        "#312e81",
    "SEL_FG":     "#a5b4fc",
    "HEADER_BG":  "#020617",
    "EVEN_ROW":   "#1e293b",
    "ODD_ROW":    "#111827",
    "TAG_BG":     "#1e1040",
    "TAG_FG":     "#a78bfa",
    "BULK_BG":    "#0f172a",
    "BULK_BORDER":"#1e3a5f",
    "CARD":       "#1e293b",
    "CARD_BORDER":"#334155",
    "INPUT_BG":   "#0f172a",
    "BTN_FG":     "#ffffff",
    "SHADOW":     "#00000060",
    "GRADIENT_START": "#0a0e17",
    "GRADIENT_END": "#111827",
}

LIGHT_PALETTE = {
    "BG":         "#f1f5f9",
    "PANEL":      "#ffffff",
    "SIDEBAR":    "#f8fafc",
    "ACCENT":     "#4f46e5",
    "ACCENT2":    "#7c3aed",
    "ACCENT3":    "#0891b2",
    "SUCCESS":    "#059669",
    "WARNING":    "#d97706",
    "DANGER":     "#dc2626",
    "TEXT":       "#0f172a",
    "SUBTEXT":    "#64748b",
    "BORDER":     "#e2e8f0",
    "HOVER":      "#f1f5f9",
    "SEL":        "#e0e7ff",
    "SEL_FG":     "#4338ca",
    "HEADER_BG":  "#f8fafc",
    "EVEN_ROW":   "#f8fafc",
    "ODD_ROW":    "#f1f5f9",
    "TAG_BG":     "#ede9fe",
    "TAG_FG":     "#5b21b6",
    "BULK_BG":    "#f0f9ff",
    "BULK_BORDER":"#bae6fd",
    "CARD":       "#ffffff",
    "CARD_BORDER":"#e2e8f0",
    "INPUT_BG":   "#f8fafc",
    "BTN_FG":     "#ffffff",
    "SHADOW":     "#00000015",
    "GRADIENT_START": "#f1f5f9",
    "GRADIENT_END": "#ffffff",
}

P = dict(DARK_PALETTE)
CONFIG_FILE = os.path.join(Path.home(), ".nexfile_config.json")

# ── Theme-aware Mark Colors ────────────────────────────────────────────────────
MARK_COLORS_DARK = {
    "red":    {"bg": "#3a1010", "fg": "#f87171"},
    "orange": {"bg": "#2e1a08", "fg": "#fb923c"},
    "yellow": {"bg": "#2a2000", "fg": "#facc15"},
    "green":  {"bg": "#0a2a18", "fg": "#4ade80"},
    "blue":   {"bg": "#0a1a3a", "fg": "#60a5fa"},
    "purple": {"bg": "#1a0a30", "fg": "#c084fc"},
    "pink":   {"bg": "#2a0a18", "fg": "#f472b6"},
    "cyan":   {"bg": "#001e20", "fg": "#22d3ee"},
    "none":   {"bg": None,      "fg": None},
}
MARK_COLORS_LIGHT = {
    "red":    {"bg": "#fee2e2", "fg": "#b91c1c"},
    "orange": {"bg": "#ffedd5", "fg": "#c2410c"},
    "yellow": {"bg": "#fef9c3", "fg": "#854d0e"},
    "green":  {"bg": "#dcfce7", "fg": "#15803d"},
    "blue":   {"bg": "#dbeafe", "fg": "#1d4ed8"},
    "purple": {"bg": "#f3e8ff", "fg": "#6d28d9"},
    "pink":   {"bg": "#fce7f3", "fg": "#be185d"},
    "cyan":   {"bg": "#cffafe", "fg": "#0e7490"},
    "none":   {"bg": None,      "fg": None},
}
MARK_COLORS = dict(MARK_COLORS_DARK)

EXT_ICONS = {
    ".py": "🐍", ".pyw": "🐍", ".pyx": "🐍",
    ".html": "🌐", ".htm": "🌐", ".css": "🎨", ".scss": "🎨", ".sass": "🎨", ".less": "🎨",
    ".js": "📜", ".mjs": "📜", ".cjs": "📜", ".ts": "🔷", ".tsx": "🔷", ".jsx": "⚛️",
    ".cpp": "⚙️", ".c": "⚙️", ".h": "⚙️", ".hpp": "⚙️", ".cs": "🔵", ".java": "☕",
    ".go": "🐹", ".rs": "🦀", ".swift": "🍎", ".kt": "🔶", ".scala": "♾️",
    ".php": "🐘", ".rb": "💎", ".pl": "🦪", ".lua": "🌙", ".r": "📊",
    ".json": "📋", ".yaml": "📋", ".yml": "📋", ".toml": "📋", ".xml": "📋",
    ".ini": "⚙️", ".cfg": "⚙️", ".conf": "⚙️", ".env": "🔐", ".properties": "⚙️",
    ".txt": "📄", ".md": "📝", ".rst": "📝", ".log": "📃", ".csv": "📊",
    ".pdf": "📕", ".doc": "📘", ".docx": "📘", ".odt": "📘",
    ".xls": "📗", ".xlsx": "📗", ".ods": "📗",
    ".ppt": "📙", ".pptx": "📙", ".odp": "📙",
    ".jpg": "🖼️", ".jpeg": "🖼️", ".png": "🖼️", ".gif": "🖼️",
    ".bmp": "🖼️", ".svg": "🖼️", ".webp": "🖼️", ".ico": "🖼️", ".tiff": "🖼️",
    ".mp3": "🎵", ".wav": "🎵", ".flac": "🎵", ".aac": "🎵", ".ogg": "🎵", ".m4a": "🎵",
    ".mp4": "🎬", ".avi": "🎬", ".mkv": "🎬", ".mov": "🎬", ".webm": "🎬", ".flv": "🎬",
    ".zip": "📦", ".tar": "📦", ".gz": "📦", ".bz2": "📦", ".7z": "📦",
    ".rar": "📦", ".xz": "📦", ".zst": "📦",
    ".sh": "💻", ".bash": "💻", ".zsh": "💻", ".bat": "💻", ".ps1": "💻", ".cmd": "💻",
    ".exe": "⚡", ".msi": "⚡", ".dmg": "⚡", ".deb": "⚡", ".rpm": "⚡",
    ".db": "🗄️", ".sqlite": "🗄️", ".sqlite3": "🗄️", ".sql": "🗄️",
    ".pem": "🔑", ".key": "🔑", ".crt": "🔑", ".cert": "🔑",
    ".ttf": "🔤", ".otf": "🔤", ".woff": "🔤", ".woff2": "🔤",
    ".psd": "🎭", ".ai": "🎭", ".sketch": "🎭", ".fig": "🎭",
    ".ipynb": "📓",
}

DEFAULT_EXPORT_DIR = os.path.join(Path.home(), "NexFile_Exports")

def load_config():
    try:
        with open(CONFIG_FILE, "r") as f:
            return json.load(f)
    except:
        return {}

def save_config(data):
    try:
        with open(CONFIG_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except:
        pass

def human_size(n):
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} PB"

def file_icon(path):
    if os.path.islink(path): return "🔗"
    if os.path.isdir(path): return "📁"
    name = os.path.basename(path).lower()
    if name == "dockerfile": return "🐳"
    if name in ("makefile", "rakefile", "gruntfile"): return "🔨"
    if name in (".gitignore", ".gitattributes", ".gitmodules"): return "🔧"
    if name in ("readme.md", "readme.txt", "readme"): return "📖"
    ext = Path(path).suffix.lower()
    return EXT_ICONS.get(ext, "📄")

CODE_EXTS = {".py", ".js", ".ts", ".cpp", ".c", ".h", ".java", ".go", ".rs",
             ".txt", ".md", ".json", ".yaml", ".yml", ".toml", ".xml",
             ".sh", ".bash", ".zsh", ".html", ".css", ".sql", ".log",
             ".ini", ".cfg", ".rst", ".bat", ".ps1", ".env", ".gitignore",
             ".tsx", ".jsx", ".scss", ".less", ".lua", ".rb", ".php",
             ".swift", ".kt", ".cs", ".r", ".pl", ".dockerfile"}
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}

def can_preview(path):
    return Path(path).suffix.lower() in CODE_EXTS

def is_image(path):
    return Path(path).suffix.lower() in IMAGE_EXTS

def ensure_export_dir(export_dir):
    try:
        os.makedirs(export_dir, exist_ok=True)
        return True
    except:
        return False

def make_export_filename(folder_name, ext, export_dir):
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    safe_name = re.sub(r'[\\/:*?"<>|]', '_', folder_name)
    filename = f"{safe_name}_{ts}.{ext}"
    return os.path.join(export_dir, filename)

def scan_directory_recursive(root_path, ext_filter="", show_hidden=False,
                              stop_event=None, progress_cb=None):
    results = []
    ext_f = ext_filter.strip().lstrip(".").lower()
    try:
        for dirpath, dirs, files in os.walk(root_path, followlinks=False):
            if stop_event and stop_event.is_set():
                break
            if not show_hidden:
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                files = [f for f in files if not f.startswith(".")]
            if progress_cb:
                progress_cb(f"Scanning: {dirpath[:80]}")
            for fname in files:
                if ext_f and not fname.lower().endswith("." + ext_f):
                    continue
                fpath = os.path.join(dirpath, fname)
                try:
                    st = os.stat(fpath)
                    results.append({
                        "name": fname, "full_path": fpath,
                        "size": human_size(st.st_size), "size_bytes": st.st_size,
                        "modified": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                        "type": Path(fname).suffix.upper().lstrip(".") + " File" if Path(fname).suffix else "File",
                    })
                except (PermissionError, OSError):
                    results.append({"name": fname, "full_path": fpath,
                                    "size": "N/A", "size_bytes": 0, "modified": "N/A", "type": "File"})
    except (PermissionError, OSError):
        pass
    return results

def parse_bulk_structure(text, base_path):
    lines = text.splitlines()
    items = []
    for raw in lines:
        stripped = raw.rstrip()
        if not stripped:
            continue
        content = stripped.lstrip()
        indent = len(stripped) - len(content)
        is_dir = content.endswith("/") or content.endswith("\\")
        name = content.rstrip("/\\")
        if not name:
            continue
        items.append((indent, name, is_dir))

    resolved = []
    stack = []
    for indent, name, is_dir in items:
        while stack and stack[-1][0] >= indent:
            stack.pop()
        parent = stack[-1][1] if stack else base_path
        abs_path = os.path.join(parent, name)
        resolved.append((abs_path, is_dir))
        if is_dir:
            stack.append((indent, abs_path))
    return resolved

# ── Modern styled helper builders ───────────────────────────────────────────
def make_card(parent, padding=12, rounded=True):
    """Return a card-style frame with border and optional rounded corners."""
    outer = tk.Frame(parent, bg=P["CARD_BORDER"], bd=0)
    inner = tk.Frame(outer, bg=P["CARD"], padx=padding, pady=padding)
    inner.pack(padx=1, pady=1, fill="both", expand=True)
    return outer, inner

def make_button(parent, text, command, style="primary", font_tuple=None, icon="", **kwargs):
    """Create a modern styled flat button with icon support."""
    bg_map = {
        "primary":   P["ACCENT"],
        "secondary": P["CARD"],
        "danger":    P["DANGER"],
        "success":   P["SUCCESS"],
        "warning":   P["WARNING"],
        "ghost":     P["BG"],
        "gradient":  P["ACCENT"],
    }
    fg_map = {
        "primary":   "#ffffff",
        "secondary": P["TEXT"],
        "danger":    "#ffffff",
        "success":   "#ffffff",
        "warning":   "#ffffff",
        "ghost":     P["SUBTEXT"],
        "gradient":  "#ffffff",
    }
    hover_map = {
        "primary":   P["ACCENT2"],
        "secondary": P["HOVER"],
        "danger":    "#b91c1c",
        "success":   "#059669",
        "warning":   "#b45309",
        "ghost":     P["HOVER"],
        "gradient":  P["ACCENT2"],
    }
    bg = bg_map.get(style, P["ACCENT"])
    fg = fg_map.get(style, "#ffffff")
    hover = hover_map.get(style, P["ACCENT2"])
    if font_tuple is None:
        font_tuple = ("Segoe UI", scaled(9))
    btn_text = f"{icon} {text}" if icon else text
    btn = tk.Button(parent, text=btn_text, command=command,
                    bg=bg, fg=fg,
                    activebackground=hover, activeforeground=fg,
                    relief="flat", bd=0,
                    font=font_tuple, cursor="hand2", **kwargs)
    return btn

# ── Tooltip with modern styling ──────────────────────────────────────────────
class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip = None
        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, _=None):
        if not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + 24
        self.tip = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tip_bg = P["CARD"] if P == LIGHT_PALETTE else "#1e293b"
        tip_fg = P["TEXT"]
        tip_border = P["CARD_BORDER"]
        outer = tk.Frame(tw, bg=tip_border, bd=0)
        outer.pack(padx=0, pady=0)
        tk.Label(outer, text=self.text,
                 background=tip_bg, foreground=tip_fg,
                 relief="flat", padx=12, pady=6,
                 font=("Segoe UI", scaled(9))).pack(padx=1, pady=1)

    def hide(self, _=None):
        if self.tip:
            self.tip.destroy()
            self.tip = None

# ── Modern Search Bar Widget ──────────────────────────────────────────────────
class ModernSearchBar(tk.Frame):
    def __init__(self, parent, placeholder="Search files...", command=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.command = command
        self.configure(bg=P["INPUT_BG"], highlightthickness=1, 
                       highlightbackground=P["BORDER"], highlightcolor=P["ACCENT"])
        
        self.search_icon = tk.Label(self, text="🔍", bg=P["INPUT_BG"], fg=P["SUBTEXT"])
        self.search_icon.pack(side="left", padx=(8, 4))
        
        self.entry = tk.Entry(self, bg=P["INPUT_BG"], fg=P["TEXT"],
                              insertbackground=P["ACCENT"], relief="flat",
                              font=("Segoe UI", scaled(9)),
                              selectbackground=P["SEL"], selectforeground=P["SEL_FG"])
        self.entry.pack(side="left", fill="x", expand=True, ipady=4)
        self.entry.insert(0, placeholder)
        self.entry.config(fg=P["SUBTEXT"])
        self.entry.bind("<FocusIn>", self._on_focus_in)
        self.entry.bind("<FocusOut>", self._on_focus_out)
        self.entry.bind("<Return>", lambda e: self._execute())
        
        self.clear_btn = tk.Button(self, text="✕", bg=P["INPUT_BG"], fg=P["SUBTEXT"],
                                   relief="flat", bd=0, cursor="hand2",
                                   command=self.clear)
        self.clear_btn.pack(side="right", padx=(0, 8))
        self.clear_btn.pack_forget()
        
        self.entry.bind("<KeyRelease>", self._on_key_release)
        self.placeholder = placeholder
        
    def _on_focus_in(self, e):
        if self.entry.get() == self.placeholder:
            self.entry.delete(0, "end")
            self.entry.config(fg=P["TEXT"])
    
    def _on_focus_out(self, e):
        if not self.entry.get().strip():
            self.entry.insert(0, self.placeholder)
            self.entry.config(fg=P["SUBTEXT"])
            self.clear_btn.pack_forget()
    
    def _on_key_release(self, e):
        text = self.entry.get()
        if text and text != self.placeholder:
            self.clear_btn.pack(side="right", padx=(0, 8))
        else:
            self.clear_btn.pack_forget()
        self._execute()
    
    def _execute(self):
        if self.command:
            text = self.entry.get()
            if text and text != self.placeholder:
                self.command(text)
            else:
                self.command("")
    
    def get(self):
        text = self.entry.get()
        return "" if text == self.placeholder else text
    
    def set(self, text):
        self.entry.delete(0, "end")
        if text:
            self.entry.insert(0, text)
            self.entry.config(fg=P["TEXT"])
            self.clear_btn.pack(side="right", padx=(0, 8))
        else:
            self.entry.insert(0, self.placeholder)
            self.entry.config(fg=P["SUBTEXT"])
            self.clear_btn.pack_forget()
    
    def clear(self):
        self.entry.delete(0, "end")
        self.entry.insert(0, self.placeholder)
        self.entry.config(fg=P["SUBTEXT"])
        self.clear_btn.pack_forget()
        if self.command:
            self.command("")

# ── File Watcher ───────────────────────────────────────────────────────────────
if WATCHDOG_AVAILABLE:
    class DirWatcher(FileSystemEventHandler):
        def __init__(self, callback):
            self.callback = callback
            self._last = 0

        def on_any_event(self, event):
            now = time.time()
            if now - self._last > 1.0:
                self._last = now
                self.callback()

# ── Tag Manager ────────────────────────────────────────────────────────────────
class TagManager:
    def __init__(self):
        cfg = load_config()
        self.tags = cfg.get("file_tags", {})
        self.colors = cfg.get("tag_colors", {
            "red": "#ef4444", "green": "#22c55e", "blue": "#3b82f6",
            "yellow": "#eab308", "purple": "#a855f7", "pink": "#ec4899",
            "orange": "#f97316", "cyan": "#06b6d4"
        })

    def get_tags(self, path):
        return self.tags.get(path, [])

    def add_tag(self, path, tag):
        if path not in self.tags:
            self.tags[path] = []
        if tag not in self.tags[path]:
            self.tags[path].append(tag)
        self._save()

    def remove_tag(self, path, tag):
        if path in self.tags and tag in self.tags[path]:
            self.tags[path].remove(tag)
            if not self.tags[path]:
                del self.tags[path]
        self._save()

    def _save(self):
        cfg = load_config()
        cfg["file_tags"] = self.tags
        cfg["tag_colors"] = self.colors
        save_config(cfg)

# ── Color Mark Manager ─────────────────────────────────────────────────────────
class ColorMarkManager:
    def __init__(self):
        cfg = load_config()
        self.marks = cfg.get("color_marks", {})

    def get_mark(self, path):
        return self.marks.get(path, "none")

    def set_mark(self, path, color_name):
        if color_name == "none":
            self.marks.pop(path, None)
        else:
            self.marks[path] = color_name
        self._save()

    def _save(self):
        cfg = load_config()
        cfg["color_marks"] = self.marks
        save_config(cfg)

# ── Undo/Redo Stack ────────────────────────────────────────────────────────────
class UndoStack:
    def __init__(self):
        self.stack = []
        self.redo_stack = []

    def push(self, action):
        self.stack.append(action)
        self.redo_stack.clear()

    def pop(self):
        if self.stack:
            a = self.stack.pop()
            self.redo_stack.append(a)
            return a
        return None

    def pop_redo(self):
        if self.redo_stack:
            a = self.redo_stack.pop()
            self.stack.append(a)
            return a
        return None

# ── Clipboard Path History ─────────────────────────────────────────────────────
class PathClipboard:
    def __init__(self):
        self.history = []

    def add(self, paths):
        entry = {"paths": paths, "time": datetime.now().strftime("%H:%M:%S")}
        self.history.insert(0, entry)
        if len(self.history) > 20:
            self.history.pop()

    def get_all(self):
        return self.history

# ── System Search Engine ───────────────────────────────────────────────────────
class SystemSearchEngine:
    @staticmethod
    def get_root_paths():
        roots = []
        if sys.platform == "win32":
            import string
            for letter in string.ascii_uppercase:
                drive = f"{letter}:\\"
                if os.path.exists(drive):
                    roots.append(drive)
        elif sys.platform == "darwin":
            roots = ["/"]
            vol = "/Volumes"
            if os.path.exists(vol):
                for d in os.listdir(vol):
                    fp = os.path.join(vol, d)
                    if os.path.ismount(fp):
                        roots.append(fp)
        else:
            roots = ["/"]
        return roots

    @staticmethod
    def search(query, roots, max_results=2000, show_hidden=False,
               file_type="all", min_size=None, max_size=None,
               modified_after=None, ext_filter="",
               stop_event=None, progress_cb=None, result_cb=None):
        results = []
        query_lower = query.lower()
        ext_f = ext_filter.strip().lstrip(".").lower()
        skip_dirs = {"proc", "sys", "dev", "run", "snap", "$RECYCLE.BIN", "System Volume Information"}
        for root in roots:
            try:
                for dirpath, dirs, files in os.walk(root, followlinks=False):
                    if stop_event and stop_event.is_set():
                        return results
                    dirs[:] = [d for d in dirs if d not in skip_dirs
                               and (show_hidden or not d.startswith("."))
                               and not d.startswith("$")]
                    if progress_cb:
                        progress_cb(f"Scanning: {dirpath[:70]}...")
                    all_entries = files[:]
                    if file_type in ("all", "folder"):
                        all_entries += dirs
                    for name in all_entries:
                        if not show_hidden and name.startswith("."): continue
                        if query_lower not in name.lower(): continue
                        full_path = os.path.join(dirpath, name)
                        is_dir = os.path.isdir(full_path)
                        if file_type == "file" and is_dir: continue
                        if file_type == "folder" and not is_dir: continue
                        if ext_f and not is_dir:
                            if not name.lower().endswith("." + ext_f): continue
                        if (min_size is not None or max_size is not None) and not is_dir:
                            try:
                                fsize = os.path.getsize(full_path)
                                if min_size is not None and fsize < min_size: continue
                                if max_size is not None and fsize > max_size: continue
                            except: continue
                        if modified_after is not None:
                            try:
                                mtime = os.path.getmtime(full_path)
                                if mtime < modified_after: continue
                            except: continue
                        results.append(full_path)
                        if result_cb:
                            result_cb(full_path)
                        if len(results) >= max_results:
                            return results
            except (PermissionError, OSError):
                continue
        return results

# ── NEW: Duplicate File Finder ────────────────────────────────────────────────
class DuplicateFinder:
    def __init__(self):
        self.files_by_hash = {}
        self.duplicates = []
        self.stop_event = threading.Event()
    
    def scan(self, root_path, show_hidden=False, min_size=0, compare_content=False):
        self.files_by_hash = {}
        self.duplicates = []
        self.stop_event.clear()
        
        for dirpath, dirs, files in os.walk(root_path, followlinks=False):
            if self.stop_event.is_set():
                break
            if not show_hidden:
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                files = [f for f in files if not f.startswith(".")]
            
            for fname in files:
                if self.stop_event.is_set():
                    break
                fpath = os.path.join(dirpath, fname)
                try:
                    st = os.stat(fpath)
                    if st.st_size < min_size:
                        continue
                    if compare_content:
                        # Full hash for accurate comparison
                        with open(fpath, "rb") as f:
                            file_hash = hashlib.md5(f.read()).hexdigest()
                    else:
                        # Quick check using size + first 4KB + last 4KB
                        with open(fpath, "rb") as f:
                            data = f.read(4096)
                            f.seek(max(0, st.st_size - 4096))
                            data += f.read(4096)
                            file_hash = hashlib.md5(data).hexdigest()
                        # Add size to hash key to reduce collisions
                        file_hash = f"{st.st_size}_{file_hash}"
                    
                    if file_hash not in self.files_by_hash:
                        self.files_by_hash[file_hash] = []
                    self.files_by_hash[file_hash].append(fpath)
                except (PermissionError, OSError):
                    continue
        
        # Filter to only duplicates
        self.duplicates = [files for files in self.files_by_hash.values() if len(files) > 1]
        return self.duplicates

# ── NEW: Disk Usage Analyzer ──────────────────────────────────────────────────
class DiskUsageAnalyzer:
    def __init__(self):
        self.usage_data = []
        self.total_size = 0
        self.stop_event = threading.Event()
    
    def scan(self, root_path, show_hidden=False, max_depth=10):
        self.usage_data = []
        self.total_size = 0
        self.stop_event.clear()
        
        def get_size(path):
            try:
                return os.path.getsize(path)
            except:
                return 0
        
        def scan_dir(path, depth=0):
            if self.stop_event.is_set():
                return
            if depth > max_depth:
                return
            
            try:
                entries = list(os.scandir(path))
            except (PermissionError, OSError):
                return
            
            total = 0
            sub_items = []
            
            for entry in entries:
                if self.stop_event.is_set():
                    break
                if not show_hidden and entry.name.startswith("."):
                    continue
                try:
                    if entry.is_file(follow_symlinks=False):
                        size = get_size(entry.path)
                        total += size
                        sub_items.append({
                            "name": entry.name,
                            "path": entry.path,
                            "size": size,
                            "is_dir": False,
                            "depth": depth + 1
                        })
                    elif entry.is_dir(follow_symlinks=False):
                        sub_total, sub_sub = scan_dir(entry.path, depth + 1)
                        total += sub_total
                        sub_items.append({
                            "name": entry.name,
                            "path": entry.path,
                            "size": sub_total,
                            "is_dir": True,
                            "depth": depth + 1,
                            "children": sub_sub
                        })
                except (PermissionError, OSError):
                    continue
            
            self.usage_data.append({
                "name": os.path.basename(path) or path,
                "path": path,
                "size": total,
                "is_dir": True,
                "depth": depth,
                "children": sub_items
            })
            self.total_size += total
            return total, sub_items
        
        scan_dir(root_path)
        return self.usage_data

# ── NEW: Full-Text Search Engine ──────────────────────────────────────────────
class FullTextSearch:
    def __init__(self):
        self.results = []
        self.stop_event = threading.Event()
        self.supported_exts = {".txt", ".py", ".js", ".html", ".css", ".json", ".xml", ".md", ".log",
                              ".csv", ".ini", ".cfg", ".conf", ".yaml", ".yml", ".toml", ".sql",
                              ".sh", ".bash", ".zsh", ".bat", ".ps1", ".cpp", ".c", ".h", ".java",
                              ".go", ".rs", ".swift", ".kt", ".php", ".rb", ".pl", ".lua", ".r"}
    
    def search(self, root_path, query, show_hidden=False, case_sensitive=False,
               whole_word=False, ext_filter="", max_results=500):
        self.results = []
        self.stop_event.clear()
        query_str = query if case_sensitive else query.lower()
        
        for dirpath, dirs, files in os.walk(root_path, followlinks=False):
            if self.stop_event.is_set():
                break
            if not show_hidden:
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                files = [f for f in files if not f.startswith(".")]
            
            for fname in files:
                if self.stop_event.is_set():
                    break
                fpath = os.path.join(dirpath, fname)
                ext = Path(fname).suffix.lower()
                
                # Filter by extension
                if ext_filter:
                    filters = [e.strip().lower() for e in ext_filter.split(",")]
                    if ext not in filters:
                        continue
                
                # Only search text files
                if ext not in self.supported_exts:
                    continue
                
                try:
                    # Check if file is readable and not too large
                    st = os.stat(fpath)
                    if st.st_size > 50 * 1024 * 1024:  # Skip files > 50MB
                        continue
                    
                    with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read()
                    
                    # Search in content
                    search_text = content if case_sensitive else content.lower()
                    if whole_word:
                        # Simple whole word search
                        import re
                        pattern = r'\b' + re.escape(query_str) + r'\b'
                        if re.search(pattern, search_text):
                            self.results.append(fpath)
                    else:
                        if query_str in search_text:
                            self.results.append(fpath)
                    
                    if len(self.results) >= max_results:
                        return self.results
                except (PermissionError, OSError, UnicodeDecodeError):
                    continue
        
        return self.results

# ══════════════════════════════════════════════════════════════════════════════
# ── INLINE RENAME WIDGET (modern) ─────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
class InlineRenameEntry:
    def __init__(self, tree, item, on_confirm, on_cancel):
        self.tree = tree
        self.item = item
        self.on_confirm = on_confirm
        self.on_cancel = on_cancel
        self.var = tk.StringVar()
        bbox = tree.bbox(item, column="#2")
        if not bbox:
            on_cancel()
            return
        x, y, w, h = bbox
        current_name = tree.item(item, "values")[1]
        self.var.set(current_name)
        entry_bg = P["INPUT_BG"]
        entry_fg = P["TEXT"]
        entry_sel = P["ACCENT"]
        self.entry = tk.Entry(tree, textvariable=self.var,
                              bg=entry_bg, fg=entry_fg,
                              insertbackground=entry_sel,
                              selectbackground=P["SEL"],
                              selectforeground=P["SEL_FG"],
                              relief="flat", bd=2,
                              font=("Segoe UI", scaled(9)),
                              highlightthickness=2,
                              highlightcolor=P["ACCENT"],
                              highlightbackground=P["BORDER"])
        self.entry.place(x=x-2, y=y-1, width=max(w+4, scaled(200)), height=h+2)
        self.entry.select_range(0, "end")
        self.entry.focus()
        self.entry.bind("<Return>", self._confirm)
        self.entry.bind("<Escape>", self._cancel)
        self.entry.bind("<FocusOut>", self._cancel)

    def _confirm(self, _=None):
        new_name = self.var.get().strip()
        self.entry.destroy()
        if new_name:
            self.on_confirm(new_name)
        else:
            self.on_cancel()

    def _cancel(self, _=None):
        try:
            self.entry.destroy()
        except:
            pass
        self.on_cancel()

# ══════════════════════════════════════════════════════════════════════════════
# ── MODERN BREADCRUMB BAR ────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
class ModernBreadcrumbBar(tk.Frame):
    def __init__(self, parent, navigate_cb, palette_getter, **kwargs):
        super().__init__(parent, **kwargs)
        self.navigate_cb = navigate_cb
        self.palette_getter = palette_getter
        self._buttons = []
        self.configure(bg=P["HEADER_BG"], height=scaled(32))
        
    def update(self, path):
        for w in self.winfo_children():
            w.destroy()
        self._buttons = []
        P = self.palette_getter()
        self.configure(bg=P["HEADER_BG"])
        
        # Home button
        home_btn = tk.Button(self, text="🏠", bg=P["HEADER_BG"], fg=P["ACCENT"],
                             activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                             relief="flat", bd=0, padx=4, pady=4,
                             font=("Segoe UI", scaled(10)), cursor="hand2",
                             command=lambda: self.navigate_cb(str(Path.home())))
        home_btn.pack(side="left", padx=(4, 2))
        self._buttons.append(home_btn)
        
        parts = Path(path).parts
        cumulative = ""
        for i, part in enumerate(parts):
            if sys.platform == "win32":
                cumulative = part if i == 0 else os.path.join(cumulative, part)
            else:
                cumulative = part if i == 0 else os.path.join(cumulative, part)
                if not cumulative.startswith("/"):
                    cumulative = "/" + cumulative
            
            display = part if len(part) < 15 else part[:12] + "…"
            is_last = i == len(parts) - 1
            
            # Separator
            sep = tk.Label(self, text="›", bg=P["HEADER_BG"], fg=P["SUBTEXT"],
                           font=("Segoe UI", scaled(10)))
            sep.pack(side="left", padx=2)
            
            btn = tk.Button(self, text=display,
                            bg=P["ACCENT"] if is_last else P["HEADER_BG"],
                            fg=P["TEXT"] if is_last else P["SUBTEXT"],
                            activebackground=P["HOVER"],
                            activeforeground=P["ACCENT"],
                            relief="flat", bd=0, padx=6, pady=4,
                            font=("Segoe UI", scaled(9), "bold" if is_last else "normal"),
                            cursor="hand2" if not is_last else "arrow",
                            command=lambda p=cumulative: self.navigate_cb(p) if not is_last else None)
            btn.pack(side="left")
            self._buttons.append(btn)

# ══════════════════════════════════════════════════════════════════════════════
# ── MODERN MAIN APPLICATION ──────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
class FileManager(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("NexFile Pro — Advanced File Manager")
        
        # Modern window sizing
        win_w = min(1440, max(980, int(SCREEN_W * 0.90)))
        win_h = min(900, max(650, int(SCREEN_H * 0.88)))
        x_off = (SCREEN_W - win_w) // 2
        y_off = (SCREEN_H - win_h) // 2
        self.geometry(f"{win_w}x{win_h}+{x_off}+{y_off}")
        self.minsize(scaled(850), scaled(560))
        
        # Load config
        cfg = load_config()
        self.theme = cfg.get("theme", "dark")
        self._apply_palette()
        self.configure(bg=P["BG"])
        
        self.ui_font_name = cfg.get("ui_font_name", "Segoe UI")
        self.ui_font_size = cfg.get("ui_font_size", scaled(9))
        self.export_dir = cfg.get("export_dir", DEFAULT_EXPORT_DIR)
        
        # State variables
        self.current_path = tk.StringVar(value=str(Path.home()))
        self.search_var = tk.StringVar()
        self.filter_ext_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Ready")
        self.history = [str(Path.home())]
        self.hist_idx = 0
        self.sort_col = "name"
        self.sort_rev = False
        self.show_hidden = tk.BooleanVar(value=cfg.get("show_hidden", False))
        self.view_mode = tk.StringVar(value=cfg.get("view_mode", "details"))
        self.clipboard = []
        self.selected_paths = []
        self.user_bookmarks = cfg.get("bookmarks", [])
        self.recent_files = cfg.get("recent_files", [])
        self.path_clipboard = PathClipboard()
        self.tag_manager = TagManager()
        self.color_mark_manager = ColorMarkManager()
        self.undo_stack = UndoStack()
        self._observer = None
        self._drag_data = {"paths": [], "x": 0, "y": 0}
        self._tab_dirs = [str(Path.home())]
        self._tab_idx = 0
        self._mouse_marking = False
        self._mouse_mark_color = "none"
        self._last_marked_item = None
        self._sys_search_stop_event = threading.Event()
        self._inline_rename_widget = None
        self._quick_type_filter_val = ""
        self._mark_toolbar_visible = False
        self._mini_player_visible = False
        
        self._build_styles()
        self._build_ui()
        self._bind_keys()
        self.load_dir(str(Path.home()))
        self._start_watcher()
        
        # Startup animation
        self._animate_startup()

    def _apply_palette(self):
        global P, MARK_COLORS
        P = dict(LIGHT_PALETTE if self.theme == "light" else DARK_PALETTE)
        MARK_COLORS = dict(MARK_COLORS_LIGHT if self.theme == "light" else MARK_COLORS_DARK)

    def _palette(self):
        return P

    def _f(self, size_delta=0, bold=False):
        size = self.ui_font_size + size_delta
        weight = "bold" if bold else "normal"
        return (self.ui_font_name, max(7, size), weight)

    def _animate_startup(self):
        """Subtle startup animation - fade in"""
        try:
            self.attributes("-alpha", 0.0)
            def fade_in(alpha=0.0):
                if alpha < 1.0:
                    alpha += 0.05
                    self.attributes("-alpha", min(alpha, 1.0))
                    self.after(20, lambda: fade_in(alpha))
            fade_in()
        except:
            pass

    # ── Theme ──────────────────────────────────────────────────────────────────
    def _toggle_theme(self):
        self.theme = "light" if self.theme == "dark" else "dark"
        self._apply_palette()
        self._save_prefs()
        self.rebuild_ui()

    def rebuild_ui(self):
        cur = self.current_path.get()
        hist, hidx = self.history[:], self.hist_idx
        scol, srev = self.sort_col, self.sort_rev
        shid = self.show_hidden.get()
        srch, fext = self.search_var.get(), self.filter_ext_var.get()
        bkms, recf = self.user_bookmarks[:], self.recent_files[:]
        vm = self.view_mode.get()
        tabs, tidx = self._tab_dirs[:], self._tab_idx
        self._stop_watcher()
        for w in self.winfo_children():
            w.destroy()
        self.configure(bg=P["BG"])
        self._build_styles()
        self._build_ui()
        self._bind_keys()
        self.current_path.set(cur)
        self.history = hist
        self.hist_idx = hidx
        self.sort_col = scol
        self.sort_rev = srev
        self.show_hidden.set(shid)
        self.search_var.set(srch)
        self.filter_ext_var.set(fext)
        self.user_bookmarks = bkms
        self.recent_files = recf
        self.view_mode.set(vm)
        self._tab_dirs = tabs
        self._tab_idx = tidx
        self._quick_type_filter_val = ""
        self._mark_toolbar_visible = False
        self.load_dir(cur)
        self._start_watcher()

    def _save_prefs(self):
        cfg = load_config()
        cfg["theme"] = self.theme
        cfg["show_hidden"] = self.show_hidden.get()
        cfg["view_mode"] = self.view_mode.get()
        cfg["bookmarks"] = self.user_bookmarks
        cfg["recent_files"] = self.recent_files[-50:]
        cfg["export_dir"] = self.export_dir
        cfg["ui_font_name"] = self.ui_font_name
        cfg["ui_font_size"] = self.ui_font_size
        save_config(cfg)

    # ── Watcher ────────────────────────────────────────────────────────────────
    def _start_watcher(self):
        if not WATCHDOG_AVAILABLE: return
        self._stop_watcher()
        try:
            handler = DirWatcher(lambda: self.after(200, self._auto_refresh))
            self._observer = Observer()
            self._observer.schedule(handler, self.current_path.get(), recursive=False)
            self._observer.start()
        except: pass

    def _stop_watcher(self):
        if self._observer:
            try:
                self._observer.stop()
                self._observer.join(timeout=1)
            except: pass
            self._observer = None

    def _auto_refresh(self):
        try:
            self._populate(self.current_path.get(), self.search_var.get())
        except: pass

    # ── Styles ─────────────────────────────────────────────────────────────────
    def _build_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TFrame", background=P["BG"])
        style.configure("Panel.TFrame", background=P["PANEL"])
        style.configure("Sidebar.TFrame", background=P["SIDEBAR"])
        style.configure("Header.TFrame", background=P["HEADER_BG"])
        style.configure("Card.TFrame", background=P["CARD"])
        
        # Modern Treeview
        style.configure("Treeview",
                        background=P["PANEL"],
                        foreground=P["TEXT"],
                        fieldbackground=P["PANEL"],
                        borderwidth=0,
                        rowheight=scaled(32),
                        font=self._f())
        style.configure("Treeview.Heading",
                        background=P["HEADER_BG"],
                        foreground=P["SUBTEXT"],
                        relief="flat", borderwidth=0,
                        font=self._f(bold=True))
        style.map("Treeview",
                  background=[("selected", P["SEL"]), ("!selected", P["PANEL"])],
                  foreground=[("selected", P["SEL_FG"]), ("!selected", P["TEXT"])])
        style.map("Treeview.Heading",
                  background=[("active", P["HOVER"])])
        
        # Modern scrollbars
        style.configure("Vertical.TScrollbar",
                        background=P["PANEL"],
                        troughcolor=P["BG"],
                        arrowcolor=P["SUBTEXT"],
                        borderwidth=0,
                        gripcount=0)
        style.configure("Horizontal.TScrollbar",
                        background=P["PANEL"],
                        troughcolor=P["BG"],
                        arrowcolor=P["SUBTEXT"],
                        borderwidth=0)
        
        # Modern progressbar
        style.configure("Modern.Horizontal.TProgressbar",
                        background=P["ACCENT"],
                        troughcolor=P["BORDER"],
                        borderwidth=0)
        style.configure("Success.Horizontal.TProgressbar",
                        background=P["SUCCESS"],
                        troughcolor=P["BORDER"],
                        borderwidth=0)
        
        # Modern combobox
        style.configure("Modern.TCombobox",
                        background=P["INPUT_BG"],
                        foreground=P["TEXT"],
                        fieldbackground=P["INPUT_BG"],
                        selectbackground=P["SEL"],
                        selectforeground=P["SEL_FG"],
                        bordercolor=P["BORDER"],
                        arrowcolor=P["SUBTEXT"])
        style.map("Modern.TCombobox",
                  fieldbackground=[("readonly", P["INPUT_BG"])],
                  selectbackground=[("readonly", P["INPUT_BG"])],
                  foreground=[("readonly", P["TEXT"])])

    # ══════════════════════════════════════════════════════════════════════════
    # ── MODERN UI LAYOUT ─────────────────────────────────────────────────────
    # ══════════════════════════════════════════════════════════════════════════
    def _build_ui(self):
        # ── Top Header Bar ────────────────────────────────────────────────────
        header = tk.Frame(self, bg=P["HEADER_BG"], height=scaled(56))
        header.pack(fill="x", side="top")
        header.pack_propagate(False)
        
        # App Logo
        logo_frame = tk.Frame(header, bg=P["HEADER_BG"])
        logo_frame.pack(side="left", padx=(12, 4), pady=8)
        
        # Gradient app icon
        icon_canvas = tk.Canvas(logo_frame, width=scaled(32), height=scaled(32),
                                bg=P["HEADER_BG"], highlightthickness=0)
        icon_canvas.pack(side="left")
        icon_canvas.create_rectangle(4, 4, 28, 28, fill=P["ACCENT"], outline="")
        icon_canvas.create_text(16, 16, text="N", fill="white", 
                                font=("Segoe UI", scaled(14), "bold"))
        
        tk.Label(logo_frame, text=" NexFile", bg=P["HEADER_BG"], fg=P["TEXT"],
                 font=(self.ui_font_name, scaled(14), "bold")).pack(side="left")
        tk.Label(logo_frame, text=" Pro", bg=P["HEADER_BG"], fg=P["ACCENT"],
                 font=(self.ui_font_name, scaled(12), "bold")).pack(side="left")
        
        # Divider
        tk.Frame(header, bg=P["BORDER"], width=1).pack(side="left", fill="y", padx=8, pady=10)
        
        # Navigation buttons
        nav = tk.Frame(header, bg=P["HEADER_BG"])
        nav.pack(side="left", padx=4)
        nav_btns = [
            ("◀", self.go_back, "Back (Alt+←)"),
            ("▶", self.go_forward, "Forward (Alt+→)"),
            ("↑", self.go_up, "Parent (Alt+↑)"),
            ("⟳", lambda: self.load_dir(self.current_path.get()), "Refresh (F5)"),
        ]
        for txt, cmd, tip in nav_btns:
            b = tk.Button(nav, text=txt, bg=P["HEADER_BG"], fg=P["SUBTEXT"],
                          activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                          relief="flat", bd=0, padx=scaled(10), pady=scaled(8),
                          font=("Segoe UI", scaled(13)), cursor="hand2", command=cmd)
            b.pack(side="left")
            Tooltip(b, tip)
        
        # Tabs
        self.tab_bar = tk.Frame(header, bg=P["HEADER_BG"])
        self.tab_bar.pack(side="left", padx=6, fill="y")
        self._rebuild_tabs()
        
        # Modern search bar
        self.search_bar = ModernSearchBar(
            header, placeholder="Search files...",
            command=self._live_filter_modern
        )
        self.search_bar.pack(side="left", fill="x", expand=True, padx=8, pady=10)
        
        # Right side buttons
        right_btns = tk.Frame(header, bg=P["HEADER_BG"])
        right_btns.pack(side="right", padx=8)
        
        # Theme toggle
        theme_icon = "🌙" if self.theme == "light" else "☀️"
        theme_btn = tk.Button(right_btns, text=theme_icon, bg=P["HEADER_BG"], fg=P["SUBTEXT"],
                              activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                              relief="flat", bd=0, padx=8, pady=8,
                              font=("Segoe UI", scaled(14)), cursor="hand2",
                              command=self._toggle_theme)
        theme_btn.pack(side="left", padx=2)
        Tooltip(theme_btn, "Toggle Theme")
        
        # System search
        sys_search_btn = tk.Button(right_btns, text="🌐", bg=P["HEADER_BG"], fg=P["ACCENT3"],
                                   activebackground=P["HOVER"], activeforeground=P["ACCENT3"],
                                   relief="flat", bd=0, padx=8, pady=8,
                                   font=("Segoe UI", scaled(14)), cursor="hand2",
                                   command=self.open_system_search)
        sys_search_btn.pack(side="left", padx=2)
        Tooltip(sys_search_btn, "PC-Wide Search")
        
        # Support
        support_btn = tk.Button(right_btns, text="💙", bg=P["HEADER_BG"], fg=P["ACCENT2"],
                                activebackground=P["HOVER"], activeforeground=P["ACCENT2"],
                                relief="flat", bd=0, padx=8, pady=8,
                                font=("Segoe UI", scaled(14)), cursor="hand2",
                                command=self.show_support_info)
        support_btn.pack(side="left", padx=2)
        Tooltip(support_btn, "Support & Donate")
        
        # ── Toolbar ──────────────────────────────────────────────────────────
        toolbar = tk.Frame(self, bg=P["PANEL"],
                           highlightthickness=1,
                           highlightbackground=P["BORDER"],
                           height=scaled(44))
        toolbar.pack(fill="x", side="top")
        toolbar.pack_propagate(False)
        
        # Left toolbar buttons
        left_toolbar = tk.Frame(toolbar, bg=P["PANEL"])
        left_toolbar.pack(side="left")
        
        tool_btns = [
            ("📁+", self.new_folder, "New Folder (Ctrl+N)"),
            ("📄+", self.new_file, "New File (Ctrl+Shift+N)"),
            ("📦+", self.open_bulk_create, "Bulk Create (Ctrl+Shift+B)"),
            ("✂", self.cut_files, "Cut (Ctrl+X)"),
            ("📋", self.copy_files, "Copy (Ctrl+C)"),
            ("📥", self.paste_files, "Paste (Ctrl+V)"),
            ("🗑", self.delete_files, "Delete (Del)"),
            ("✏", self.rename_file, "Rename (F2)"),
            ("↩", self.undo_action, "Undo (Ctrl+Z)"),
        ]
        for txt, cmd, tip in tool_btns:
            b = tk.Button(left_toolbar, text=txt, bg=P["PANEL"], fg=P["SUBTEXT"],
                          activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                          relief="flat", bd=0, padx=scaled(7), pady=scaled(8),
                          font=self._f(), cursor="hand2", command=cmd)
            b.pack(side="left")
            Tooltip(b, tip)
        
        tk.Frame(toolbar, bg=P["BORDER"], width=1).pack(side="left", fill="y", padx=6, pady=8)
        
        # Extension filter
        tk.Label(toolbar, text="Ext:", bg=P["PANEL"], fg=P["SUBTEXT"],
                 font=self._f()).pack(side="left", padx=(4, 2))
        ext_entry = tk.Entry(toolbar, textvariable=self.filter_ext_var,
                             width=8, bg=P["INPUT_BG"], fg=P["TEXT"],
                             relief="flat", font=self._f(),
                             insertbackground=P["ACCENT"],
                             selectbackground=P["SEL"],
                             selectforeground=P["SEL_FG"])
        ext_entry.pack(side="left")
        ext_entry.bind("<Return>", lambda e: self._apply_filter_ext())
        
        # Right toolbar buttons
        right_toolbar = tk.Frame(toolbar, bg=P["PANEL"])
        right_toolbar.pack(side="right")
        
        toolbar_right_btns = [
            ("📦", self.compress_selected, "Zip Selected"),
            ("📂", self.extract_archive, "Extract Archive"),
            ("🔗", self.create_symlink, "Create Symlink"),
            ("🏷️", self.manage_tags, "Manage Tags"),
            ("🎨", self.open_color_mark_toolbar, "Color Mark"),
            ("🔀", self.batch_rename, "Batch Rename"),
            ("🗂️", self.open_auto_organize, "Auto-Organize (Ctrl+Shift+O)"),
            ("💻", self.open_terminal_here, "Terminal Here (F4)"),
            ("🔖", self.add_bookmark, "Add Bookmark"),
            ("📊", self.show_props, "Properties"),
            ("📊", self.open_duplicate_finder, "Duplicate Finder"),
            ("💾", self.open_disk_usage, "Disk Usage"),
            ("📝", self.open_full_text_search, "Full-Text Search"),
        ]
        for txt, cmd, tip in toolbar_right_btns:
            b = tk.Button(right_toolbar, text=txt, bg=P["PANEL"], fg=P["SUBTEXT"],
                          activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                          relief="flat", bd=0, padx=scaled(6), pady=scaled(8),
                          font=self._f(), cursor="hand2", command=cmd)
            b.pack(side="left")
            Tooltip(b, tip)
        
        tk.Frame(toolbar, bg=P["BORDER"], width=1).pack(side="right", fill="y", padx=6, pady=8)
        
        # Export buttons
        tk.Button(right_toolbar, text="📊 Excel", bg=P["SUCCESS"], fg="white",
                  activebackground="#059669", relief="flat", bd=0, padx=8, pady=5,
                  font=self._f(bold=True), cursor="hand2",
                  command=self.open_export_dialog_excel).pack(side="left", padx=2)
        tk.Button(right_toolbar, text="📑 PDF", bg="#e11d48", fg="white",
                  activebackground="#be123c", relief="flat", bd=0, padx=8, pady=5,
                  font=self._f(bold=True), cursor="hand2",
                  command=self.open_export_dialog_pdf).pack(side="left", padx=2)
        
        tk.Frame(toolbar, bg=P["BORDER"], width=1).pack(side="right", fill="y", padx=6, pady=8)
        
        # Hidden toggle
        tk.Checkbutton(right_toolbar, text="👁 Hidden", variable=self.show_hidden,
                       bg=P["PANEL"], fg=P["SUBTEXT"], selectcolor=P["PANEL"],
                       activebackground=P["PANEL"], activeforeground=P["TEXT"],
                       font=self._f(), cursor="hand2",
                       command=lambda: self.load_dir(self.current_path.get())
                       ).pack(side="left", padx=6)
        
        # View mode toggle
        view_frame = tk.Frame(right_toolbar, bg=P["PANEL"])
        view_frame.pack(side="left", padx=4)
        for mode, icon, tip in [("details", "≡", "List view"), ("icons", "⊞", "Icon grid")]:
            b = tk.Button(view_frame, text=icon, bg=P["PANEL"], fg=P["SUBTEXT"],
                          activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                          relief="flat", padx=6, pady=4, font=("Segoe UI", scaled(13)),
                          cursor="hand2", command=lambda m=mode: self._set_view_mode(m))
            b.pack(side="left")
            Tooltip(b, tip)
        
        # ── Breadcrumb ─────────────────────────────────────────────────────
        self.breadcrumb = ModernBreadcrumbBar(
            self, navigate_cb=self.load_dir,
            palette_getter=self._palette
        )
        self.breadcrumb.pack(fill="x", side="top")
        
        # ── Color Mark Toolbar ─────────────────────────────────────────────
        self.mark_toolbar = tk.Frame(self, bg=P["HEADER_BG"], height=scaled(36))
        tk.Label(self.mark_toolbar, text="  🎨 Mark:", bg=P["HEADER_BG"],
                 fg=P["SUBTEXT"], font=self._f(bold=True)).pack(side="left", padx=4)
        self._mark_mode_active = tk.BooleanVar(value=False)
        tk.Checkbutton(self.mark_toolbar, text="Drag-mark",
                       variable=self._mark_mode_active,
                       bg=P["HEADER_BG"], fg=P["TEXT"], selectcolor=P["HEADER_BG"],
                       activebackground=P["HEADER_BG"],
                       font=self._f()).pack(side="left", padx=4)
        self._active_mark_color = tk.StringVar(value="none")
        for cname, cdata in MARK_COLORS.items():
            dot_color = cdata["fg"] if cname != "none" else P["SUBTEXT"]
            label_text = "✕" if cname == "none" else "●"
            b = tk.Button(self.mark_toolbar, text=f"{label_text} {cname[:3]}",
                          bg=P["HEADER_BG"], fg=dot_color,
                          activebackground=P["HOVER"], activeforeground=dot_color,
                          relief="flat", bd=0, padx=5, pady=4,
                          font=self._f(), cursor="hand2",
                          command=lambda c=cname: self._set_mark_color_and_apply(c))
            b.pack(side="left")
        tk.Button(self.mark_toolbar, text="Clear All",
                  bg=P["DANGER"], fg="white", relief="flat", bd=0,
                  padx=6, pady=4, font=self._f(), cursor="hand2",
                  command=self._clear_all_marks).pack(side="left", padx=4)
        tk.Button(self.mark_toolbar, text="✕ Close", bg=P["HEADER_BG"], fg=P["SUBTEXT"],
                  relief="flat", bd=0, font=self._f(),
                  command=self.open_color_mark_toolbar).pack(side="right", padx=8)
        
        # ── Sort / Filter Bar ──────────────────────────────────────────────
        sort_bar = tk.Frame(self, bg=P["PANEL"],
                            highlightthickness=1,
                            highlightbackground=P["BORDER"],
                            height=scaled(34))
        sort_bar.pack(fill="x", side="top")
        sort_bar.pack_propagate(False)
        
        tk.Label(sort_bar, text="  Sort:", bg=P["PANEL"], fg=P["SUBTEXT"],
                 font=self._f()).pack(side="left", padx=(8, 4))
        for col, label in [("name", "Name"), ("size", "Size"), ("modified", "Date"), ("type", "Type")]:
            is_active = self.sort_col == col
            b = tk.Button(sort_bar, text=label,
                          bg=P["ACCENT"] if is_active else P["PANEL"],
                          fg="white" if is_active else P["SUBTEXT"],
                          relief="flat", padx=10, pady=4, font=self._f(),
                          cursor="hand2",
                          command=lambda c=col: self._sort_by(c))
            b.pack(side="left", padx=2)
        
        tk.Frame(sort_bar, bg=P["BORDER"], width=1).pack(side="left", fill="y", padx=8, pady=6)
        tk.Label(sort_bar, text="Filter:", bg=P["PANEL"], fg=P["SUBTEXT"],
                 font=self._f()).pack(side="left", padx=4)
        for label, val in [("All", ""), ("📄 Files", "__files__"), ("📁 Folders", "__dirs__")]:
            b = tk.Button(sort_bar, text=label, bg=P["PANEL"], fg=P["SUBTEXT"],
                          activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                          relief="flat", padx=8, pady=4, font=self._f(),
                          cursor="hand2",
                          command=lambda v=val: self._quick_type_filter(v))
            b.pack(side="left", padx=2)
        
        # ── Main Content ────────────────────────────────────────────────────
        main = tk.Frame(self, bg=P["BG"])
        main.pack(fill="both", expand=True)
        
        # Sidebar
        sidebar = tk.Frame(main, bg=P["SIDEBAR"], width=scaled(210))
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        
        # Sidebar tabs
        sb_tabs = tk.Frame(sidebar, bg=P["SIDEBAR"])
        sb_tabs.pack(fill="x")
        self._sb_frames = {}
        for key, label in [("quick", "⚡ Quick"), ("recent", "📋 Recent"), ("tags", "🏷️ Tags")]:
            b = tk.Button(sb_tabs, text=label, bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                          activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                          relief="flat", bd=0, padx=8, pady=8, font=self._f(),
                          cursor="hand2", command=lambda k=key: self._show_sidebar(k))
            b.pack(side="left", expand=True, fill="x")
        
        self._sb_content = tk.Frame(sidebar, bg=P["SIDEBAR"])
        self._sb_content.pack(fill="both", expand=True)
        self._build_sidebar_quick()
        self._build_sidebar_recent()
        self._build_sidebar_tags()
        self._show_sidebar("quick")
        
        # Content area
        content = tk.Frame(main, bg=P["BG"])
        content.pack(side="left", fill="both", expand=True)
        
        self.list_panel = tk.Frame(content, bg=P["PANEL"])
        self.list_panel.pack(side="left", fill="both", expand=True)
        self._build_tree_view()
        
        # Preview panel
        self.preview = tk.Frame(content, bg=P["PANEL"], width=scaled(320))
        self.preview.pack(side="right", fill="y")
        self.preview.pack_propagate(False)
        
        preview_header = tk.Frame(self.preview, bg=P["HEADER_BG"], height=scaled(34))
        preview_header.pack(fill="x")
        preview_header.pack_propagate(False)
        tk.Label(preview_header, text="  👁 Preview", bg=P["HEADER_BG"], fg=P["SUBTEXT"],
                 font=self._f(bold=True)).pack(side="left", pady=8)
        tk.Button(preview_header, text="✕", bg=P["HEADER_BG"], fg=P["SUBTEXT"],
                  activebackground=P["HOVER"], activeforeground=P["DANGER"],
                  relief="flat", bd=0, font=self._f(),
                  command=self._toggle_preview).pack(side="right", padx=8)
        
        self.preview_image_label = tk.Label(self.preview, bg=P["PANEL"])
        self.preview_text = tk.Text(self.preview,
                                    bg=P["INPUT_BG"], fg=P["TEXT"],
                                    font=self._f(), relief="flat", wrap="none",
                                    state="disabled", insertbackground=P["ACCENT"],
                                    bd=8, selectbackground=P["SEL"],
                                    selectforeground=P["SEL_FG"])
        self.preview_text.pack(fill="both", expand=True, padx=4, pady=4)
        
        # ── Status Bar ──────────────────────────────────────────────────────
        statusbar = tk.Frame(self, bg=P["HEADER_BG"], height=scaled(28))
        statusbar.pack(fill="x", side="bottom")
        statusbar.pack_propagate(False)
        
        tk.Label(statusbar, textvariable=self.status_var,
                 bg=P["HEADER_BG"], fg=P["SUBTEXT"], font=self._f(), anchor="w"
                 ).pack(side="left", padx=12, fill="y")
        
        # Status indicators
        status_right = tk.Frame(statusbar, bg=P["HEADER_BG"])
        status_right.pack(side="right", padx=8)
        
        self.sel_label = tk.Label(status_right, text="",
                                  bg=P["HEADER_BG"], fg=P["ACCENT"], font=self._f())
        self.sel_label.pack(side="left", padx=8)
        
        self.watcher_label = tk.Label(status_right,
                                      text="● Live" if WATCHDOG_AVAILABLE else "",
                                      bg=P["HEADER_BG"], 
                                      fg=P["SUCCESS"] if WATCHDOG_AVAILABLE else P["SUBTEXT"],
                                      font=self._f())
        self.watcher_label.pack(side="left", padx=4)
        
        # Version info
        tk.Label(status_right, text="v3.0", bg=P["HEADER_BG"], fg=P["SUBTEXT"],
                 font=self._f()).pack(side="left", padx=8)

    # ── Sidebar ────────────────────────────────────────────────────────────────
    def _show_sidebar(self, key):
        for k, f in self._sb_frames.items():
            f.pack_forget()
        if key in self._sb_frames:
            self._sb_frames[key].pack(fill="both", expand=True)

    def _build_sidebar_quick(self):
        f = tk.Frame(self._sb_content, bg=P["SIDEBAR"])
        self._sb_frames["quick"] = f
        
        tk.Label(f, text="QUICK ACCESS", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                 font=self._f(bold=True), pady=8).pack(anchor="w", padx=12)
        for name, path in self._default_bookmarks():
            self._add_sidebar_btn(f, name, path)
        
        tk.Label(f, text="BOOKMARKS", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                 font=self._f(bold=True), pady=6).pack(anchor="w", padx=12)
        if self.user_bookmarks:
            for name, path in self.user_bookmarks:
                bf = tk.Frame(f, bg=P["SIDEBAR"])
                bf.pack(fill="x", padx=6, pady=1)
                btn = tk.Button(bf, text=f"  {name}", bg=P["SIDEBAR"], fg=P["TEXT"],
                                activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                                relief="flat", bd=0, anchor="w", padx=4, pady=4,
                                font=self._f(), cursor="hand2",
                                command=lambda p=path: self.load_dir(p))
                btn.pack(side="left", fill="x", expand=True)
                tk.Button(bf, text="✕", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                          activebackground=P["HOVER"], activeforeground=P["DANGER"],
                          relief="flat", bd=0, font=self._f(),
                          command=lambda n=name, p=path: self._remove_bookmark(n, p)).pack(side="right")
        else:
            tk.Label(f, text="No bookmarks yet", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                     font=self._f(), pady=4).pack(anchor="w", padx=16)
        
        tk.Label(f, text="DRIVES", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                 font=self._f(bold=True), pady=6).pack(anchor="w", padx=12)
        for drive in self._get_drives():
            self._add_sidebar_btn(f, drive, drive)

    def _build_sidebar_recent(self):
        f = tk.Frame(self._sb_content, bg=P["SIDEBAR"])
        self._sb_frames["recent"] = f
        
        tk.Label(f, text="RECENT FILES", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                 font=self._f(bold=True), pady=8).pack(anchor="w", padx=12)
        
        if not self.recent_files:
            tk.Label(f, text="No recent files", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                     font=self._f(), pady=4).pack(anchor="w", padx=16)
            return
        
        canvas = tk.Canvas(f, bg=P["SIDEBAR"], highlightthickness=0)
        vsb = ttk.Scrollbar(f, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        inner = tk.Frame(canvas, bg=P["SIDEBAR"])
        canvas.create_window((0, 0), window=inner, anchor="nw")
        
        for path in reversed(self.recent_files[-30:]):
            name = os.path.basename(path)
            btn = tk.Button(inner, text=f"  {file_icon(path)} {name[:20]}",
                            bg=P["SIDEBAR"], fg=P["TEXT"],
                            activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                            relief="flat", bd=0, anchor="w", padx=4, pady=3,
                            font=self._f(), cursor="hand2",
                            command=lambda p=path: self._open_recent(p))
            btn.pack(fill="x", padx=6, pady=1)
            Tooltip(btn, path)
        inner.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

    def _build_sidebar_tags(self):
        f = tk.Frame(self._sb_content, bg=P["SIDEBAR"])
        self._sb_frames["tags"] = f
        
        tk.Label(f, text="TAG LABELS", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                 font=self._f(bold=True), pady=8).pack(anchor="w", padx=12)
        
        if not self.tag_manager.colors:
            tk.Label(f, text="No tags defined", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                     font=self._f(), pady=4).pack(anchor="w", padx=16)
        
        for tag_name, color in self.tag_manager.colors.items():
            bf = tk.Frame(f, bg=P["SIDEBAR"])
            bf.pack(fill="x", padx=8, pady=2)
            tk.Label(bf, text="●", bg=P["SIDEBAR"], fg=color,
                     font=("Segoe UI", scaled(12))).pack(side="left")
            tk.Button(bf, text=tag_name, bg=P["SIDEBAR"], fg=P["TEXT"],
                      activebackground=P["HOVER"], activeforeground=color,
                      relief="flat", bd=0, anchor="w", padx=4, pady=3,
                      font=self._f(), cursor="hand2",
                      command=lambda t=tag_name: self._filter_by_tag(t)
                      ).pack(side="left", fill="x", expand=True)
        
        tk.Label(f, text="COLOR MARKS", bg=P["SIDEBAR"], fg=P["SUBTEXT"],
                 font=self._f(bold=True), pady=8).pack(anchor="w", padx=12)
        for cname, cdata in MARK_COLORS.items():
            if cname == "none": continue
            bf = tk.Frame(f, bg=P["SIDEBAR"])
            bf.pack(fill="x", padx=8, pady=2)
            tk.Label(bf, text="●", bg=P["SIDEBAR"], fg=cdata["fg"],
                     font=("Segoe UI", scaled(12))).pack(side="left")
            tk.Button(bf, text=cname, bg=P["SIDEBAR"], fg=P["TEXT"],
                      activebackground=P["HOVER"], activeforeground=cdata["fg"],
                      relief="flat", bd=0, anchor="w", padx=4, pady=3,
                      font=self._f(), cursor="hand2",
                      command=lambda c=cname: self._filter_by_color_mark(c)
                      ).pack(side="left", fill="x", expand=True)

    def _filter_by_color_mark(self, color_name):
        if not hasattr(self, "tree") or not self.tree: return
        marked = [p for p, c in self.color_mark_manager.marks.items() if c == color_name]
        self.tree.delete(*self.tree.get_children())
        for i, path in enumerate(marked):
            if not os.path.exists(path): continue
            try:
                st = os.stat(path)
                is_dir = os.path.isdir(path)
                size = "" if is_dir else human_size(st.st_size)
                mtime = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
                ftype = "Folder" if is_dir else (Path(path).suffix.upper().lstrip(".") + " File")
                perms = stat.filemode(st.st_mode)
            except:
                size = mtime = ftype = perms = ""
            icon = file_icon(path)
            tags_list = self.tag_manager.get_tags(path)
            tag_str = " ".join(f"[{t}]" for t in tags_list)
            self.tree.insert("", "end",
                             values=(icon, os.path.basename(path), size, ftype, mtime, perms, tag_str),
                             tags=[f"mark_{color_name}"], iid=path)
        self.status_var.set(f"Color mark filter: {color_name} — {len(marked)} item(s)")

    def _open_recent(self, path):
        if os.path.exists(path):
            if os.path.isdir(path): self.load_dir(path)
            else:
                self.load_dir(os.path.dirname(path))
                self._open_file(path)

    def _filter_by_tag(self, tag):
        if not hasattr(self, "tree") or not self.tree: return
        tagged = [p for p, tags in self.tag_manager.tags.items() if tag in tags]
        self.tree.delete(*self.tree.get_children())
        for i, path in enumerate(tagged):
            if not os.path.exists(path): continue
            try:
                st = os.stat(path)
                is_dir = os.path.isdir(path)
                size = "" if is_dir else human_size(st.st_size)
                mtime = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
                ftype = "Folder" if is_dir else (Path(path).suffix.upper().lstrip(".") + " File")
                perms = stat.filemode(st.st_mode)
            except:
                size = mtime = ftype = perms = ""
            icon = file_icon(path)
            tags_list = self.tag_manager.get_tags(path)
            tag_str = " ".join(f"[{t}]" for t in tags_list)
            row_tag = ["even", "tagged"] if i % 2 == 0 else ["odd", "tagged"]
            self.tree.insert("", "end",
                             values=(icon, os.path.basename(path), size, ftype, mtime, perms, tag_str),
                             tags=row_tag, iid=path)
        self.status_var.set(f"Tag filter: #{tag} — {len(tagged)} item(s)")

    def _default_bookmarks(self):
        h = str(Path.home())
        items = [
            ("🏠 Home", h), ("🖥️ Desktop", os.path.join(h, "Desktop")),
            ("📥 Downloads", os.path.join(h, "Downloads")),
            ("📄 Documents", os.path.join(h, "Documents")),
            ("🎵 Music", os.path.join(h, "Music")),
            ("🖼️ Pictures", os.path.join(h, "Pictures")),
            ("🎬 Videos", os.path.join(h, "Videos")),
        ]
        return [(n, p) for n, p in items if os.path.exists(p)]

    def _get_drives(self):
        if sys.platform == "win32":
            import string
            return [f"{d}:\\" for d in string.ascii_uppercase if os.path.exists(f"{d}:\\")]
        elif sys.platform == "darwin":
            vol = "/Volumes"
            return [os.path.join(vol, d) for d in os.listdir(vol)] if os.path.exists(vol) else ["/"]
        else:
            drives = ["/"]
            mnt = "/mnt"
            if os.path.exists(mnt):
                drives += [os.path.join(mnt, d) for d in os.listdir(mnt)]
            return drives

    def _add_sidebar_btn(self, parent, name, path):
        btn = tk.Button(parent, text=f"  {name}", bg=P["SIDEBAR"], fg=P["TEXT"],
                        activebackground=P["HOVER"], activeforeground=P["ACCENT"],
                        relief="flat", bd=0, anchor="w", padx=6, pady=4, font=self._f(),
                        cursor="hand2", command=lambda p=path: self.load_dir(p))
        btn.pack(fill="x", padx=6, pady=1)
        return btn

    def add_bookmark(self):
        current = self.current_path.get()
        name = os.path.basename(current) or current
        for n, p in self.user_bookmarks:
            if p == current:
                messagebox.showinfo("Bookmark", "Already bookmarked!")
                return
        self.user_bookmarks.append((f"🔖 {name}", current))
        self._save_prefs()
        self.rebuild_ui()

    def _remove_bookmark(self, name, path):
        self.user_bookmarks = [(n, p) for n, p in self.user_bookmarks if p != path]
        self._save_prefs()
        self.rebuild_ui()

    # ── Key bindings ──────────────────────────────────────────────────────────
    def _bind_keys(self):
        self.bind("<F5>", lambda e: self.load_dir(self.current_path.get()))
        self.bind("<Alt-Left>", lambda e: self.go_back())
        self.bind("<Alt-Right>", lambda e: self.go_forward())
        self.bind("<Alt-Up>", lambda e: self.go_up())
        self.bind("<BackSpace>", lambda e: self.go_up())
        self.bind("<Delete>", lambda e: self.delete_files())
        self.bind("<Control-c>", lambda e: self.copy_files())
        self.bind("<Control-x>", lambda e: self.cut_files())
        self.bind("<Control-v>", lambda e: self.paste_files())
        self.bind("<Control-z>", lambda e: self.undo_action())
        self.bind("<Control-y>", lambda e: self.redo_action())
        self.bind("<F2>", lambda e: self._start_inline_rename())
        self.bind("<Control-f>", lambda e: self.search_bar.entry.focus())
        self.bind("<Escape>", lambda e: self._clear_search())
        self.bind("<Control-t>", lambda e: self._new_tab())
        self.bind("<Control-w>", lambda e: self._close_tab(self._tab_idx))
        self.bind("<Control-a>", lambda e: self._select_all())
        self.bind("<F4>", lambda e: self.open_terminal_here())
        self.bind("<Control-Shift-F>", lambda e: self.open_system_search())
        self.bind("<Control-Shift-V>", lambda e: self.paste_to_multiple_dirs())
        self.bind("<Control-e>", lambda e: self.open_export_dialog_excel())
        self.bind("<Control-p>", lambda e: self.open_export_dialog_pdf())
        self.bind("<Control-n>", lambda e: self.new_folder())
        self.bind("<Control-Shift-N>", lambda e: self.new_file())
        self.bind("<Control-Shift-B>", lambda e: self.open_bulk_create())
        self.bind("<Control-Shift-O>", lambda e: self.open_auto_organize())
        self.bind("<Control-r>", lambda e: self.rename_file())
        self.bind("<Control-b>", lambda e: self.add_bookmark())
        self.bind("<F1>", lambda e: self.show_support_info())
        self.bind("<F3>", lambda e: self.search_bar.entry.focus())
        self.bind("<F6>", lambda e: self.open_font_settings())
        self.bind("<Control-h>", lambda e: [
            self.show_hidden.set(not self.show_hidden.get()),
            self.load_dir(self.current_path.get())
        ])
        self.bind("<Control-Home>", lambda e: self.load_dir(str(Path.home())))
        self.bind("<Control-Tab>", lambda e: self._switch_tab(
            (self._tab_idx + 1) % len(self._tab_dirs)))
        self.bind("<Alt-Return>", lambda e: self.show_props())
        self.bind("<Control-m>", lambda e: self.compress_selected())
        self.bind("<Control-Shift-F>", lambda e: self.open_system_search())

    # ── Mouse handlers ─────────────────────────────────────────────────────────
    def _on_mouse_press(self, event):
        if not self.tree: return
        item = self.tree.identify_row(event.y)
        self._drag_data["x"] = event.x
        self._drag_data["y"] = event.y
        self._drag_data["paths"] = []
        self._last_marked_item = None
        if self._mark_mode_active.get() and item:
            color = self._active_mark_color.get()
            self.color_mark_manager.set_mark(item, color)
            self._last_marked_item = item
            self._apply_mark_tag(item, color)
        else:
            if item:
                sel = list(self.tree.selection())
                if item not in sel: sel = [item]
                self._drag_data["paths"] = sel

    def _on_mouse_drag(self, event):
        if not self.tree: return
        item = self.tree.identify_row(event.y)
        if self._mark_mode_active.get():
            if item and item != self._last_marked_item:
                color = self._active_mark_color.get()
                self.color_mark_manager.set_mark(item, color)
                self._apply_mark_tag(item, color)
                self._last_marked_item = item
        else:
            if self._drag_data["paths"]:
                dx = abs(event.x - self._drag_data["x"])
                dy = abs(event.y - self._drag_data["y"])
                if dx > 5 or dy > 5:
                    self.config(cursor="fleur")

    def _on_mouse_release(self, event):
        self.config(cursor="")
        if not self.tree: return
        if self._mark_mode_active.get():
            self._last_marked_item = None
            return
        if not self._drag_data["paths"]: return
        target_item = self.tree.identify_row(event.y)
        dx = abs(event.x - self._drag_data["x"])
        dy = abs(event.y - self._drag_data["y"])
        if (dx > 8 or dy > 8) and target_item and os.path.isdir(target_item):
            paths = self._drag_data["paths"]
            if target_item not in paths:
                answer = messagebox.askyesnocancel(
                    "Drag & Drop",
                    f"Move {len(paths)} item(s) to\n{os.path.basename(target_item)}?\n\n(No = Copy, Cancel = abort)")
                if answer is None:
                    self._drag_data["paths"] = []
                    return
                op = "move" if answer else "copy"
                errors = []
                for src in paths:
                    name = os.path.basename(src)
                    dst = os.path.join(target_item, name)
                    try:
                        if op == "move": shutil.move(src, dst)
                        else:
                            if os.path.isdir(src): shutil.copytree(src, dst)
                            else: shutil.copy2(src, dst)
                    except Exception as e:
                        errors.append(str(e))
                if errors:
                    messagebox.showerror("Drag & Drop Error", "\n".join(errors))
                self.load_dir(self.current_path.get())
        self._drag_data["paths"] = []

    def _apply_mark_tag(self, item, color_name):
        if not self.tree: return
        try:
            existing_tags = [t for t in list(self.tree.item(item, "tags"))
                             if not t.startswith("mark_")]
            if color_name != "none":
                existing_tags.append(f"mark_{color_name}")
            self.tree.item(item, tags=existing_tags)
        except: pass

    # ── Tab Management ─────────────────────────────────────────────────────────
    def _rebuild_tabs(self):
        for w in self.tab_bar.winfo_children():
            w.destroy()
        for i, path in enumerate(self._tab_dirs):
            name = os.path.basename(path) or path
            color = P["ACCENT"] if i == self._tab_idx else P["SUBTEXT"]
            bg = P["HOVER"] if i == self._tab_idx else P["HEADER_BG"]
            f = tk.Frame(self.tab_bar, bg=bg)
            f.pack(side="left")
            btn = tk.Button(f, text=f"  {name[:12]}  ", bg=bg, fg=color,
                            relief="flat", bd=0, padx=4, pady=8, font=self._f(),
                            cursor="hand2", command=lambda idx=i: self._switch_tab(idx))
            btn.pack(side="left")
            if len(self._tab_dirs) > 1:
                tk.Button(f, text="✕", bg=bg, fg=P["SUBTEXT"],
                          activebackground=P["HOVER"], activeforeground=P["DANGER"],
                          relief="flat", bd=0, font=self._f(), cursor="hand2",
                          command=lambda idx=i: self._close_tab(idx)).pack(side="left")
        tk.Button(self.tab_bar, text="+", bg=P["HEADER_BG"], fg=P["SUCCESS"],
                  activebackground=P["HOVER"],
                  relief="flat", bd=0, padx=6, pady=8, font=(self.ui_font_name, scaled(12)),
                  cursor="hand2", command=self._new_tab).pack(side="left")

    def _new_tab(self):
        self._tab_dirs.append(str(Path.home()))
        self._tab_idx = len(self._tab_dirs) - 1
        self._rebuild_tabs()
        self.load_dir(self._tab_dirs[self._tab_idx])

    def _switch_tab(self, idx):
        self._tab_dirs[self._tab_idx] = self.current_path.get()
        self._tab_idx = idx
        self._rebuild_tabs()
        self.load_dir(self._tab_dirs[idx])

    def _close_tab(self, idx):
        if len(self._tab_dirs) <= 1: return
        self._tab_dirs.pop(idx)
        self._tab_idx = max(0, min(idx, len(self._tab_dirs) - 1))
        self._rebuild_tabs()
        self.load_dir(self._tab_dirs[self._tab_idx])

    # ── Navigation ─────────────────────────────────────────────────────────────
    def go_back(self):
        if self.hist_idx > 0:
            self.hist_idx -= 1
            self._load_no_history(self.history[self.hist_idx])

    def go_forward(self):
        if self.hist_idx < len(self.history) - 1:
            self.hist_idx += 1
            self._load_no_history(self.history[self.hist_idx])

    def go_up(self):
        self.load_dir(str(Path(self.current_path.get()).parent))

    def load_dir(self, path):
        if not os.path.isdir(path): return
        path = str(Path(path).resolve())
        self.history = self.history[:self.hist_idx + 1]
        self.history.append(path)
        self.hist_idx = len(self.history) - 1
        self._load_no_history(path)
        self._tab_dirs[self._tab_idx] = path
        try: self._rebuild_tabs()
        except: pass
        self._restart_watcher(path)
        try:
            self.breadcrumb.update(path)
        except: pass

    def _restart_watcher(self, path):
        if WATCHDOG_AVAILABLE:
            self._stop_watcher()
            try:
                handler = DirWatcher(lambda: self.after(300, self._auto_refresh))
                self._observer = Observer()
                self._observer.schedule(handler, path, recursive=False)
                self._observer.start()
            except: pass

    def _load_no_history(self, path):
        self.current_path.set(path)
        self.search_bar.set("")
        self._populate(path)

    def _populate(self, path, filter_str=""):
        try:
            entries = list(os.scandir(path))
        except PermissionError:
            self.status_var.set("Permission denied.")
            return
        except Exception as e:
            self.status_var.set(f"Error: {e}")
            return

        dirs, files = [], []
        ext_filter = self.filter_ext_var.get().strip().lower()
        qtf = getattr(self, "_quick_type_filter_val", "")
        for e in entries:
            if not self.show_hidden.get() and e.name.startswith("."): continue
            if filter_str and filter_str.lower() not in e.name.lower(): continue
            if ext_filter and not e.is_dir():
                ext = ext_filter.lstrip(".")
                if not e.name.lower().endswith("." + ext) and not e.name.lower().endswith(ext): continue
            if qtf == "__files__" and e.is_dir(follow_symlinks=False): continue
            if qtf == "__dirs__" and not e.is_dir(follow_symlinks=False): continue
            (dirs if e.is_dir(follow_symlinks=False) else files).append(e)

        def sort_key(e):
            if self.sort_col == "name": return e.name.lower()
            if self.sort_col == "size":
                try: return e.stat().st_size
                except: return 0
            if self.sort_col == "modified":
                try: return e.stat().st_mtime
                except: return 0
            if self.sort_col == "type": return Path(e.name).suffix.lower()
            return e.name.lower()

        dirs.sort(key=sort_key, reverse=self.sort_rev)
        files.sort(key=sort_key, reverse=self.sort_rev)
        all_entries = dirs + files

        if self.view_mode.get() == "icons":
            self._build_icons_view(all_entries)
            self.status_var.set(f"  {len(dirs)} folder(s)  •  {len(files)} file(s)   ({path})")
            return

        if not hasattr(self, "tree") or not self.tree:
            self._build_tree_view()

        self.tree.delete(*self.tree.get_children())
        row = 0
        for e in all_entries:
            try:
                st = e.stat()
                size = "" if e.is_dir() else human_size(st.st_size)
                mtime = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d  %H:%M")
                perms = stat.filemode(st.st_mode)
                ftype = "Folder" if e.is_dir() else (
                    Path(e.name).suffix.upper().lstrip(".") + " File"
                    if Path(e.name).suffix else "File")
            except:
                size = mtime = perms = ftype = ""

            tags_list = self.tag_manager.get_tags(e.path)
            tag_str = " ".join(f"[{t}]" for t in tags_list) if tags_list else ""
            icon = file_icon(e.path)
            row_tags = [("even", "odd")[row % 2]]
            if os.path.islink(e.path): row_tags.append("link")
            elif e.is_dir(): row_tags.append("dir")
            elif os.access(e.path, os.X_OK): row_tags.append("exec")
            if tags_list: row_tags.append("tagged")
            mark = self.color_mark_manager.get_mark(e.path)
            if mark != "none": row_tags.append(f"mark_{mark}")
            self.tree.insert("", "end",
                             values=(icon, e.name, size, ftype, mtime, perms, tag_str),
                             tags=row_tags, iid=e.path)
            row += 1

        self.status_var.set(
            f"  {len(dirs)} folder{'s' if len(dirs) != 1 else ''}"
            f"  •  {len(files)} file{'s' if len(files) != 1 else ''}   ({path})")

    def _apply_filter_ext(self):
        self._populate(self.current_path.get(), self.search_bar.get())

    def _sort_by(self, col):
        if self.sort_col == col: self.sort_rev = not self.sort_rev
        else:
            self.sort_col = col
            self.sort_rev = False
        self._populate(self.current_path.get(), self.search_bar.get())

    def _select_all(self):
        if self.tree:
            self.tree.selection_set(self.tree.get_children())
            self.sel_label.config(text=f"{len(self.tree.get_children())} items selected")

    # ── Selection ──────────────────────────────────────────────────────────────
    def _on_select(self, _=None):
        if not self.tree: return
        sel = self.tree.selection()
        self.selected_paths = list(sel)
        count = len(sel)
        if count == 1:
            self._show_preview(sel[0])
            self.sel_label.config(text="1 item selected")
        elif count > 1:
            total_size = 0
            for p in sel:
                try:
                    if os.path.isfile(p):
                        total_size += os.path.getsize(p)
                except: pass
            size_info = f" ({human_size(total_size)})" if total_size > 0 else ""
            self.sel_label.config(text=f"{count} items selected{size_info}")
            self._clear_preview()
        else:
            self.sel_label.config(text="")
            self._clear_preview()

    def _on_double_click(self, _=None):
        if not self.tree: return
        sel = self.tree.selection()
        if not sel: return
        path = sel[0]
        if os.path.isdir(path): self.load_dir(path)
        else: self._open_file(path)

    def _open_file(self, path):
        if path not in self.recent_files:
            self.recent_files.append(path)
            self.recent_files = self.recent_files[-50:]
            self._save_prefs()
        try:
            if sys.platform == "win32": os.startfile(path)
            elif sys.platform == "darwin": subprocess.Popen(["open", path])
            else: subprocess.Popen(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Open Error", str(e))

    # ── Preview ────────────────────────────────────────────────────────────────
    def _show_preview(self, path):
        self.preview_image_label.pack_forget()
        self.preview_text.pack_forget()
        if PIL_AVAILABLE and is_image(path):
            try:
                img = Image.open(path)
                img.thumbnail((280, 220))
                photo = ImageTk.PhotoImage(img)
                self.preview_image_label.config(image=photo, bg=P["PANEL"])
                self.preview_image_label.image = photo
                self.preview_image_label.pack(fill="x", padx=4, pady=4)
            except: pass
        self.preview_text.pack(fill="both", expand=True, padx=4, pady=4)
        self.preview_text.config(state="normal",
                                  bg=P["INPUT_BG"], fg=P["TEXT"],
                                  selectbackground=P["SEL"],
                                  selectforeground=P["SEL_FG"])
        self.preview_text.delete("1.0", "end")
        if os.path.isdir(path):
            try:
                items = os.listdir(path)[:30]
                content = f"📁 {os.path.basename(path)}\n{'─'*28}\n" + "\n".join(items)
                if len(os.listdir(path)) > 30:
                    content += f"\n... +{len(os.listdir(path))-30} more"
            except: content = "(Permission denied)"
        elif can_preview(path):
            try:
                with open(path, "r", encoding="utf-8", errors="replace") as f:
                    content = f.read(6000)
                if len(content) == 6000: content += "\n\n[Truncated...]"
            except: content = "(Cannot read file)"
        else:
            try:
                st = os.stat(path)
                content = (f"📄 {os.path.basename(path)}\n{'─'*28}\n"
                           f"Size:     {human_size(st.st_size)}\n"
                           f"Modified: {datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M')}\n"
                           f"Perms:    {stat.filemode(st.st_mode)}\n\n[Binary / Unsupported]")
            except: content = "(Cannot stat file)"
        tags = self.tag_manager.get_tags(path)
        if tags: content += f"\n\n🏷️ Tags: {', '.join(tags)}"
        mark = self.color_mark_manager.get_mark(path)
        if mark != "none": content += f"\n🎨 Color mark: {mark}"
        self.preview_text.insert("1.0", content)
        self.preview_text.config(state="disabled")

    def _clear_preview(self):
        self.preview_image_label.pack_forget()
        self.preview_text.config(state="normal")
        self.preview_text.delete("1.0", "end")
        self.preview_text.config(state="disabled")

    def _toggle_preview(self):
        if self.preview.winfo_width() > 10: self.preview.pack_forget()
        else: self.preview.pack(side="right", fill="y")

    # ── Search ─────────────────────────────────────────────────────────────────
    def _live_filter_modern(self, text):
        self._populate(self.current_path.get(), text)

    def _do_search(self):
        q = self.search_bar.get().strip()
        if not q:
            self._populate(self.current_path.get())
            return
        root = self.current_path.get()
        self.status_var.set(f"Searching '{q}'...")
        if self.tree:
            self.tree.delete(*self.tree.get_children())

        def run():
            results = []
            try:
                for dirpath, dirs, files in os.walk(root):
                    if not self.show_hidden.get():
                        dirs[:] = [d for d in dirs if not d.startswith(".")]
                    for name in dirs + files:
                        if q.lower() in name.lower():
                            results.append(os.path.join(dirpath, name))
                        if len(results) >= 500: break
                    if len(results) >= 500: break
            except: pass
            self.after(0, lambda: self._show_search_results(results, q))

        threading.Thread(target=run, daemon=True).start()

    def _show_search_results(self, results, q):
        if not self.tree:
            self._build_tree_view()
        self.tree.delete(*self.tree.get_children())
        for i, path in enumerate(results):
            try:
                st = os.stat(path)
                is_dir = os.path.isdir(path)
                size = "" if is_dir else human_size(st.st_size)
                mtime = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d  %H:%M")
                ftype = "Folder" if is_dir else (Path(path).suffix.upper().lstrip(".") + " File")
                perms = stat.filemode(st.st_mode)
            except: size = mtime = ftype = perms = ""
            icon = file_icon(path)
            tags_list = self.tag_manager.get_tags(path)
            tag_str = " ".join(f"[{t}]" for t in tags_list)
            row_tags = [("even", "odd")[i % 2]]
            if os.path.isdir(path): row_tags.append("dir")
            mark = self.color_mark_manager.get_mark(path)
            if mark != "none": row_tags.append(f"mark_{mark}")
            self.tree.insert("", "end",
                             values=(icon, os.path.relpath(path, self.current_path.get()),
                                     size, ftype, mtime, perms, tag_str),
                             tags=row_tags, iid=path)
        self.status_var.set(f"Search '{q}': {len(results)} result(s)"
                            + (" (max 500)" if len(results) == 500 else ""))

    def _clear_search(self):
        self.search_bar.clear()
        self._populate(self.current_path.get())

    # ── Color Mark Toolbar ─────────────────────────────────────────────────────
    def open_color_mark_toolbar(self):
        if self._mark_toolbar_visible:
            self.mark_toolbar.pack_forget()
            self._mark_toolbar_visible = False
        else:
            self.mark_toolbar.pack(fill="x", side="top", after=self.breadcrumb)
            self._mark_toolbar_visible = True

    def _set_mark_color_and_apply(self, color_name):
        self._active_mark_color.set(color_name)
        sel = self._get_selection()
        if sel:
            for path in sel:
                self.color_mark_manager.set_mark(path, color_name)
                self._apply_mark_tag(path, color_name)
            self.status_var.set(f"Applied '{color_name}' mark to {len(sel)} item(s)")
        else:
            self.status_var.set(f"Active mark color: {color_name}")

    def _clear_all_marks(self):
        if not messagebox.askyesno("Clear Marks", "Clear all color marks in the current folder?"):
            return
        if self.tree:
            for item in self.tree.get_children():
                self.color_mark_manager.set_mark(item, "none")
        self._populate(self.current_path.get(), self.search_bar.get())

    # ── PC-Wide Search ─────────────────────────────────────────────────────────
    def open_system_search(self):
        win = tk.Toplevel(self)
        win.title("🌐 PC-Wide System Search")
        win.geometry(f"{scaled(980)}x{scaled(700)}")
        win.configure(bg=P["BG"])
        win.grab_set()
        win.resizable(True, True)

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(50))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  🌐  SEARCH ENTIRE PC", bg=P["HEADER_BG"], fg=P["ACCENT"],
                 font=self._f(4, bold=True)).pack(side="left", padx=12, pady=12)

        search_frame = tk.Frame(win, bg=P["INPUT_BG"],
                                highlightthickness=1, highlightbackground=P["BORDER"])
        search_frame.pack(fill="x", padx=16, pady=(12, 0))
        sys_search_var = tk.StringVar()
        entry = tk.Entry(search_frame, textvariable=sys_search_var,
                         bg=P["INPUT_BG"], fg=P["TEXT"],
                         insertbackground=P["ACCENT"],
                         font=self._f(4), relief="flat", bd=8,
                         selectbackground=P["SEL"], selectforeground=P["SEL_FG"])
        entry.pack(side="left", fill="x", expand=True)
        entry.focus()

        filter_frame = tk.Frame(win, bg=P["BG"])
        filter_frame.pack(fill="x", padx=16, pady=6)
        tk.Label(filter_frame, text="Type:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left", padx=(0, 4))
        type_var = tk.StringVar(value="all")
        for t in ["all", "file", "folder"]:
            tk.Radiobutton(filter_frame, text=t.capitalize(), variable=type_var, value=t,
                           bg=P["BG"], fg=P["TEXT"], selectcolor=P["CARD"],
                           activebackground=P["BG"], font=self._f()).pack(side="left", padx=4)
        tk.Label(filter_frame, text="  Ext:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left", padx=(8, 4))
        sys_ext_var = tk.StringVar()
        tk.Entry(filter_frame, textvariable=sys_ext_var, width=8,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=self._f(), insertbackground=P["ACCENT"]).pack(side="left")
        tk.Label(filter_frame, text="  Max MB:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left", padx=(8, 4))
        max_size_var = tk.StringVar()
        tk.Entry(filter_frame, textvariable=max_size_var, width=6,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=self._f(), insertbackground=P["ACCENT"]).pack(side="left")
        tk.Label(filter_frame, text="  Days:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left", padx=(8, 4))
        days_var = tk.StringVar()
        tk.Entry(filter_frame, textvariable=days_var, width=5,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=self._f(), insertbackground=P["ACCENT"]).pack(side="left")
        show_hidden_sys = tk.BooleanVar(value=False)
        tk.Checkbutton(filter_frame, text="Hidden", variable=show_hidden_sys,
                       bg=P["BG"], fg=P["SUBTEXT"], selectcolor=P["BG"],
                       activebackground=P["BG"], font=self._f()).pack(side="left", padx=8)

        scope_frame = tk.Frame(win, bg=P["BG"])
        scope_frame.pack(fill="x", padx=16, pady=2)
        tk.Label(scope_frame, text="Search in:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left", padx=(0, 6))
        scope_var = tk.StringVar(value="home")
        for s, label in [("home", "Home"), ("current", "Current Folder"), ("pc", "Entire PC")]:
            tk.Radiobutton(scope_frame, text=label, variable=scope_var, value=s,
                           bg=P["BG"], fg=P["TEXT"], selectcolor=P["CARD"],
                           activebackground=P["BG"], font=self._f()).pack(side="left", padx=6)

        btn_frame = tk.Frame(win, bg=P["BG"])
        btn_frame.pack(fill="x", padx=16, pady=6)
        self._sys_search_stop_event = threading.Event()
        search_btn_var = tk.StringVar(value="🔍 Search")
        status_sys = tk.StringVar(value="Enter a search term and press Search")
        tk.Label(btn_frame, textvariable=status_sys, bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left")
        stop_btn = tk.Button(btn_frame, text="⏹ Stop", bg=P["DANGER"], fg="white",
                             relief="flat", padx=12, pady=6, font=self._f(),
                             cursor="hand2", state="disabled")
        stop_btn.pack(side="right", padx=4)
        search_btn = tk.Button(btn_frame, textvariable=search_btn_var,
                               bg=P["ACCENT"], fg="white", relief="flat", padx=16, pady=6,
                               font=self._f(bold=True), cursor="hand2")
        search_btn.pack(side="right", padx=4)

        prog_frame = tk.Frame(win, bg=P["BG"])
        prog_frame.pack(fill="x", padx=16, pady=2)
        progress = ttk.Progressbar(prog_frame, mode="indeterminate")
        progress.pack(fill="x")

        result_frame = tk.Frame(win, bg=P["PANEL"])
        result_frame.pack(fill="both", expand=True, padx=16, pady=4)
        res_cols = ("icon", "name", "path", "size", "type", "modified")
        res_tree = ttk.Treeview(result_frame, columns=res_cols, show="headings", selectmode="extended")
        for col, label, w, anchor in [
            ("icon", "", scaled(30), "center"), ("name", "Name", scaled(220), "w"),
            ("path", "Location", scaled(340), "w"), ("size", "Size", scaled(80), "e"),
            ("type", "Type", scaled(90), "w"), ("modified", "Modified", scaled(140), "w"),
        ]:
            res_tree.heading(col, text=label)
            res_tree.column(col, width=w, anchor=anchor, stretch=(col == "path"))
        res_tree.tag_configure("even", background=P["EVEN_ROW"], foreground=P["TEXT"])
        res_tree.tag_configure("odd", background=P["ODD_ROW"], foreground=P["TEXT"])
        res_tree.tag_configure("dir", foreground=P["ACCENT"])
        res_vsb = ttk.Scrollbar(result_frame, orient="vertical", command=res_tree.yview)
        res_tree.configure(yscrollcommand=res_vsb.set)
        res_tree.pack(side="left", fill="both", expand=True)
        res_vsb.pack(side="right", fill="y")
        result_count = [0]

        def add_result(path):
            try:
                st = os.stat(path)
                is_dir = os.path.isdir(path)
                size = "" if is_dir else human_size(st.st_size)
                mtime = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
                ftype = "Folder" if is_dir else (Path(path).suffix.upper().lstrip(".") + " File" if Path(path).suffix else "File")
                icon = file_icon(path)
                row_tags = [("even", "odd")[result_count[0] % 2]]
                if is_dir: row_tags.append("dir")
                res_tree.insert("", "end",
                                values=(icon, os.path.basename(path), os.path.dirname(path), size, ftype, mtime),
                                tags=row_tags, iid=path)
                result_count[0] += 1
            except: pass

        def do_search():
            q = sys_search_var.get().strip()
            if not q:
                messagebox.showwarning("Empty Query", "Enter a search term.", parent=win)
                return
            self._sys_search_stop_event.set()
            time.sleep(0.1)
            self._sys_search_stop_event.clear()
            res_tree.delete(*res_tree.get_children())
            result_count[0] = 0
            scope = scope_var.get()
            if scope == "home": roots = [str(Path.home())]
            elif scope == "current": roots = [self.current_path.get()]
            else: roots = SystemSearchEngine.get_root_paths()
            max_mb = None
            try: max_mb = int(float(max_size_var.get()) * 1024 * 1024)
            except: pass
            mod_after = None
            try: mod_after = time.time() - int(days_var.get()) * 86400
            except: pass
            search_btn.config(state="disabled")
            stop_btn.config(state="normal")
            progress.start(10)
            search_btn_var.set("🔍 Searching...")
            status_sys.set(f"Searching for '{q}'...")

            def run():
                SystemSearchEngine.search(
                    query=q, roots=roots, max_results=3000,
                    show_hidden=show_hidden_sys.get(), file_type=type_var.get(),
                    max_size=max_mb, modified_after=mod_after, ext_filter=sys_ext_var.get(),
                    stop_event=self._sys_search_stop_event,
                    progress_cb=lambda msg: win.after(0, lambda m=msg: status_sys.set(m[:80])),
                    result_cb=lambda p: win.after(0, lambda path=p: add_result(path)),
                )
                win.after(0, search_done)

            def search_done():
                progress.stop()
                search_btn.config(state="normal")
                stop_btn.config(state="disabled")
                search_btn_var.set("🔍 Search")
                n = result_count[0]
                status_sys.set(f"✅ Found {n} result(s)" + (" (limit 3000)" if n >= 3000 else ""))

            threading.Thread(target=run, daemon=True).start()

        def stop_search():
            self._sys_search_stop_event.set()
            progress.stop()
            search_btn.config(state="normal")
            stop_btn.config(state="disabled")
            search_btn_var.set("🔍 Search")
            status_sys.set("Search stopped.")

        search_btn.config(command=do_search)
        stop_btn.config(command=stop_search)
        entry.bind("<Return>", lambda e: do_search())

        act_frame = tk.Frame(win, bg=P["PANEL"])
        act_frame.pack(fill="x", padx=16, pady=(0, 8))

        def navigate_to():
            sel = res_tree.selection()
            if not sel: return
            path = sel[0]
            target = path if os.path.isdir(path) else os.path.dirname(path)
            win.destroy()
            self.load_dir(target)

        def open_selected_sys():
            for path in res_tree.selection()[:5]: self._open_file(path)

        def delete_files_sys():
            sel = res_tree.selection()
            if not sel: return
            names = "\n".join(os.path.basename(p) for p in sel[:6])
            if not messagebox.askyesno("Delete", f"Delete {len(sel)} item(s)?\n\n{names}", parent=win):
                return
            errors = []
            for p in sel:
                try:
                    if TRASH_AVAILABLE: send2trash.send2trash(p)
                    else:
                        if os.path.isdir(p): shutil.rmtree(p)
                        else: os.remove(p)
                    res_tree.delete(p)
                    result_count[0] -= 1
                except Exception as e:
                    errors.append(str(e))
            if errors: messagebox.showerror("Delete Errors", "\n".join(errors), parent=win)

        count_label = tk.Label(act_frame, text="", bg=P["PANEL"], fg=P["SUBTEXT"], font=self._f())
        count_label.pack(side="right", padx=8)

        for txt, cmd, color in [
            ("📂 Go To", navigate_to, P["ACCENT"]),
            ("🚀 Open", open_selected_sys, P["ACCENT2"]),
            ("🗑 Delete", delete_files_sys, P["DANGER"]),
        ]:
            tk.Button(act_frame, text=txt, bg=color, fg="white",
                      relief="flat", padx=8, pady=6, font=self._f(),
                      cursor="hand2", command=cmd).pack(side="left", padx=2, pady=6)

        def on_res_select(_=None):
            n = len(res_tree.selection())
            count_label.config(text=f"{n} selected" if n else "")
        res_tree.bind("<<TreeviewSelect>>", on_res_select)
        res_tree.bind("<Double-1>", lambda e: navigate_to())
        win.protocol("WM_DELETE_WINDOW",
                     lambda: [self._sys_search_stop_event.set(), win.destroy()])

    # ── NEW: Duplicate File Finder ─────────────────────────────────────────────
    def open_duplicate_finder(self):
        win = tk.Toplevel(self)
        win.title("🔍 Duplicate File Finder")
        win.geometry(f"{scaled(900)}x{scaled(700)}")
        win.configure(bg=P["BG"])
        win.grab_set()
        win.resizable(True, True)

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(50))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  🔍  DUPLICATE FILE FINDER", bg=P["HEADER_BG"], fg=P["ACCENT"],
                 font=self._f(4, bold=True)).pack(side="left", padx=12, pady=12)

        # Settings
        settings = tk.Frame(win, bg=P["BG"])
        settings.pack(fill="x", padx=16, pady=8)

        # Source folder
        src_frame = tk.Frame(settings, bg=P["BG"])
        src_frame.pack(fill="x", pady=4)
        tk.Label(src_frame, text="Scan Folder:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left")
        src_var = tk.StringVar(value=self.current_path.get())
        tk.Entry(src_frame, textvariable=src_var, bg=P["INPUT_BG"], fg=P["TEXT"],
                 relief="flat", font=self._f(), width=50,
                 insertbackground=P["ACCENT"]).pack(side="left", padx=8, ipady=4)
        tk.Button(src_frame, text="Browse", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=10, pady=4, font=self._f(),
                  cursor="hand2",
                  command=lambda: src_var.set(
                      filedialog.askdirectory(title="Select Folder to Scan for Duplicates", parent=win)
                      or src_var.get())).pack(side="left")

        # Options
        opts = tk.Frame(settings, bg=P["BG"])
        opts.pack(fill="x", pady=4)
        tk.Label(opts, text="Options:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left")
        min_size_var = tk.IntVar(value=1024)
        tk.Label(opts, text="Min Size (bytes):", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left", padx=(12, 4))
        tk.Entry(opts, textvariable=min_size_var, width=8,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=self._f(), insertbackground=P["ACCENT"]).pack(side="left", padx=4)
        
        compare_content_var = tk.BooleanVar(value=False)
        tk.Checkbutton(opts, text="Full content comparison (slower but accurate)",
                       variable=compare_content_var, bg=P["BG"], fg=P["SUBTEXT"],
                       selectcolor=P["BG"], font=self._f()).pack(side="left", padx=12)

        # Results area
        result_frame = tk.Frame(win, bg=P["PANEL"])
        result_frame.pack(fill="both", expand=True, padx=16, pady=4)
        
        # Treeview for duplicates
        cols = ("group", "name", "path", "size", "modified")
        dupe_tree = ttk.Treeview(result_frame, columns=cols, show="headings", selectmode="extended")
        for col, label, w, anchor in [
            ("group", "Group", scaled(50), "center"),
            ("name", "Name", scaled(200), "w"),
            ("path", "Path", scaled(350), "w"),
            ("size", "Size", scaled(80), "e"),
            ("modified", "Modified", scaled(140), "w"),
        ]:
            dupe_tree.heading(col, text=label)
            dupe_tree.column(col, width=w, anchor=anchor, stretch=(col == "path"))
        dupe_tree.tag_configure("even", background=P["EVEN_ROW"], foreground=P["TEXT"])
        dupe_tree.tag_configure("odd", background=P["ODD_ROW"], foreground=P["TEXT"])
        dupe_tree.tag_configure("selected", background=P["SEL"], foreground=P["SEL_FG"])
        dupe_vsb = ttk.Scrollbar(result_frame, orient="vertical", command=dupe_tree.yview)
        dupe_tree.configure(yscrollcommand=dupe_vsb.set)
        dupe_tree.pack(side="left", fill="both", expand=True)
        dupe_vsb.pack(side="right", fill="y")

        status_var = tk.StringVar(value="Configure options and click 'Find Duplicates'")
        tk.Label(win, textvariable=status_var, bg=P["BG"], fg=P["SUBTEXT"],
                 font=self._f(), anchor="w").pack(fill="x", padx=16, pady=4)
        progress = ttk.Progressbar(win, mode="indeterminate")
        progress.pack(fill="x", padx=16, pady=4)

        # Action buttons
        action_frame = tk.Frame(win, bg=P["BG"])
        action_frame.pack(fill="x", padx=16, pady=8)

        finder = DuplicateFinder()
        duplicates_data = []

        def find_duplicates():
            root = src_var.get()
            if not os.path.isdir(root):
                messagebox.showerror("Invalid Folder", "Select a valid folder.", parent=win)
                return
            
            dupe_tree.delete(*dupe_tree.get_children())
            status_var.set("Scanning for duplicates...")
            progress.start(10)
            
            def run():
                try:
                    dups = finder.scan(
                        root_path=root,
                        show_hidden=self.show_hidden.get(),
                        min_size=min_size_var.get(),
                        compare_content=compare_content_var.get()
                    )
                    win.after(0, lambda: show_results(dups))
                except Exception as e:
                    win.after(0, lambda: status_var.set(f"Error: {e}"))

            def show_results(dups):
                progress.stop()
                if not dups:
                    status_var.set("✅ No duplicates found!")
                    return
                
                duplicates_data.clear()
                group_num = 1
                total_files = 0
                total_wasted = 0
                
                for group in dups:
                    if len(group) < 2:
                        continue
                    duplicates_data.append(group)
                    total_files += len(group)
                    try:
                        total_wasted += sum(os.path.getsize(p) for p in group[1:])
                    except:
                        pass
                    
                    for idx, path in enumerate(group):
                        try:
                            st = os.stat(path)
                            size = human_size(st.st_size)
                            mtime = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
                            name = os.path.basename(path)
                            display_path = os.path.dirname(path)
                            if len(display_path) > 60:
                                display_path = "..." + display_path[-57:]
                        except:
                            size = "N/A"
                            mtime = "N/A"
                            name = os.path.basename(path)
                            display_path = os.path.dirname(path)
                        
                        tags = ("even", "odd")[idx % 2]
                        if idx == 0:
                            tags = (tags, "first")
                        dupe_tree.insert("", "end",
                                       values=(str(group_num), name, display_path, size, mtime),
                                       tags=tags, iid=path)
                    group_num += 1
                
                status_var.set(f"✅ Found {len(dups)} duplicate group(s), {total_files} file(s), "
                             f"~{human_size(total_wasted)} wasted space")
                dupe_tree.tag_configure("first", background=P["ACCENT2"], foreground="white")

            threading.Thread(target=run, daemon=True).start()

        def delete_selected_duplicates():
            sel = dupe_tree.selection()
            if not sel:
                messagebox.showwarning("No Selection", "Select duplicate files to delete.", parent=win)
                return
            if not messagebox.askyesno("Delete Duplicates", 
                                      f"Delete {len(sel)} selected duplicate file(s)?\n"
                                      "This will permanently delete the files.", parent=win):
                return
            
            errors = []
            for path in sel:
                try:
                    if TRASH_AVAILABLE:
                        send2trash.send2trash(path)
                    else:
                        os.remove(path)
                    dupe_tree.delete(path)
                except Exception as e:
                    errors.append(str(e))
            if errors:
                messagebox.showerror("Delete Errors", "\n".join(errors), parent=win)
            status_var.set(f"Deleted {len(sel) - len(errors)} duplicate file(s)")

        def keep_one_per_group():
            # Keep only the first file in each group, delete the rest
            to_delete = []
            for group in duplicates_data:
                if len(group) > 1:
                    to_delete.extend(group[1:])
            
            if not to_delete:
                messagebox.showinfo("No Duplicates", "No duplicates to clean up.", parent=win)
                return
            
            if not messagebox.askyesno("Keep One Per Group",
                                      f"Keep only 1 file from each duplicate group?\n"
                                      f"{len(to_delete)} file(s) will be deleted.", parent=win):
                return
            
            errors = []
            for path in to_delete:
                try:
                    if TRASH_AVAILABLE:
                        send2trash.send2trash(path)
                    else:
                        os.remove(path)
                    dupe_tree.delete(path)
                except Exception as e:
                    errors.append(str(e))
            if errors:
                messagebox.showerror("Delete Errors", "\n".join(errors), parent=win)
            status_var.set(f"Cleaned up {len(to_delete) - len(errors)} duplicate file(s)")

        def select_group(group_idx):
            # Select all files in a group
            items = []
            for item in dupe_tree.get_children():
                if dupe_tree.item(item, "values")[0] == str(group_idx):
                    items.append(item)
            if items:
                dupe_tree.selection_set(items)
                status_var.set(f"Selected {len(items)} files from group {group_idx}")

        # Buttons
        tk.Button(action_frame, text="🔍 Find Duplicates", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=16, pady=8, font=self._f(bold=True),
                  cursor="hand2", command=find_duplicates).pack(side="left", padx=4)
        tk.Button(action_frame, text="🗑 Delete Selected", bg=P["DANGER"], fg="white",
                  relief="flat", padx=12, pady=8, font=self._f(),
                  cursor="hand2", command=delete_selected_duplicates).pack(side="left", padx=4)
        tk.Button(action_frame, text="✨ Keep One Per Group", bg=P["SUCCESS"], fg="white",
                  relief="flat", padx=12, pady=8, font=self._f(),
                  cursor="hand2", command=keep_one_per_group).pack(side="left", padx=4)
        tk.Button(action_frame, text="Close", bg=P["CARD"], fg=P["TEXT"],
                  relief="flat", padx=10, pady=8, font=self._f(),
                  cursor="hand2", command=win.destroy).pack(side="right", padx=4)

        # Double click to open file location
        def on_dupe_double_click(event):
            sel = dupe_tree.selection()
            if sel:
                path = sel[0]
                if os.path.exists(path):
                    self.load_dir(os.path.dirname(path))
                    win.destroy()
        dupe_tree.bind("<Double-1>", on_dupe_double_click)

    # ── NEW: Disk Usage Analyzer ──────────────────────────────────────────────
    def open_disk_usage(self):
        win = tk.Toplevel(self)
        win.title("💾 Disk Usage Analyzer")
        win.geometry(f"{scaled(900)}x{scaled(700)}")
        win.configure(bg=P["BG"])
        win.grab_set()
        win.resizable(True, True)

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(50))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  💾  DISK USAGE ANALYZER", bg=P["HEADER_BG"], fg=P["ACCENT3"],
                 font=self._f(4, bold=True)).pack(side="left", padx=12, pady=12)

        # Settings
        settings = tk.Frame(win, bg=P["BG"])
        settings.pack(fill="x", padx=16, pady=8)

        src_frame = tk.Frame(settings, bg=P["BG"])
        src_frame.pack(fill="x", pady=4)
        tk.Label(src_frame, text="Analyze Folder:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left")
        src_var = tk.StringVar(value=self.current_path.get())
        tk.Entry(src_frame, textvariable=src_var, bg=P["INPUT_BG"], fg=P["TEXT"],
                 relief="flat", font=self._f(), width=50,
                 insertbackground=P["ACCENT"]).pack(side="left", padx=8, ipady=4)
        tk.Button(src_frame, text="Browse", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=10, pady=4, font=self._f(),
                  cursor="hand2",
                  command=lambda: src_var.set(
                      filedialog.askdirectory(title="Select Folder to Analyze", parent=win)
                      or src_var.get())).pack(side="left")

        # Options
        opts = tk.Frame(settings, bg=P["BG"])
        opts.pack(fill="x", pady=4)
        depth_var = tk.IntVar(value=5)
        tk.Label(opts, text="Max Depth:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left")
        tk.Spinbox(opts, from_=1, to=20, textvariable=depth_var, width=5,
                   bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                   font=self._f()).pack(side="left", padx=4)

        # Results
        result_frame = tk.Frame(win, bg=P["PANEL"])
        result_frame.pack(fill="both", expand=True, padx=16, pady=4)

        # Treeview for disk usage
        cols = ("name", "size", "human_size", "percentage", "depth")
        usage_tree = ttk.Treeview(result_frame, columns=cols, show="headings", selectmode="browse")
        for col, label, w, anchor in [
            ("name", "Name", scaled(300), "w"),
            ("size", "Size (bytes)", scaled(120), "e"),
            ("human_size", "Human Size", scaled(100), "e"),
            ("percentage", "% of Total", scaled(80), "e"),
            ("depth", "Depth", scaled(60), "center"),
        ]:
            usage_tree.heading(col, text=label)
            usage_tree.column(col, width=w, anchor=anchor, stretch=(col == "name"))
        usage_tree.tag_configure("dir", foreground=P["ACCENT"])
        usage_tree.tag_configure("file", foreground=P["TEXT"])
        usage_tree.tag_configure("large", background=P["DANGER"], foreground="white")
        usage_tree.tag_configure("medium", background=P["WARNING"], foreground="black")
        usage_vsb = ttk.Scrollbar(result_frame, orient="vertical", command=usage_tree.yview)
        usage_tree.configure(yscrollcommand=usage_vsb.set)
        usage_tree.pack(side="left", fill="both", expand=True)
        usage_vsb.pack(side="right", fill="y")

        status_var = tk.StringVar(value="Select a folder and click 'Analyze'")
        tk.Label(win, textvariable=status_var, bg=P["BG"], fg=P["SUBTEXT"],
                 font=self._f(), anchor="w").pack(fill="x", padx=16, pady=4)
        progress = ttk.Progressbar(win, mode="indeterminate")
        progress.pack(fill="x", padx=16, pady=4)

        # Summary
        summary_frame = tk.Frame(win, bg=P["HEADER_BG"])
        summary_frame.pack(fill="x", padx=16, pady=4)
        summary_labels = []
        for i in range(4):
            lbl = tk.Label(summary_frame, text="", bg=P["HEADER_BG"], fg=P["SUBTEXT"],
                          font=self._f(), padx=10)
            lbl.pack(side="left", expand=True)
            summary_labels.append(lbl)

        def analyze_disk():
            root = src_var.get()
            if not os.path.isdir(root):
                messagebox.showerror("Invalid Folder", "Select a valid folder.", parent=win)
                return
            
            usage_tree.delete(*usage_tree.get_children())
            status_var.set("Analyzing disk usage...")
            progress.start(10)
            
            analyzer = DiskUsageAnalyzer()
            
            def run():
                try:
                    data = analyzer.scan(
                        root_path=root,
                        show_hidden=self.show_hidden.get(),
                        max_depth=depth_var.get()
                    )
                    win.after(0, lambda: show_results(data, analyzer.total_size))
                except Exception as e:
                    win.after(0, lambda: status_var.set(f"Error: {e}"))

            def show_results(data, total_size):
                progress.stop()
                if not data:
                    status_var.set("No data found.")
                    return
                
                # Summary
                summary_labels[0].config(text=f"📁 Total: {human_size(total_size)}", fg=P["TEXT"])
                summary_labels[1].config(text=f"📂 Items: {len(usage_tree.get_children())}", fg=P["SUBTEXT"])
                
                def add_items(items, parent="", depth=0):
                    for item in items:
                        if not item.get("is_dir", True):
                            continue
                        name = item.get("name", "")
                        size = item.get("size", 0)
                        pct = (size / total_size * 100) if total_size > 0 else 0
                        
                        tag = "dir"
                        if pct > 20:
                            tag = "large"
                        elif pct > 10:
                            tag = "medium"
                        
                        node = usage_tree.insert(parent, "end",
                                               values=(name, size, human_size(size), f"{pct:.1f}%", depth),
                                               tags=(tag,))
                        
                        if "children" in item:
                            add_items(item["children"], node, depth + 1)
                
                add_items(data)
                usage_tree.tag_configure("dir", foreground=P["ACCENT"])
                usage_tree.tag_configure("large", background=P["DANGER"], foreground="white")
                usage_tree.tag_configure("medium", background=P["WARNING"], foreground="black")
                status_var.set(f"✅ Analysis complete! Total: {human_size(total_size)}")

            threading.Thread(target=run, daemon=True).start()

        def open_selected_folder():
            sel = usage_tree.selection()
            if sel:
                path = sel[0]
                if os.path.exists(path):
                    self.load_dir(path)
                    win.destroy()

        # Buttons
        action_frame = tk.Frame(win, bg=P["BG"])
        action_frame.pack(fill="x", padx=16, pady=8)

        tk.Button(action_frame, text="💾 Analyze", bg=P["ACCENT3"], fg="white",
                  relief="flat", padx=16, pady=8, font=self._f(bold=True),
                  cursor="hand2", command=analyze_disk).pack(side="left", padx=4)
        tk.Button(action_frame, text="📂 Open Folder", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=12, pady=8, font=self._f(),
                  cursor="hand2", command=open_selected_folder).pack(side="left", padx=4)
        tk.Button(action_frame, text="Close", bg=P["CARD"], fg=P["TEXT"],
                  relief="flat", padx=10, pady=8, font=self._f(),
                  cursor="hand2", command=win.destroy).pack(side="right", padx=4)

        usage_tree.bind("<Double-1>", lambda e: open_selected_folder())

    # ── NEW: Full-Text Search ──────────────────────────────────────────────────
    def open_full_text_search(self):
        win = tk.Toplevel(self)
        win.title("📝 Full-Text Search")
        win.geometry(f"{scaled(900)}x{scaled(700)}")
        win.configure(bg=P["BG"])
        win.grab_set()
        win.resizable(True, True)

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(50))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  📝  FULL-TEXT SEARCH", bg=P["HEADER_BG"], fg=P["ACCENT2"],
                 font=self._f(4, bold=True)).pack(side="left", padx=12, pady=12)

        # Search input
        search_frame = tk.Frame(win, bg=P["INPUT_BG"],
                                highlightthickness=1, highlightbackground=P["BORDER"])
        search_frame.pack(fill="x", padx=16, pady=(8, 4))
        query_var = tk.StringVar()
        entry = tk.Entry(search_frame, textvariable=query_var,
                         bg=P["INPUT_BG"], fg=P["TEXT"],
                         insertbackground=P["ACCENT"],
                         font=self._f(2), relief="flat", bd=8,
                         selectbackground=P["SEL"], selectforeground=P["SEL_FG"])
        entry.pack(side="left", fill="x", expand=True)
        entry.focus()

        # Settings
        settings = tk.Frame(win, bg=P["BG"])
        settings.pack(fill="x", padx=16, pady=4)

        # Folder
        src_frame = tk.Frame(settings, bg=P["BG"])
        src_frame.pack(fill="x", pady=2)
        tk.Label(src_frame, text="Search in:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left")
        src_var = tk.StringVar(value=self.current_path.get())
        tk.Entry(src_frame, textvariable=src_var, bg=P["INPUT_BG"], fg=P["TEXT"],
                 relief="flat", font=self._f(), width=40,
                 insertbackground=P["ACCENT"]).pack(side="left", padx=8, ipady=2)
        tk.Button(src_frame, text="Browse", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=8, pady=2, font=self._f(),
                  cursor="hand2",
                  command=lambda: src_var.set(
                      filedialog.askdirectory(title="Select Folder to Search", parent=win)
                      or src_var.get())).pack(side="left")

        # Options
        opts = tk.Frame(settings, bg=P["BG"])
        opts.pack(fill="x", pady=2)
        
        case_var = tk.BooleanVar(value=False)
        tk.Checkbutton(opts, text="Case Sensitive", variable=case_var,
                       bg=P["BG"], fg=P["SUBTEXT"], selectcolor=P["BG"], font=self._f()).pack(side="left", padx=4)
        
        whole_var = tk.BooleanVar(value=False)
        tk.Checkbutton(opts, text="Whole Word", variable=whole_var,
                       bg=P["BG"], fg=P["SUBTEXT"], selectcolor=P["BG"], font=self._f()).pack(side="left", padx=12)
        
        tk.Label(opts, text="Ext (comma):", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left", padx=(12, 4))
        ext_var = tk.StringVar()
        tk.Entry(opts, textvariable=ext_var, width=15,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=self._f(), insertbackground=P["ACCENT"]).pack(side="left", padx=4)

        # Results
        result_frame = tk.Frame(win, bg=P["PANEL"])
        result_frame.pack(fill="both", expand=True, padx=16, pady=4)

        result_tree = ttk.Treeview(result_frame, columns=("name", "path", "size", "modified"), 
                                   show="headings", selectmode="extended")
        for col, label, w, anchor in [
            ("name", "Name", scaled(250), "w"),
            ("path", "Path", scaled(380), "w"),
            ("size", "Size", scaled(80), "e"),
            ("modified", "Modified", scaled(140), "w"),
        ]:
            result_tree.heading(col, text=label)
            result_tree.column(col, width=w, anchor=anchor, stretch=(col == "path"))
        result_tree.tag_configure("even", background=P["EVEN_ROW"], foreground=P["TEXT"])
        result_tree.tag_configure("odd", background=P["ODD_ROW"], foreground=P["TEXT"])
        result_vsb = ttk.Scrollbar(result_frame, orient="vertical", command=result_tree.yview)
        result_tree.configure(yscrollcommand=result_vsb.set)
        result_tree.pack(side="left", fill="both", expand=True)
        result_vsb.pack(side="right", fill="y")

        status_var = tk.StringVar(value="Enter search term and click Search")
        tk.Label(win, textvariable=status_var, bg=P["BG"], fg=P["SUBTEXT"],
                 font=self._f(), anchor="w").pack(fill="x", padx=16, pady=4)
        progress = ttk.Progressbar(win, mode="indeterminate")
        progress.pack(fill="x", padx=16, pady=4)

        searcher = FullTextSearch()

        def do_search():
            q = query_var.get().strip()
            if not q:
                messagebox.showwarning("Empty Query", "Enter text to search for.", parent=win)
                return
            
            root = src_var.get()
            if not os.path.isdir(root):
                messagebox.showerror("Invalid Folder", "Select a valid folder.", parent=win)
                return
            
            result_tree.delete(*result_tree.get_children())
            status_var.set(f"Searching for '{q}'...")
            progress.start(10)
            
            def run():
                try:
                    results = searcher.search(
                        root_path=root,
                        query=q,
                        show_hidden=self.show_hidden.get(),
                        case_sensitive=case_var.get(),
                        whole_word=whole_var.get(),
                        ext_filter=ext_var.get(),
                        max_results=500
                    )
                    win.after(0, lambda: show_results(results))
                except Exception as e:
                    win.after(0, lambda: status_var.set(f"Error: {e}"))

            def show_results(results):
                progress.stop()
                if not results:
                    status_var.set("No matching files found.")
                    return
                
                for idx, path in enumerate(results):
                    try:
                        st = os.stat(path)
                        size = human_size(st.st_size)
                        mtime = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
                        name = os.path.basename(path)
                        dirname = os.path.dirname(path)
                        if len(dirname) > 60:
                            dirname = "..." + dirname[-57:]
                    except:
                        size = "N/A"
                        mtime = "N/A"
                        name = os.path.basename(path)
                        dirname = os.path.dirname(path)
                    
                    tag = ("even", "odd")[idx % 2]
                    result_tree.insert("", "end", values=(name, dirname, size, mtime), 
                                      tags=(tag,), iid=path)
                
                status_var.set(f"✅ Found {len(results)} file(s) containing '{q}'")

            threading.Thread(target=run, daemon=True).start()

        def open_selected():
            sel = result_tree.selection()
            if sel:
                path = sel[0]
                if os.path.exists(path):
                    self.load_dir(os.path.dirname(path))
                    self._open_file(path)
                    win.destroy()

        def show_context():
            sel = result_tree.selection()
            if sel:
                path = sel[0]
                try:
                    with open(path, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read(2000)
                    preview_win = tk.Toplevel(win)
                    preview_win.title(f"Preview: {os.path.basename(path)}")
                    preview_win.geometry(f"{scaled(600)}x{scaled(400)}")
                    preview_win.configure(bg=P["BG"])
                    text = tk.Text(preview_win, bg=P["INPUT_BG"], fg=P["TEXT"],
                                  font=("Courier New", scaled(9)), wrap="none",
                                  relief="flat", bd=10,
                                  selectbackground=P["SEL"], selectforeground=P["SEL_FG"])
                    text.pack(fill="both", expand=True)
                    text.insert("1.0", content)
                    text.config(state="disabled")
                    
                    # Highlight search term
                    q = query_var.get().strip()
                    if q:
                        text.config(state="normal")
                        start = "1.0"
                        while True:
                            pos = text.search(q, start, nocase=not case_var.get())
                            if not pos:
                                break
                            end = f"{pos}+{len(q)}c"
                            text.tag_add("highlight", pos, end)
                            start = end
                        text.tag_config("highlight", background=P["WARNING"], foreground="black")
                        text.config(state="disabled")
                except Exception as e:
                    messagebox.showerror("Preview Error", str(e), parent=win)

        # Buttons
        action_frame = tk.Frame(win, bg=P["BG"])
        action_frame.pack(fill="x", padx=16, pady=8)

        tk.Button(action_frame, text="🔍 Search", bg=P["ACCENT2"], fg="white",
                  relief="flat", padx=16, pady=8, font=self._f(bold=True),
                  cursor="hand2", command=do_search).pack(side="left", padx=4)
        entry.bind("<Return>", lambda e: do_search())
        
        tk.Button(action_frame, text="📂 Open File", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=12, pady=8, font=self._f(),
                  cursor="hand2", command=open_selected).pack(side="left", padx=4)
        
        tk.Button(action_frame, text="👁 Preview", bg=P["SUCCESS"], fg="white",
                  relief="flat", padx=12, pady=8, font=self._f(),
                  cursor="hand2", command=show_context).pack(side="left", padx=4)
        
        tk.Button(action_frame, text="Close", bg=P["CARD"], fg=P["TEXT"],
                  relief="flat", padx=10, pady=8, font=self._f(),
                  cursor="hand2", command=win.destroy).pack(side="right", padx=4)

        result_tree.bind("<Double-1>", lambda e: open_selected())

    # ── Paste to Multiple Dirs ─────────────────────────────────────────────────
    def paste_to_multiple_dirs(self):
        if not self.clipboard:
            messagebox.showwarning("No Clipboard", "Nothing in clipboard.")
            return
        win = tk.Toplevel(self)
        win.title("📥+ Paste to Multiple Directories")
        win.geometry(f"{scaled(720)}x{scaled(520)}")
        win.configure(bg=P["BG"])
        win.grab_set()

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(46))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"  📥  Paste {len(self.clipboard)} item(s) to Multiple Folders",
                 bg=P["HEADER_BG"], fg=P["ACCENT"],
                 font=self._f(3, bold=True)).pack(side="left", padx=12, pady=10)

        src_frame = tk.Frame(win, bg=P["CARD"])
        src_frame.pack(fill="x", padx=16, pady=8)
        tk.Label(src_frame, text="Files to paste:", bg=P["CARD"], fg=P["SUBTEXT"],
                 font=self._f(bold=True)).pack(anchor="w", padx=8, pady=4)
        for src, op in self.clipboard[:8]:
            color = P["WARNING"] if op == "cut" else P["SUCCESS"]
            tk.Label(src_frame, text=f"  {'✂' if op=='cut' else '📋'} {os.path.basename(src)}",
                     bg=P["CARD"], fg=color, font=self._f(), anchor="w").pack(anchor="w", padx=12)
        if len(self.clipboard) > 8:
            tk.Label(src_frame, text=f"  ... and {len(self.clipboard)-8} more",
                     bg=P["CARD"], fg=P["SUBTEXT"], font=self._f()).pack(anchor="w", padx=12)

        tk.Label(win, text="Target Directories:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=self._f(bold=True)).pack(anchor="w", padx=16, pady=(8, 2))
        dir_list_frame = tk.Frame(win, bg=P["CARD"])
        dir_list_frame.pack(fill="both", expand=True, padx=16)
        dir_listbox = tk.Listbox(dir_list_frame, bg=P["CARD"], fg=P["TEXT"],
                                 selectbackground=P["SEL"], selectforeground=P["SEL_FG"],
                                 font=self._f(), relief="flat", bd=0,
                                 selectmode=tk.MULTIPLE, activestyle="none")
        dir_sb = ttk.Scrollbar(dir_list_frame, orient="vertical", command=dir_listbox.yview)
        dir_listbox.configure(yscrollcommand=dir_sb.set)
        dir_listbox.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        dir_sb.pack(side="right", fill="y")

        default_dirs = [self.current_path.get()]
        for name, path in self.user_bookmarks:
            if path not in default_dirs and os.path.isdir(path): default_dirs.append(path)
        for d in default_dirs: dir_listbox.insert("end", d)
        dir_listbox.selection_set(0)

        dir_btn_frame = tk.Frame(win, bg=P["BG"])
        dir_btn_frame.pack(fill="x", padx=16, pady=4)

        def add_dir():
            path = filedialog.askdirectory(title="Add Target Directory", parent=win)
            if path:
                items = list(dir_listbox.get(0, "end"))
                if path not in items: dir_listbox.insert("end", path)

        def remove_dir():
            for idx in reversed(list(dir_listbox.curselection())): dir_listbox.delete(idx)

        for txt, cmd, color in [
            ("➕ Browse...", add_dir, P["ACCENT"]),
            ("➖ Remove", remove_dir, P["DANGER"]),
        ]:
            tk.Button(dir_btn_frame, text=txt, bg=color, fg="white",
                      relief="flat", padx=10, pady=5, font=self._f(),
                      cursor="hand2", command=cmd).pack(side="left", padx=4)

        paste_status = tk.StringVar(value="Select target directories and click Paste")
        tk.Label(win, textvariable=paste_status, bg=P["BG"], fg=P["SUBTEXT"],
                 font=self._f()).pack(anchor="w", padx=16, pady=2)
        paste_progress = ttk.Progressbar(win, mode="determinate")
        paste_progress.pack(padx=16, fill="x")

        def do_paste():
            selected_indices = dir_listbox.curselection()
            if not selected_indices:
                messagebox.showwarning("No Selection", "Select at least one target directory.", parent=win)
                return
            targets = [dir_listbox.get(i) for i in selected_indices]
            total = len(self.clipboard) * len(targets)
            done = [0]
            errors = []
            paste_progress["maximum"] = total

            def run():
                for target in targets:
                    for src, op in self.clipboard:
                        name = os.path.basename(src)
                        dst = os.path.join(target, name)
                        if os.path.exists(dst):
                            base, ext = os.path.splitext(name)
                            dst = os.path.join(target, f"{base}_copy{ext}")
                        try:
                            if os.path.isdir(src): shutil.copytree(src, dst)
                            else: shutil.copy2(src, dst)
                        except Exception as e:
                            errors.append(f"{name} → {target}: {e}")
                        done[0] += 1
                        win.after(0, lambda d=done[0]: paste_progress.config(value=d))
                win.after(0, paste_done)

            def paste_done():
                paste_status.set(f"✅ Pasted to {len(targets)} dir(s)" +
                                 (f" | {len(errors)} error(s)" if errors else ""))
                if errors: messagebox.showerror("Paste Errors", "\n".join(errors[:10]), parent=win)
                if self.clipboard and self.clipboard[0][1] == "cut": self.clipboard.clear()
                self.load_dir(self.current_path.get())

            threading.Thread(target=run, daemon=True).start()

        bottom = tk.Frame(win, bg=P["BG"])
        bottom.pack(fill="x", padx=16, pady=10)
        tk.Button(bottom, text="📥 Paste Now", bg=P["SUCCESS"], fg="white",
                  relief="flat", padx=20, pady=8, font=self._f(1, bold=True),
                  cursor="hand2", command=do_paste).pack(side="left", padx=4)
        tk.Button(bottom, text="Cancel", bg=P["CARD"], fg=P["TEXT"],
                  relief="flat", padx=14, pady=8, font=self._f(),
                  cursor="hand2", command=win.destroy).pack(side="left", padx=4)

    # ── File Operations ────────────────────────────────────────────────────────
    def _get_selection(self):
        if self.tree:
            return self.tree.selection()
        return tuple(self.selected_paths)

    def new_folder(self):
        name = simpledialog.askstring("New Folder", "Folder name:", parent=self)
        if not name: return
        path = os.path.join(self.current_path.get(), name)
        try:
            os.makedirs(path, exist_ok=True)
            self.undo_stack.push(("delete", path))
            self.load_dir(self.current_path.get())
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def new_file(self):
        name = simpledialog.askstring("New File", "File name:", parent=self)
        if not name: return
        path = os.path.join(self.current_path.get(), name)
        try:
            Path(path).touch()
            self.undo_stack.push(("delete", path))
            self.load_dir(self.current_path.get())
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def copy_files(self):
        sel = self._get_selection()
        if not sel: return
        self.clipboard = [(p, "copy") for p in sel]
        self.path_clipboard.add(list(sel))
        self.status_var.set(f"Copied {len(sel)} item(s)")

    def cut_files(self):
        sel = self._get_selection()
        if not sel: return
        self.clipboard = [(p, "cut") for p in sel]
        self.path_clipboard.add(list(sel))
        self.status_var.set(f"Cut {len(sel)} item(s)")

    def paste_files(self):
        if not self.clipboard: return
        dest = self.current_path.get()
        errors = []
        pasted = []
        for src, op in self.clipboard:
            name = os.path.basename(src)
            dst = os.path.join(dest, name)
            if os.path.exists(dst):
                base, ext = os.path.splitext(name)
                dst = os.path.join(dest, f"{base}_copy{ext}")
            try:
                if os.path.isdir(src):
                    if op == "copy": shutil.copytree(src, dst)
                    else: shutil.move(src, dst)
                else:
                    if op == "copy": shutil.copy2(src, dst)
                    else: shutil.move(src, dst)
                pasted.append((dst, op))
            except Exception as e:
                errors.append(str(e))
        if self.clipboard and self.clipboard[0][1] == "cut":
            self.clipboard.clear()
        if pasted: self.undo_stack.push(("paste_undo", pasted))
        self.load_dir(dest)
        if errors: messagebox.showerror("Paste errors", "\n".join(errors))

    def delete_files(self):
        sel = self._get_selection()
        if not sel: return
        names = "\n".join(os.path.basename(p) for p in sel[:8])
        if len(sel) > 8: names += f"\n... and {len(sel)-8} more"
        trash_msg = " (to Trash)" if TRASH_AVAILABLE else " (PERMANENT)"
        if not messagebox.askyesno("Delete", f"Delete {len(sel)} item(s){trash_msg}?\n\n{names}"):
            return
        errors = []
        for p in sel:
            try:
                if TRASH_AVAILABLE: send2trash.send2trash(p)
                else:
                    if os.path.isdir(p): shutil.rmtree(p)
                    else: os.remove(p)
            except Exception as e:
                errors.append(str(e))
        self.load_dir(self.current_path.get())
        if errors: messagebox.showerror("Delete errors", "\n".join(errors))

    def rename_file(self):
        self._start_inline_rename()

    def _duplicate_files(self):
        sel = self._get_selection()
        if not sel: return
        for src in sel:
            base, ext = os.path.splitext(os.path.basename(src))
            dst = os.path.join(os.path.dirname(src), f"{base}_copy{ext}")
            counter = 1
            while os.path.exists(dst):
                dst = os.path.join(os.path.dirname(src), f"{base}_copy{counter}{ext}")
                counter += 1
            try:
                if os.path.isdir(src): shutil.copytree(src, dst)
                else: shutil.copy2(src, dst)
            except Exception as e:
                messagebox.showerror("Duplicate Error", str(e))
        self.load_dir(self.current_path.get())
        self.status_var.set(f"Duplicated {len(sel)} item(s)")

    def undo_action(self):
        action = self.undo_stack.pop()
        if not action:
            self.status_var.set("Nothing to undo.")
            return
        try:
            op = action[0]
            if op == "delete":
                path = action[1]
                if os.path.isdir(path): shutil.rmtree(path)
                elif os.path.exists(path): os.remove(path)
            elif op == "rename":
                os.rename(action[1], action[2])
            elif op == "paste_undo":
                for dst, op_type in action[1]:
                    if op_type == "copy":
                        if os.path.isdir(dst): shutil.rmtree(dst)
                        elif os.path.exists(dst): os.remove(dst)
            elif op == "batch_rename":
                for new_path, old_path in action[1]:
                    if os.path.exists(new_path):
                        os.rename(new_path, old_path)
            elif op == "auto_organize":
                for new_path, old_path in action[1]:
                    if os.path.exists(new_path):
                        os.makedirs(os.path.dirname(old_path), exist_ok=True)
                        shutil.move(new_path, old_path)
            self.load_dir(self.current_path.get())
            self.status_var.set("Undo successful.")
        except Exception as e:
            messagebox.showerror("Undo Error", str(e))

    def redo_action(self):
        self.status_var.set("Redo: limited support — use manual operations.")

    # ── Batch Rename ──────────────────────────────────────────────────────────
    def batch_rename(self):
        sel = self._get_selection()
        if not sel:
            messagebox.showwarning("No Selection", "Select files to batch rename.")
            return
        self.batch_rename_paths(list(sel), parent=self)
        self.load_dir(self.current_path.get())

    def batch_rename_paths(self, paths, parent=None):
        if not paths: return
        if parent is None: parent = self
        win = tk.Toplevel(parent)
        win.title("🔀 Batch Rename")
        win.geometry(f"{scaled(640)}x{scaled(540)}")
        win.configure(bg=P["BG"])
        win.grab_set()

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(44))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"  🔀  Batch Rename — {len(paths)} file(s)",
                 bg=P["HEADER_BG"], fg=P["ACCENT"],
                 font=(self.ui_font_name, scaled(12), "bold")).pack(side="left", padx=12, pady=10)

        body = tk.Frame(win, bg=P["BG"])
        body.pack(fill="both", expand=True, padx=16, pady=10)

        fields_frame = tk.Frame(body, bg=P["CARD"])
        fields_frame.pack(fill="x", pady=(0, 8))
        fields = {}
        for label, key, default in [
            ("Find (text):", "find", ""), ("Replace with:", "replace", ""),
            ("Add prefix:", "prefix", ""), ("Add suffix (before ext):", "suffix", ""),
            ("Start counter at:", "counter", "1"), ("Counter step:", "step", "1"),
        ]:
            row = tk.Frame(fields_frame, bg=P["CARD"])
            row.pack(fill="x", pady=3)
            tk.Label(row, text=label, bg=P["CARD"], fg=P["SUBTEXT"],
                     font=self._f(), width=26, anchor="w").pack(side="left", padx=8)
            var = tk.StringVar(value=default)
            fields[key] = var
            tk.Entry(row, textvariable=var, bg=P["INPUT_BG"], fg=P["TEXT"],
                     relief="flat", font=self._f(), width=28,
                     insertbackground=P["ACCENT"],
                     selectbackground=P["SEL"],
                     selectforeground=P["SEL_FG"]).pack(side="left", padx=4)

        check_frame = tk.Frame(body, bg=P["BG"])
        check_frame.pack(fill="x", pady=4)
        use_counter = tk.BooleanVar(value=False)
        change_ext = tk.BooleanVar(value=False)
        to_lower = tk.BooleanVar(value=False)
        to_upper = tk.BooleanVar(value=False)
        new_ext_var = tk.StringVar()
        for var, lbl in [(use_counter, "Append counter"), (to_lower, "To lowercase"),
                         (to_upper, "To UPPERCASE"), (change_ext, "Change ext:")]:
            tk.Checkbutton(check_frame, text=lbl, variable=var,
                           bg=P["BG"], fg=P["TEXT"], selectcolor=P["BG"],
                           activebackground=P["BG"], font=self._f()).pack(side="left", padx=6)
        tk.Entry(check_frame, textvariable=new_ext_var, width=6,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=self._f(), insertbackground=P["ACCENT"]).pack(side="left", padx=2)

        tk.Label(body, text="Preview:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=self._f(bold=True)).pack(anchor="w", pady=(6, 2))
        preview_box = tk.Text(body, bg=P["INPUT_BG"], fg=P["TEXT"],
                              font=self._f(), height=10, relief="flat", bd=8,
                              selectbackground=P["SEL"], selectforeground=P["SEL_FG"])
        preview_box.pack(fill="both", expand=True)
        preview_box.tag_config("old", foreground=P["SUBTEXT"])
        preview_box.tag_config("arr", foreground=P["ACCENT"])
        preview_box.tag_config("new", foreground=P["SUCCESS"])

        def compute_new_name(old_name, idx):
            base, ext = os.path.splitext(old_name)
            new = base
            if fields["find"].get():
                new = new.replace(fields["find"].get(), fields["replace"].get())
            new = fields["prefix"].get() + new + fields["suffix"].get()
            if to_lower.get(): new = new.lower()
            if to_upper.get(): new = new.upper()
            if use_counter.get():
                try:
                    start = int(fields["counter"].get() or 1)
                    step = int(fields["step"].get() or 1)
                    cnt = start + idx * step
                except: cnt = idx + 1
                new = f"{new}_{cnt:03d}"
            if change_ext.get() and new_ext_var.get().strip():
                ext_new = new_ext_var.get().strip()
                if not ext_new.startswith("."): ext_new = "." + ext_new
                ext = ext_new
            return new + ext

        def do_preview():
            preview_box.config(state="normal")
            preview_box.delete("1.0", "end")
            for idx, path in enumerate(paths):
                old = os.path.basename(path)
                new = compute_new_name(old, idx)
                preview_box.insert("end", f"  {old}", "old")
                preview_box.insert("end", "  →  ", "arr")
                preview_box.insert("end", f"{new}\n", "new")
            preview_box.config(state="disabled")

        def do_apply():
            errors = []
            renamed = []
            for idx, path in enumerate(paths):
                old = os.path.basename(path)
                new = compute_new_name(old, idx)
                if new == old: continue
                new_path = os.path.join(os.path.dirname(path), new)
                try:
                    os.rename(path, new_path)
                    renamed.append((new_path, path))
                except Exception as e:
                    errors.append(f"{old}: {e}")
            if renamed:
                self.undo_stack.push(("batch_rename", renamed))
            win.destroy()
            self.load_dir(self.current_path.get())
            if errors: messagebox.showerror("Batch Rename Errors", "\n".join(errors))
            else: self.status_var.set(f"Batch renamed {len(renamed)} file(s)")

        btn_row = tk.Frame(win, bg=P["BG"])
        btn_row.pack(pady=8)
        for txt, cmd, color in [
            ("👁 Preview", do_preview, P["ACCENT"]),
            ("✅ Apply", do_apply, P["SUCCESS"]),
            ("Cancel", win.destroy, P["DANGER"])
        ]:
            tk.Button(btn_row, text=txt, bg=color, fg="white",
                      relief="flat", padx=16, pady=7, font=self._f(bold=True),
                      cursor="hand2", command=cmd).pack(side="left", padx=6)
        do_preview()

    def compress_selected(self):
        sel = self._get_selection()
        if not sel:
            messagebox.showwarning("No Selection", "Select files/folders to compress.")
            return
        fmt = simpledialog.askstring("Archive Format", "Format? (zip / tar.gz / tar.bz2):",
                                     initialvalue="zip", parent=self)
        if not fmt: return
        fmt = fmt.strip().lower()
        default_name = f"archive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{fmt}"
        zip_path = filedialog.asksaveasfilename(
            title="Save Archive", defaultextension=f".{fmt}",
            initialfile=default_name,
            filetypes=[("Archives", "*.zip *.tar.gz *.tar.bz2")])
        if not zip_path: return
        try:
            if fmt == "zip":
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    for item in sel:
                        if os.path.isdir(item):
                            for root, dirs, files in os.walk(item):
                                for file in files:
                                    fp = os.path.join(root, file)
                                    rp = os.path.relpath(fp, os.path.dirname(item))
                                    zf.write(fp, arcname=rp)
                        else: zf.write(item, arcname=os.path.basename(item))
            elif fmt in ("tar.gz", "tar.bz2"):
                mode = "w:gz" if fmt == "tar.gz" else "w:bz2"
                with tarfile.open(zip_path, mode) as tf:
                    for item in sel: tf.add(item, arcname=os.path.basename(item))
            self.status_var.set(f"Created: {os.path.basename(zip_path)}")
            self.load_dir(self.current_path.get())
        except Exception as e:
            messagebox.showerror("Compression Error", str(e))

    def extract_archive(self):
        sel = self._get_selection()
        path = sel[0] if sel else None
        if not path:
            path = filedialog.askopenfilename(
                title="Select Archive",
                filetypes=[("Archives", "*.zip *.tar *.tar.gz *.tar.bz2 *.tgz")])
        if not path: return
        dest = filedialog.askdirectory(title="Extract to:", initialdir=self.current_path.get())
        if not dest: return
        try:
            ext = path.lower()
            if ext.endswith(".zip"):
                with zipfile.ZipFile(path, "r") as zf: zf.extractall(dest)
            elif any(ext.endswith(e) for e in (".tar.gz", ".tgz", ".tar.bz2", ".tar")):
                with tarfile.open(path, "r:*") as tf: tf.extractall(dest)
            self.status_var.set(f"Extracted to: {dest}")
            self.load_dir(dest)
        except Exception as e:
            messagebox.showerror("Extract Error", str(e))

    def create_symlink(self):
        sel = self._get_selection()
        if not sel:
            messagebox.showwarning("No Selection", "Select a file/folder to symlink.")
            return
        src = sel[0]
        link_name = simpledialog.askstring("Create Symlink", "Link name:",
                                           initialvalue=os.path.basename(src) + "_link", parent=self)
        if not link_name: return
        link_path = os.path.join(self.current_path.get(), link_name)
        try:
            os.symlink(src, link_path)
            self.status_var.set(f"Symlink created: {link_name}")
            self.load_dir(self.current_path.get())
        except Exception as e:
            messagebox.showerror("Symlink Error", str(e))

    def compare_files(self):
        sel = self._get_selection()
        if len(sel) == 2: f1, f2 = sel[0], sel[1]
        elif len(sel) == 1:
            f1 = sel[0]
            f2 = filedialog.askopenfilename(title="Select second file to compare")
            if not f2: return
        else:
            messagebox.showwarning("Compare", "Select 1 or 2 files to compare.")
            return
        win = tk.Toplevel(self)
        win.title(f"Compare: {os.path.basename(f1)}  vs  {os.path.basename(f2)}")
        win.geometry(f"{scaled(900)}x{scaled(600)}")
        win.configure(bg=P["BG"])
        header = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(32))
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text=f"  {os.path.basename(f1)}", bg=P["HEADER_BG"],
                 fg=P["ACCENT"], font=self._f(1, bold=True)).pack(side="left", padx=10, pady=8)
        tk.Label(header, text=f"  {os.path.basename(f2)}", bg=P["HEADER_BG"],
                 fg=P["ACCENT2"], font=self._f(1, bold=True)).pack(side="right", padx=10, pady=8)
        txt = tk.Text(win, bg=P["PANEL"], fg=P["TEXT"], font=self._f(),
                      wrap="none", relief="flat", bd=8,
                      selectbackground=P["SEL"], selectforeground=P["SEL_FG"])
        txt.pack(fill="both", expand=True, padx=6, pady=6)
        if self.theme == "light":
            txt.tag_config("add", background="#dcfce7", foreground="#15803d")
            txt.tag_config("rem", background="#fee2e2", foreground="#b91c1c")
        else:
            txt.tag_config("add", background="#1a3a1a", foreground="#4ade80")
            txt.tag_config("rem", background="#3a1a1a", foreground="#f87171")
        txt.tag_config("hdr", foreground=P["ACCENT2"], font=self._f(bold=True))
        try:
            with open(f1, "r", encoding="utf-8", errors="replace") as fh: lines1 = fh.readlines()
            with open(f2, "r", encoding="utf-8", errors="replace") as fh: lines2 = fh.readlines()
            diff = difflib.unified_diff(lines1, lines2,
                                        fromfile=os.path.basename(f1), tofile=os.path.basename(f2))
            txt.config(state="normal")
            for line in diff:
                if line.startswith("---") or line.startswith("+++"): txt.insert("end", line, "hdr")
                elif line.startswith("+"): txt.insert("end", line, "add")
                elif line.startswith("-"): txt.insert("end", line, "rem")
                else: txt.insert("end", line)
            txt.config(state="disabled")
        except Exception as e:
            txt.config(state="normal")
            txt.insert("end", f"Error: {e}")
            txt.config(state="disabled")

    def manage_tags(self):
        sel = self._get_selection()
        if not sel:
            messagebox.showwarning("No Selection", "Select files to tag.")
            return
        win = tk.Toplevel(self)
        win.title("Manage Tags")
        win.geometry(f"{scaled(380)}x{scaled(320)}")
        win.configure(bg=P["BG"])
        win.grab_set()
        tk.Label(win, text="🏷️  Tag Management", bg=P["BG"], fg=P["TEXT"],
                 font=self._f(3, bold=True)).pack(pady=10)
        tk.Label(win, text=f"{len(sel)} file(s) selected", bg=P["BG"], fg=P["SUBTEXT"],
                 font=self._f()).pack()
        current_tags = set()
        for p in sel: current_tags.update(self.tag_manager.get_tags(p))
        tag_frame = tk.Frame(win, bg=P["CARD"])
        tag_frame.pack(fill="x", padx=20, pady=10)
        tk.Label(tag_frame, text="Color labels:", bg=P["CARD"], fg=P["SUBTEXT"],
                 font=self._f()).pack(anchor="w", padx=8, pady=4)
        vars_map = {}
        for tag_name, color in self.tag_manager.colors.items():
            var = tk.BooleanVar(value=tag_name in current_tags)
            vars_map[tag_name] = var
            row = tk.Frame(tag_frame, bg=P["CARD"])
            row.pack(fill="x", padx=8, pady=2)
            tk.Label(row, text="●", bg=P["CARD"], fg=color, font=("Segoe UI", scaled(12))).pack(side="left")
            tk.Checkbutton(row, text=tag_name, variable=var,
                           bg=P["CARD"], fg=P["TEXT"], selectcolor=P["CARD"],
                           activebackground=P["CARD"], font=self._f()).pack(side="left", padx=4)
        custom_var = tk.StringVar()
        row = tk.Frame(win, bg=P["BG"])
        row.pack(fill="x", padx=20, pady=4)
        tk.Label(row, text="Custom tag:", bg=P["BG"], fg=P["SUBTEXT"], font=self._f()).pack(side="left")
        tk.Entry(row, textvariable=custom_var, bg=P["INPUT_BG"], fg=P["TEXT"],
                 relief="flat", font=self._f(), width=18,
                 insertbackground=P["ACCENT"],
                 selectbackground=P["SEL"]).pack(side="left", padx=6)

        def apply_tags():
            for p in sel:
                for t in list(self.tag_manager.colors.keys()): self.tag_manager.remove_tag(p, t)
                for tag_name, var in vars_map.items():
                    if var.get(): self.tag_manager.add_tag(p, tag_name)
                ct = custom_var.get().strip()
                if ct: self.tag_manager.add_tag(p, ct)
            win.destroy()
            self._populate(self.current_path.get(), self.search_bar.get())

        tk.Button(win, text="Apply Tags", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=20, pady=6, font=self._f(1, bold=True),
                  cursor="hand2", command=apply_tags).pack(pady=10)

    def show_props(self):
        sel = self._get_selection()
        if not sel: return
        self.show_props_for(sel[0])

    def show_props_for(self, path):
        try:
            st = os.stat(path)
            is_dir = os.path.isdir(path)
            if is_dir:
                total = sum(os.path.getsize(os.path.join(dp, f))
                            for dp, dn, fn in os.walk(path) for f in fn)
                size_str = human_size(total)
                count_str = str(sum(1 for _ in os.walk(path)))
            else:
                size_str = human_size(st.st_size)
                count_str = ""
            md5 = sha256 = ""
            if not is_dir and st.st_size < 100 * 1024 * 1024:
                try:
                    h_md5 = hashlib.md5()
                    h_sha = hashlib.sha256()
                    with open(path, "rb") as f:
                        for chunk in iter(lambda: f.read(65536), b""):
                            h_md5.update(chunk)
                            h_sha.update(chunk)
                    md5 = h_md5.hexdigest()
                    sha256 = h_sha.hexdigest()
                except: pass
            is_link = os.path.islink(path)
            link_target = os.readlink(path) if is_link else ""
            tags = self.tag_manager.get_tags(path)
            mark = self.color_mark_manager.get_mark(path)
            info = (
                f"Name:        {os.path.basename(path)}\n"
                f"Path:        {path}\n"
                f"Type:        {'Directory' if is_dir else Path(path).suffix.upper()+' File'}\n"
                f"Size:        {size_str}\n"
                + (f"Subfolders:  {count_str}\n" if is_dir else "")
                + (f"Symlink to:  {link_target}\n" if is_link else "")
                + f"Created:     {datetime.fromtimestamp(st.st_ctime).strftime('%Y-%m-%d %H:%M:%S')}\n"
                  f"Modified:    {datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S')}\n"
                  f"Accessed:    {datetime.fromtimestamp(st.st_atime).strftime('%Y-%m-%d %H:%M:%S')}\n"
                  f"Permissions: {stat.filemode(st.st_mode)}\n"
                + (f"MD5:         {md5}\n" if md5 else "")
                + (f"SHA256:      {sha256}\n" if sha256 else "")
                + (f"Tags:        {', '.join(tags)}\n" if tags else "")
                + (f"Color Mark:  {mark}\n" if mark != "none" else "")
            )
            messagebox.showinfo(f"Properties — {os.path.basename(path)}", info)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def copy_selected_path(self):
        sel = self._get_selection()
        if not sel: return
        paths = "\n".join(sel)
        self.clipboard_clear()
        self.clipboard_append(paths)
        self.path_clipboard.add(list(sel))
        self.status_var.set(f"Copied {len(sel)} path(s) to clipboard")

    def open_terminal_here(self):
        current = self.current_path.get()
        try:
            if sys.platform == "win32":
                subprocess.Popen(f'start cmd /K "cd /d {current}"', shell=True)
            elif sys.platform == "darwin":
                script = f'tell application "Terminal" to do script "cd \\"{current}\\""'
                subprocess.Popen(["osascript", "-e", script])
            else:
                terminals = ["gnome-terminal", "xfce4-terminal", "konsole", "lxterminal", "xterm"]
                for term in terminals:
                    if shutil.which(term):
                        if term == "gnome-terminal":
                            subprocess.Popen([term, f"--working-directory={current}"])
                        else:
                            subprocess.Popen([term, "--working-directory", current])
                        return
                subprocess.Popen(["xterm", "-e", f"cd '{current}'; exec bash"])
        except Exception as e:
            messagebox.showerror("Terminal Error", str(e))

    def folder_sync(self):
        src = filedialog.askdirectory(title="Source folder:")
        if not src: return
        dst = filedialog.askdirectory(title="Destination folder:")
        if not dst: return
        copied = skipped = errors = 0
        for dirpath, dirs, files in os.walk(src):
            rel = os.path.relpath(dirpath, src)
            dest_dir = os.path.join(dst, rel)
            os.makedirs(dest_dir, exist_ok=True)
            for fname in files:
                src_f = os.path.join(dirpath, fname)
                dst_f = os.path.join(dest_dir, fname)
                try:
                    if not os.path.exists(dst_f):
                        shutil.copy2(src_f, dst_f); copied += 1
                    else: skipped += 1
                except: errors += 1
        messagebox.showinfo("Sync Complete",
                            f"Copied: {copied}\nSkipped (exists): {skipped}\nErrors: {errors}")

    # ── Context Menu ──────────────────────────────────────────────────────────
    def _context_menu(self, event):
        if not self.tree: return
        item = self.tree.identify_row(event.y)
        if item and item not in self.tree.selection():
            self.tree.selection_set(item)
        self._context_menu_for_path(event, item)

    def _context_menu_for_path(self, event, item=None):
        menu = tk.Menu(self, tearoff=0,
                       bg=P["CARD"], fg=P["TEXT"],
                       activebackground=P["SEL"],
                       activeforeground=P["SEL_FG"],
                       font=self._f(), bd=0, relief="flat")
        menu.add_command(label="🚀  Open",                 command=self._on_double_click)
        menu.add_separator()
        menu.add_command(label="✂   Cut  (Ctrl+X)",        command=self.cut_files)
        menu.add_command(label="📋  Copy  (Ctrl+C)",       command=self.copy_files)
        menu.add_command(label="📥  Paste  (Ctrl+V)",      command=self.paste_files)
        menu.add_command(label="📥+  Paste to Multiple",   command=self.paste_to_multiple_dirs)
        menu.add_command(label="⬛  Duplicate",             command=self._duplicate_files)
        menu.add_separator()
        menu.add_command(label="✏   Rename  (F2)",         command=self._start_inline_rename)
        menu.add_command(label="🔀  Batch Rename",         command=self.batch_rename)
        menu.add_command(label="🗑   Delete  (Del)",        command=self.delete_files)
        menu.add_command(label="↩   Undo  (Ctrl+Z)",       command=self.undo_action)
        menu.add_command(label="☑   Select All  (Ctrl+A)", command=self._select_all)
        menu.add_separator()
        mark_menu = tk.Menu(menu, tearoff=0, bg=P["CARD"], fg=P["TEXT"],
                            activebackground=P["SEL"], activeforeground=P["SEL_FG"],
                            font=self._f())
        for cname, cdata in MARK_COLORS.items():
            dot = "✕" if cname == "none" else "●"
            fg = cdata["fg"] if cname != "none" else P["SUBTEXT"]
            mark_menu.add_command(label=f"{dot}  {cname.capitalize()}", foreground=fg,
                                  command=lambda c=cname: self._quick_mark_selection(c))
        menu.add_cascade(label="🎨  Color Mark", menu=mark_menu)
        menu.add_separator()
        menu.add_command(label="📦  Compress  (Ctrl+M)",    command=self.compress_selected)
        menu.add_command(label="📂  Extract Here",           command=self.extract_archive)
        menu.add_command(label="🔗  Create Symlink",         command=self.create_symlink)
        menu.add_separator()
        menu.add_command(label="⚖️   Compare",               command=self.compare_files)
        menu.add_command(label="🏷️   Manage Tags",           command=self.manage_tags)
        menu.add_separator()
        menu.add_command(label="📦+  Bulk Create",           command=self.open_bulk_create)
        menu.add_command(label="🗂️   Auto-Organize",         command=self.open_auto_organize)
        menu.add_command(label="📑   Export PDF",            command=self.open_export_dialog_pdf)
        menu.add_command(label="📊   Export Excel",          command=self.open_export_dialog_excel)
        menu.add_separator()
        menu.add_command(label="🔍   Find Duplicates",       command=self.open_duplicate_finder)
        menu.add_command(label="💾   Disk Usage",            command=self.open_disk_usage)
        menu.add_command(label="📝   Full-Text Search",      command=self.open_full_text_search)
        menu.add_separator()
        menu.add_command(label="🌐   PC-Wide Search",        command=self.open_system_search)
        menu.add_command(label="📋   Copy Path",             command=self.copy_selected_path)
        menu.add_command(label="💻   Terminal Here  (F4)",   command=self.open_terminal_here)
        menu.add_command(label="🔄   Folder Sync",           command=self.folder_sync)
        menu.add_separator()
        menu.add_command(label="📊   Properties  (Alt+Enter)", command=self.show_props)
        menu.tk_popup(event.x_root, event.y_root)

    def _quick_mark_selection(self, color_name):
        sel = self._get_selection()
        for path in sel:
            self.color_mark_manager.set_mark(path, color_name)
            self._apply_mark_tag(path, color_name)
        self.status_var.set(f"Marked {len(sel)} item(s) as '{color_name}'")

    # ══════════════════════════════════════════════════════════════════════════
    # ── AUTO-ORGANIZE FILES (Modern) ─────────────────────────────────────────
    # ══════════════════════════════════════════════════════════════════════════
    def open_auto_organize(self):
        win = tk.Toplevel(self)
        win.title("🗂️ Auto-Organize Files")
        win.geometry(f"{scaled(720)}x{scaled(700)}")
        win.configure(bg=P["BG"])
        win.grab_set()
        win.resizable(True, True)

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(52))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  🗂️  Auto-Organize — Smart File Sorting",
                 bg=P["HEADER_BG"], fg=P["ACCENT"],
                 font=(self.ui_font_name, scaled(13), "bold")).pack(side="left", padx=12, pady=12)

        body = tk.Frame(win, bg=P["BG"])
        body.pack(fill="both", expand=True, padx=20, pady=12)

        # Source folder with modern style
        tk.Label(body, text="Folder to organize:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10), "bold")).pack(anchor="w")
        src_frame = tk.Frame(body, bg=P["BG"])
        src_frame.pack(fill="x", pady=4)
        src_var = tk.StringVar(value=self.current_path.get())
        tk.Entry(src_frame, textvariable=src_var, bg=P["INPUT_BG"], fg=P["TEXT"],
                 relief="flat", font=(self.ui_font_name, scaled(9)),
                 insertbackground=P["ACCENT"],
                 selectbackground=P["SEL"], selectforeground=P["SEL_FG"]
                 ).pack(side="left", fill="x", expand=True, ipady=4)
        tk.Button(src_frame, text="Browse", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=10, pady=4, font=(self.ui_font_name, scaled(9)),
                  cursor="hand2",
                  command=lambda: src_var.set(
                      filedialog.askdirectory(title="Select Folder to Organize", parent=win)
                      or src_var.get())).pack(side="left", padx=6)

        # Organize mode with better UX
        tk.Label(body, text="Auto-generate subfolders by:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10), "bold")).pack(anchor="w", pady=(12, 4))
        mode_outer = tk.Frame(body, bg=P["CARD_BORDER"])
        mode_outer.pack(fill="x", pady=(0, 8))
        mode_frame = tk.Frame(mode_outer, bg=P["CARD"])
        mode_frame.pack(fill="x", padx=1, pady=1)
        mode_var = tk.StringVar(value="ext")
        mode_options = [
            ("ext", "🧩  File Extension    — e.g.  PDF Files/, JPG Files/, No Extension/"),
            ("date_ym", "📅  Modified Date (Year-Month)  — e.g.  2026-07/"),
            ("date_y", "📅  Modified Date (Year only)  — e.g.  2026/"),
            ("name", "🔤  First Letter of Name  — e.g.  A/, B/, 0-9/, #/"),
        ]
        for val, lbl in mode_options:
            tk.Radiobutton(mode_frame, text=lbl, variable=mode_var, value=val,
                           bg=P["CARD"], fg=P["TEXT"], selectcolor=P["CARD"],
                           activebackground=P["CARD"], font=(self.ui_font_name, scaled(9)),
                           pady=5, anchor="w", cursor="hand2").pack(fill="x", padx=12)

        # Options
        opts_frame = tk.Frame(body, bg=P["BG"])
        opts_frame.pack(fill="x", pady=6)
        recursive_var = tk.BooleanVar(value=False)
        tk.Checkbutton(opts_frame, text="Include files in subfolders too", variable=recursive_var,
                       bg=P["BG"], fg=P["SUBTEXT"], selectcolor=P["BG"],
                       activebackground=P["BG"], font=(self.ui_font_name, scaled(9))
                       ).pack(side="left", padx=(0, 14))
        copy_mode_var = tk.BooleanVar(value=False)
        tk.Checkbutton(opts_frame, text="Copy instead of move", variable=copy_mode_var,
                       bg=P["BG"], fg=P["SUBTEXT"], selectcolor=P["BG"],
                       activebackground=P["BG"], font=(self.ui_font_name, scaled(9))
                       ).pack(side="left", padx=(0, 14))
        hidden_var = tk.BooleanVar(value=False)
        tk.Checkbutton(opts_frame, text="Include hidden files", variable=hidden_var,
                       bg=P["BG"], fg=P["SUBTEXT"], selectcolor=P["BG"],
                       activebackground=P["BG"], font=(self.ui_font_name, scaled(9))
                       ).pack(side="left")

        # Quick Apply button
        quick_apply_frame = tk.Frame(body, bg=P["BG"])
        quick_apply_frame.pack(fill="x", pady=10)
        tk.Label(quick_apply_frame, text="Quick Apply:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10), "bold")).pack(side="left", padx=(0, 10))
        quick_apply_btn = tk.Button(quick_apply_frame, text="✅ Apply Now", bg=P["SUCCESS"], fg="white",
                                    relief="flat", padx=20, pady=8,
                                    font=(self.ui_font_name, scaled(11), "bold"),
                                    cursor="hand2", command=lambda: do_organize())
        quick_apply_btn.pack(side="left")
        Tooltip(quick_apply_btn, "Apply organization immediately without preview")

        tk.Label(body, text="Preview:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10), "bold")).pack(anchor="w", pady=(8, 2))
        preview_outer = tk.Frame(body, bg=P["BORDER"])
        preview_outer.pack(fill="both", expand=True)
        preview_box = tk.Text(preview_outer, bg=P["INPUT_BG"], fg=P["TEXT"],
                              font=("Courier New", scaled(9)), relief="flat", bd=8,
                              selectbackground=P["SEL"], selectforeground=P["SEL_FG"],
                              wrap="none")
        preview_sb = ttk.Scrollbar(preview_outer, orient="vertical", command=preview_box.yview)
        preview_box.configure(yscrollcommand=preview_sb.set)
        preview_box.pack(side="left", fill="both", expand=True)
        preview_sb.pack(side="right", fill="y")
        preview_box.tag_config("folder", foreground=P["ACCENT"])
        preview_box.tag_config("file", foreground=P["SUBTEXT"])

        status_var = tk.StringVar(value="Choose options above, then click Preview or Apply Now.")
        tk.Label(win, textvariable=status_var, bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(9)), anchor="w").pack(fill="x", padx=20)

        computed = {"plan": []}

        def get_files():
            base = src_var.get()
            if not os.path.isdir(base):
                return None
            results = []
            if recursive_var.get():
                for dirpath, dirs, files in os.walk(base):
                    if not hidden_var.get():
                        dirs[:] = [d for d in dirs if not d.startswith(".")]
                    for fname in files:
                        if not hidden_var.get() and fname.startswith("."):
                            continue
                        results.append(os.path.join(dirpath, fname))
            else:
                for entry in os.scandir(base):
                    if entry.is_file(follow_symlinks=False):
                        if not hidden_var.get() and entry.name.startswith("."):
                            continue
                        results.append(entry.path)
            return results

        def target_subfolder(fpath):
            mode = mode_var.get()
            name = os.path.basename(fpath)
            if mode == "ext":
                ext = Path(name).suffix.lower().lstrip(".")
                return (ext.upper() + " Files") if ext else "No Extension"
            elif mode == "date_ym":
                try:
                    mt = os.path.getmtime(fpath)
                    return datetime.fromtimestamp(mt).strftime("%Y-%m")
                except Exception:
                    return "Unknown Date"
            elif mode == "date_y":
                try:
                    mt = os.path.getmtime(fpath)
                    return datetime.fromtimestamp(mt).strftime("%Y")
                except Exception:
                    return "Unknown Date"
            elif mode == "name":
                c = name[0].upper() if name else "#"
                if c.isdigit():
                    return "0-9"
                elif c.isalpha():
                    return c
                else:
                    return "#"
            return "Other"

        def do_preview():
            base = src_var.get()
            if not os.path.isdir(base):
                messagebox.showerror("Invalid Folder", "Select a valid folder first.", parent=win)
                return
            files = get_files()
            if not files:
                preview_box.config(state="normal")
                preview_box.delete("1.0", "end")
                preview_box.config(state="disabled")
                status_var.set("No files found to organize in this folder.")
                computed["plan"] = []
                return
            plan = [(fp, target_subfolder(fp)) for fp in files]
            computed["plan"] = plan
            preview_box.config(state="normal")
            preview_box.delete("1.0", "end")
            grouped = {}
            for fp, sub in plan:
                grouped.setdefault(sub, []).append(fp)
            for sub in sorted(grouped.keys()):
                items = grouped[sub]
                preview_box.insert("end", f"📁 {sub}/   ({len(items)} file(s))\n", "folder")
                for fp in items[:6]:
                    preview_box.insert("end", f"      • {os.path.basename(fp)}\n", "file")
                if len(items) > 6:
                    preview_box.insert("end", f"      ... +{len(items) - 6} more\n", "file")
            preview_box.config(state="disabled")
            status_var.set(f"Preview ready: {len(plan)} file(s) → {len(grouped)} folder(s)")

        def do_organize():
            if not computed["plan"]:
                do_preview()
            plan = computed["plan"]
            if not plan:
                messagebox.showinfo("Nothing to do", "No files found to organize.", parent=win)
                return
            base = src_var.get()
            is_copy = copy_mode_var.get()
            folder_count = len(set(sub for _, sub in plan))
            if not messagebox.askyesno(
                "Confirm Auto-Organize",
                f"{'Copy' if is_copy else 'Move'} {len(plan)} file(s) into "
                f"{folder_count} auto-generated subfolder(s) inside:\n\n{base}\n\nProceed?",
                parent=win):
                return
            moved = []
            errors = []
            for fp, sub in plan:
                target_dir = os.path.join(base, sub)
                try:
                    os.makedirs(target_dir, exist_ok=True)
                    dst = os.path.join(target_dir, os.path.basename(fp))
                    if os.path.abspath(os.path.dirname(fp)) == os.path.abspath(target_dir):
                        continue
                    if os.path.exists(dst):
                        base_name, ext = os.path.splitext(os.path.basename(fp))
                        counter = 1
                        while os.path.exists(dst):
                            dst = os.path.join(target_dir, f"{base_name}_{counter}{ext}")
                            counter += 1
                    if is_copy:
                        shutil.copy2(fp, dst)
                    else:
                        shutil.move(fp, dst)
                        moved.append((dst, fp))
                except Exception as e:
                    errors.append(f"{os.path.basename(fp)}: {e}")

            if moved:
                self.undo_stack.push(("auto_organize", moved))

            ok_count = len(plan) - len(errors)
            status_var.set(f"✅ Organized {ok_count} file(s) into {folder_count} folder(s)"
                            + (f" | {len(errors)} error(s)" if errors else ""))
            self.load_dir(self.current_path.get())
            if errors:
                messagebox.showerror("Some errors occurred", "\n".join(errors[:10]), parent=win)
            else:
                messagebox.showinfo(
                    "Auto-Organize Complete",
                    f"Successfully {'copied' if is_copy else 'moved'} {ok_count} file(s) into "
                    f"{folder_count} folder(s).", parent=win)

        btn_frame = tk.Frame(win, bg=P["BG"])
        btn_frame.pack(fill="x", padx=20, pady=(8, 16))
        tk.Button(btn_frame, text="👁 Preview", bg=P["SUBTEXT"], fg="white",
                  relief="flat", padx=16, pady=8, font=(self.ui_font_name, scaled(10), "bold"),
                  cursor="hand2", command=do_preview).pack(side="left", padx=4)
        tk.Button(btn_frame, text="🗂️ Organize Now", bg=P["SUCCESS"], fg="white",
                  relief="flat", padx=20, pady=8, font=(self.ui_font_name, scaled(11), "bold"),
                  cursor="hand2", command=do_organize).pack(side="left", padx=4)
        tk.Button(btn_frame, text="Cancel", bg=P["CARD"], fg=P["TEXT"],
                  relief="flat", padx=14, pady=8, font=(self.ui_font_name, scaled(9)),
                  cursor="hand2", command=win.destroy).pack(side="left", padx=4)

    # ══════════════════════════════════════════════════════════════════════════
    # ── BULK CREATE DIALOG (Modern) ──────────────────────────────────────────
    # ══════════════════════════════════════════════════════════════════════════
    def open_bulk_create(self):
        win = tk.Toplevel(self)
        win.title("📦+ Bulk Create Files & Folders")
        win.geometry(f"{scaled(700)}x{scaled(580)}")
        win.configure(bg=P["BG"])
        win.grab_set()
        win.resizable(True, True)

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(52))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  📦+  Bulk Create — Files & Folder Structure",
                 bg=P["HEADER_BG"], fg=P["ACCENT2"],
                 font=(self.ui_font_name, scaled(13), "bold")).pack(side="left", padx=12, pady=12)

        info = tk.Frame(win, bg=P["BULK_BG"])
        info.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(info, text=(
            "Paste a structure below. Use indentation (spaces or tabs) for nesting.\n"
            "End folder names with  /   •  Files are created automatically  •  Mixed nesting supported"
        ), bg=P["BULK_BG"], fg=P["SUBTEXT"], font=(self.ui_font_name, scaled(8)),
                 justify="left", padx=8, pady=6).pack(anchor="w")

        ex_frame = tk.Frame(win, bg=P["BULK_BG"])
        ex_frame.pack(fill="x", padx=16, pady=4)
        tk.Label(ex_frame, text="Example:", bg=P["BULK_BG"], fg=P["ACCENT"],
                 font=(self.ui_font_name, scaled(8), "bold")).pack(side="left", padx=(8, 6))
        ex_text = "Folder A/\n  file1.txt\n  SubFolder/\n    file2.js\nnotes.md"
        tk.Label(ex_frame, text=ex_text, bg=P["BULK_BG"], fg=P["SUCCESS"],
                 font=("Courier New", scaled(8)), justify="left", padx=8).pack(side="left")

        tk.Label(win, text="Your structure:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(9), "bold")).pack(anchor="w", padx=16, pady=(10, 2))

        text_frame = tk.Frame(win, bg=P["BORDER"])
        text_frame.pack(fill="both", expand=True, padx=16, pady=4)
        input_text = tk.Text(text_frame, bg=P["INPUT_BG"], fg=P["TEXT"],
                             font=("Courier New", scaled(10)), relief="flat", bd=8,
                             insertbackground=P["ACCENT"],
                             selectbackground=P["SEL"],
                             selectforeground=P["SEL_FG"],
                             wrap="none")
        txt_sb = ttk.Scrollbar(text_frame, orient="vertical", command=input_text.yview)
        input_text.configure(yscrollcommand=txt_sb.set)
        input_text.pack(side="left", fill="both", expand=True)
        txt_sb.pack(side="right", fill="y")

        tpl_frame = tk.Frame(win, bg=P["BG"])
        tpl_frame.pack(fill="x", padx=16, pady=4)
        tk.Label(tpl_frame, text="Templates:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(8))).pack(side="left", padx=(0, 6))
        templates = {
            "Python Project": "src/\n  __init__.py\n  main.py\n  utils.py\ntests/\n  test_main.py\nrequirements.txt\nREADME.md\n.gitignore",
            "Web App": "public/\n  index.html\n  style.css\n  app.js\nsrc/\n  components/\n    Header.jsx\npackage.json\nREADME.md",
            "Data Project": "data/\n  raw/\n  processed/\nsrc/\n  analysis.py\nnotebooks/\n  exploration.ipynb\nREADME.md",
        }
        for name, content in templates.items():
            tk.Button(tpl_frame, text=name, bg=P["CARD"], fg=P["ACCENT"],
                      relief="flat", padx=8, pady=3, font=(self.ui_font_name, scaled(8)),
                      cursor="hand2",
                      command=lambda c=content: [
                          input_text.delete("1.0", "end"),
                          input_text.insert("1.0", c)
                      ]).pack(side="left", padx=3)

        dest_frame = tk.Frame(win, bg=P["BG"])
        dest_frame.pack(fill="x", padx=16, pady=6)
        tk.Label(dest_frame, text="Create in:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(9), "bold")).pack(side="left", padx=(0, 8))
        dest_var = tk.StringVar(value=self.current_path.get())
        tk.Entry(dest_frame, textvariable=dest_var, width=42,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=(self.ui_font_name, scaled(9)),
                 insertbackground=P["ACCENT"],
                 selectbackground=P["SEL"],
                 selectforeground=P["SEL_FG"]).pack(side="left", ipady=4)
        tk.Button(dest_frame, text="Browse", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=8, pady=4, font=(self.ui_font_name, scaled(9)),
                  cursor="hand2",
                  command=lambda: dest_var.set(
                      filedialog.askdirectory(title="Create in folder", parent=win)
                      or dest_var.get())).pack(side="left", padx=6)

        status_var = tk.StringVar(value="")
        tk.Label(win, textvariable=status_var, bg=P["BG"], fg=P["SUCCESS"],
                 font=(self.ui_font_name, scaled(9)), anchor="w").pack(fill="x", padx=16)

        btn_frame = tk.Frame(win, bg=P["BG"])
        btn_frame.pack(fill="x", padx=16, pady=(4, 14))

        def do_preview():
            text = input_text.get("1.0", "end").strip()
            if not text:
                messagebox.showwarning("Empty", "Enter a file/folder structure.", parent=win)
                return
            base = dest_var.get()
            items = parse_bulk_structure(text, base)
            preview_lines = []
            for abs_path, is_dir in items:
                rel = os.path.relpath(abs_path, base)
                icon = "📁" if is_dir else file_icon(abs_path)
                preview_lines.append(f"{icon} {rel}")
            messagebox.showinfo("Preview — Items to Create",
                                "\n".join(preview_lines[:50]) +
                                (f"\n... +{len(preview_lines)-50} more" if len(preview_lines) > 50 else ""),
                                parent=win)

        def do_create():
            text = input_text.get("1.0", "end").strip()
            if not text:
                messagebox.showwarning("Empty", "Enter a file/folder structure.", parent=win)
                return
            base = dest_var.get()
            if not os.path.isdir(base):
                messagebox.showerror("Invalid", f"Destination not found:\n{base}", parent=win)
                return
            items = parse_bulk_structure(text, base)
            created_dirs = created_files = 0
            errors = []
            for abs_path, is_dir in items:
                try:
                    if is_dir:
                        os.makedirs(abs_path, exist_ok=True)
                        created_dirs += 1
                    else:
                        os.makedirs(os.path.dirname(abs_path), exist_ok=True)
                        Path(abs_path).touch()
                        created_files += 1
                except Exception as e:
                    errors.append(f"{os.path.basename(abs_path)}: {e}")
            msg = f"✅ Created {created_dirs} folder(s), {created_files} file(s)"
            if errors:
                msg += f" | {len(errors)} error(s)"
            status_var.set(msg)
            self.load_dir(base)
            if errors:
                messagebox.showerror("Errors", "\n".join(errors[:10]), parent=win)

        tk.Button(btn_frame, text="👁 Preview", bg=P["SUBTEXT"], fg="white",
                  relief="flat", padx=16, pady=8, font=(self.ui_font_name, scaled(10), "bold"),
                  cursor="hand2", command=do_preview).pack(side="left", padx=4)
        tk.Button(btn_frame, text="✅ Create All", bg=P["SUCCESS"], fg="white",
                  relief="flat", padx=20, pady=8, font=(self.ui_font_name, scaled(11), "bold"),
                  cursor="hand2", command=do_create).pack(side="left", padx=4)
        tk.Button(btn_frame, text="Cancel", bg=P["CARD"], fg=P["TEXT"],
                  relief="flat", padx=14, pady=8, font=(self.ui_font_name, scaled(9)),
                  cursor="hand2", command=win.destroy).pack(side="left", padx=4)

    # ── Quick type filter ──────────────────────────────────────────────────────
    def _quick_type_filter(self, val):
        self._quick_type_filter_val = val
        self._populate(self.current_path.get(), self.search_bar.get())

    # ── Icons View ─────────────────────────────────────────────────────────────
    def _build_icons_view(self, entries):
        for w in self.list_panel.winfo_children():
            w.destroy()

        canvas = tk.Canvas(self.list_panel, bg=P["PANEL"], highlightthickness=0)
        vsb = ttk.Scrollbar(self.list_panel, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        inner = tk.Frame(canvas, bg=P["PANEL"])
        canvas.create_window((0, 0), window=inner, anchor="nw")

        ICON_W = scaled(120)
        ICON_H = scaled(100)

        def make_icon_cell(parent, entry):
            path = entry.path
            icon_txt = file_icon(path)
            mark = self.color_mark_manager.get_mark(path)
            mc = MARK_COLORS.get(mark, {"bg": None, "fg": None})
            cell_bg = mc["bg"] if mark != "none" and mc["bg"] else P["PANEL"]

            frame = tk.Frame(parent, bg=cell_bg, width=ICON_W,
                             highlightthickness=1, highlightbackground=P["BORDER"],
                             cursor="hand2")
            frame.pack_propagate(False)

            icon_lbl = tk.Label(frame, text=icon_txt, bg=cell_bg,
                                font=("Segoe UI Emoji", scaled(30)), pady=scaled(6))
            icon_lbl.pack()

            raw_name = entry.name
            display = raw_name if len(raw_name) <= 13 else raw_name[:11] + "…"
            fg_color = P["ACCENT"] if entry.is_dir() else P["TEXT"]
            name_lbl = tk.Label(frame, text=display,
                                bg=cell_bg, fg=fg_color,
                                font=(self.ui_font_name, scaled(8)),
                                wraplength=ICON_W - 8, justify="center")
            name_lbl.pack(fill="x", padx=2, pady=(0, 4))

            if len(raw_name) > 13:
                Tooltip(name_lbl, raw_name)
                Tooltip(icon_lbl, raw_name)

            tags_list = self.tag_manager.get_tags(path)
            if tags_list:
                tk.Label(frame, text="🏷️", bg=cell_bg,
                         font=("Segoe UI Emoji", scaled(8))).pack()

            def on_click(e, p=path):
                if entry.is_dir():
                    self.load_dir(p)
                else:
                    self._open_file(p)

            def on_hover_enter(e, f=frame, lbl=name_lbl, ibg=cell_bg):
                f.config(highlightbackground=P["ACCENT"])

            def on_hover_leave(e, f=frame, lbl=name_lbl, ibg=cell_bg):
                f.config(highlightbackground=P["BORDER"])

            def on_right_click(e, p=path):
                self.selected_paths = [p]
                self._context_menu_for_path(e, p)

            for w in [frame, icon_lbl, name_lbl]:
                w.bind("<Double-1>", on_click)
                w.bind("<Button-3>", on_right_click)
                w.bind("<Enter>", on_hover_enter)
                w.bind("<Leave>", on_hover_leave)
            return frame

        dirs = [e for e in entries if e.is_dir(follow_symlinks=False)]
        files = [e for e in entries if not e.is_dir(follow_symlinks=False)]
        all_cells = []
        for entry in dirs + files:
            cell = make_icon_cell(inner, entry)
            all_cells.append(cell)

        def relayout(event=None):
            w = canvas.winfo_width()
            if w < 10:
                return
            cols = max(2, w // (ICON_W + 12))
            for child in inner.winfo_children():
                child.grid_forget()
            for idx, cell in enumerate(all_cells):
                r, c = divmod(idx, cols)
                cell.config(width=ICON_W, height=ICON_H)
                cell.grid(row=r, column=c, padx=6, pady=6, sticky="n")
            inner.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))

        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", relayout)
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))
        canvas.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))
        self.after(50, relayout)
        self.tree = None

    # ── Tree / Details View ────────────────────────────────────────────────────
    def _build_tree_view(self):
        for w in self.list_panel.winfo_children():
            w.destroy()
        cols = ("icon", "name", "size", "type", "modified", "perms", "tags")
        self.tree = ttk.Treeview(self.list_panel, columns=cols, show="headings",
                                 selectmode="extended")

        for col, label, w, anchor in [
            ("icon", "", scaled(34), "center"),
            ("name", "Name", scaled(300), "w"),
            ("size", "Size", scaled(85), "e"),
            ("type", "Type", scaled(100), "w"),
            ("modified", "Modified", scaled(155), "w"),
            ("perms", "Permissions", scaled(95), "w"),
            ("tags", "Tags", scaled(110), "w"),
        ]:
            self.tree.heading(col, text=label, command=lambda c=col: self._sort_by(c))
            self.tree.column(col, width=w, anchor=anchor, stretch=(col == "name"))

        # Theme-aware tag configuration
        self.tree.tag_configure("even", background=P["EVEN_ROW"], foreground=P["TEXT"])
        self.tree.tag_configure("odd", background=P["ODD_ROW"], foreground=P["TEXT"])
        self.tree.tag_configure("dir", foreground=P["ACCENT"])
        self.tree.tag_configure("exec", foreground=P["SUCCESS"])
        self.tree.tag_configure("link", foreground=P["ACCENT2"])
        self.tree.tag_configure("tagged", foreground=P["TAG_FG"])

        for cname, cdata in MARK_COLORS.items():
            if cname != "none" and cdata["bg"]:
                self.tree.tag_configure(f"mark_{cname}",
                                        background=cdata["bg"],
                                        foreground=cdata["fg"])

        vsb = ttk.Scrollbar(self.list_panel, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.list_panel, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Return>", self._on_double_click)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Button-3>", self._context_menu)
        self.tree.bind("<ButtonPress-1>", self._on_mouse_press)
        self.tree.bind("<B1-Motion>", self._on_mouse_drag)
        self.tree.bind("<ButtonRelease-1>", self._on_mouse_release)
        self.tree.bind("<F2>", lambda e: self._start_inline_rename())

    def _set_view_mode(self, mode):
        self.view_mode.set(mode)
        self._save_prefs()
        self._populate(self.current_path.get(), self.search_bar.get())

    # ── Inline Rename ──────────────────────────────────────────────────────────
    def _start_inline_rename(self):
        if not self.tree:
            self.rename_file()
            return
        sel = self.tree.selection()
        if not sel:
            return
        item = sel[0]

        def on_confirm(new_name):
            old_path = item
            new_path = os.path.join(os.path.dirname(old_path), new_name)
            try:
                os.rename(old_path, new_path)
                self.undo_stack.push(("rename", new_path, old_path))
                self.load_dir(self.current_path.get())
                self.status_var.set(f"Renamed → {new_name}")
            except Exception as e:
                messagebox.showerror("Rename Error", str(e))

        def on_cancel():
            self._inline_rename_widget = None

        self._inline_rename_widget = InlineRenameEntry(self.tree, item, on_confirm, on_cancel)

    # ── Font Settings ──────────────────────────────────────────────────────────
    def open_font_settings(self):
        win = tk.Toplevel(self)
        win.title("🔤 Font & Interface Settings")
        win.geometry(f"{scaled(540)}x{scaled(480)}")
        win.configure(bg=P["BG"])
        win.grab_set()
        win.resizable(False, False)

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(46))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  🔤  Font & Interface Customization",
                 bg=P["HEADER_BG"], fg=P["ACCENT"],
                 font=(self.ui_font_name, scaled(12), "bold")).pack(side="left", padx=12, pady=10)

        body = tk.Frame(win, bg=P["BG"])
        body.pack(fill="both", expand=True, padx=20, pady=16)

        tk.Label(body, text="Font Family:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10), "bold")).grid(row=0, column=0, sticky="w", pady=8)

        available_fonts = sorted(set([
            "Segoe UI", "Consolas", "Courier New", "Lucida Console", "Monaco", "Menlo",
            "DejaVu Sans Mono", "Source Code Pro", "Fira Code",
            "Arial", "Calibri", "Helvetica", "Verdana", "Tahoma",
        ] + list(tkfont.families())))

        font_var = tk.StringVar(value=self.ui_font_name)
        font_cb = ttk.Combobox(body, textvariable=font_var, values=available_fonts,
                               width=28, state="readonly", font=(self.ui_font_name, scaled(10)))
        font_cb.grid(row=0, column=1, sticky="w", padx=12, pady=8)

        tk.Label(body, text="Font Size:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10), "bold")).grid(row=1, column=0, sticky="w", pady=8)
        size_var = tk.IntVar(value=self.ui_font_size)
        size_frame = tk.Frame(body, bg=P["BG"])
        size_frame.grid(row=1, column=1, sticky="w", padx=12, pady=8)

        def dec_size():
            v = size_var.get()
            if v > 7: size_var.set(v - 1)
            update_preview()

        def inc_size():
            v = size_var.get()
            if v < 20: size_var.set(v + 1)
            update_preview()

        tk.Button(size_frame, text="−", bg=P["DANGER"], fg="white",
                  relief="flat", padx=10, pady=4, font=(self.ui_font_name, scaled(10), "bold"),
                  cursor="hand2", command=dec_size).pack(side="left")
        tk.Label(size_frame, textvariable=size_var, bg=P["INPUT_BG"],
                 fg=P["TEXT"], width=4, font=(self.ui_font_name, scaled(12), "bold"),
                 relief="flat", pady=4).pack(side="left", padx=4)
        tk.Button(size_frame, text="+", bg=P["SUCCESS"], fg="white",
                  relief="flat", padx=10, pady=4, font=(self.ui_font_name, scaled(10), "bold"),
                  cursor="hand2", command=inc_size).pack(side="left")

        tk.Label(body, text="Quick Sizes:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10))).grid(row=2, column=0, sticky="w", pady=4)
        preset_frame = tk.Frame(body, bg=P["BG"])
        preset_frame.grid(row=2, column=1, sticky="w", padx=12, pady=4)
        for ps in [8, 9, 10, 11, 12, 14]:
            tk.Button(preset_frame, text=str(ps), bg=P["CARD"], fg=P["TEXT"],
                      relief="flat", padx=8, pady=3, font=(self.ui_font_name, scaled(9)),
                      cursor="hand2",
                      command=lambda s=ps: [size_var.set(s), update_preview()]
                      ).pack(side="left", padx=2)

        tk.Label(body, text="Preview:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10), "bold")).grid(row=3, column=0, sticky="nw", pady=8)
        preview_frame = tk.Frame(body, bg=P["CARD_BORDER"], bd=0)
        preview_frame.grid(row=3, column=1, sticky="ew", padx=12, pady=8)
        preview_lbl = tk.Label(preview_frame,
                               text="AaBbCcDd 0123456789\n📁 My Documents  📄 file.txt\n✂ Cut  📋 Copy  📥 Paste",
                               bg=P["CARD"], fg=P["TEXT"], justify="left",
                               padx=12, pady=10, anchor="w",
                               font=(self.ui_font_name, self.ui_font_size))
        preview_lbl.pack(fill="x", padx=1, pady=1)

        def update_preview(*_):
            try:
                preview_lbl.config(font=(font_var.get(), size_var.get()))
            except: pass
        font_cb.bind("<<ComboboxSelected>>", update_preview)

        tk.Label(body, text="Export Folder:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10), "bold")).grid(row=4, column=0, sticky="w", pady=8)
        export_dir_var = tk.StringVar(value=self.export_dir)
        ed_frame = tk.Frame(body, bg=P["BG"])
        ed_frame.grid(row=4, column=1, sticky="ew", padx=12, pady=8)
        tk.Entry(ed_frame, textvariable=export_dir_var, width=26,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=(self.ui_font_name, scaled(9)),
                 insertbackground=P["ACCENT"],
                 selectbackground=P["SEL"]).pack(side="left")
        tk.Button(ed_frame, text="Browse", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=6, pady=2, font=(self.ui_font_name, scaled(9)),
                  cursor="hand2",
                  command=lambda: export_dir_var.set(
                      filedialog.askdirectory(title="Choose Export Folder", parent=win)
                      or export_dir_var.get())).pack(side="left", padx=4)

        btn_frame = tk.Frame(win, bg=P["BG"])
        btn_frame.pack(fill="x", padx=20, pady=(0, 16))

        def apply_font():
            self.ui_font_name = font_var.get()
            self.ui_font_size = size_var.get()
            new_dir = export_dir_var.get().strip()
            if new_dir: self.export_dir = new_dir
            self._save_prefs()
            win.destroy()
            self.rebuild_ui()

        tk.Button(btn_frame, text="✅ Apply & Restart UI", bg=P["SUCCESS"], fg="white",
                  relief="flat", padx=16, pady=8, font=(self.ui_font_name, scaled(10), "bold"),
                  cursor="hand2", command=apply_font).pack(side="left", padx=4)
        tk.Button(btn_frame, text="Cancel", bg=P["CARD"], fg=P["TEXT"],
                  relief="flat", padx=12, pady=8, font=(self.ui_font_name, scaled(9)),
                  cursor="hand2", command=win.destroy).pack(side="left", padx=4)

    # ══════════════════════════════════════════════════════════════════════════
    # ── SUPPORT / DONATE DIALOG ──────────────────────────────────────────────
    # ══════════════════════════════════════════════════════════════════════════
    def show_support_info(self):
        win = tk.Toplevel(self)
        win.title("💙 Support ")
        win.geometry(f"{scaled(420)}x{scaled(480)}")
        win.configure(bg=P["BG"])
        win.resizable(False, False)
        win.grab_set()

        x = self.winfo_rootx() + (self.winfo_width() - scaled(420)) // 2
        y = self.winfo_rooty() + (self.winfo_height() - scaled(480)) // 2
        win.geometry(f"{scaled(420)}x{scaled(480)}+{x}+{y}")

        # Top accent strip
        tk.Frame(win, bg=P["ACCENT"], height=4).pack(fill="x")

        # Hero section
        hero = tk.Frame(win, bg=P["HEADER_BG"])
        hero.pack(fill="x")
        tk.Label(hero, text="💙", bg=P["HEADER_BG"],
                 font=("Segoe UI Emoji", scaled(32))).pack(pady=(scaled(16), 4))
        tk.Label(hero, text="Support NexFile Pro",
                 bg=P["HEADER_BG"], fg=P["TEXT"],
                 font=(self.ui_font_name, scaled(15), "bold")).pack()
        tk.Label(hero, text="If you find this tool helpful, your support keeps\ndevelopment going and new features coming!",
                 bg=P["HEADER_BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(9)), justify="center").pack(pady=(4, scaled(14)))

        # Developer card
        dev_card_outer = tk.Frame(win, bg=P["CARD_BORDER"])
        dev_card_outer.pack(fill="x", padx=16, pady=8)
        dev_card = tk.Frame(dev_card_outer, bg=P["CARD"], padx=16, pady=12)
        dev_card.pack(fill="x", padx=1, pady=1)

        tk.Label(dev_card, text="👨‍💻  Developer", bg=P["CARD"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(8), "bold"), anchor="w").pack(fill="x")
        tk.Label(dev_card, text="Asif Miah", bg=P["CARD"], fg=P["TEXT"],
                 font=(self.ui_font_name, scaled(12), "bold"), anchor="w").pack(fill="x")
        tk.Frame(dev_card, bg=P["BORDER"], height=1).pack(fill="x", pady=8)

        for icon, label, val in [
            ("📞", "Phone / WhatsApp", "01677460147"),
            ("✉️", "Email", "asifmiah12@gmail.com"),
        ]:
            row = tk.Frame(dev_card, bg=P["CARD"])
            row.pack(fill="x", pady=3)
            tk.Label(row, text=icon, bg=P["CARD"],
                     font=("Segoe UI Emoji", scaled(13)), width=3).pack(side="left")
            lbl_frame = tk.Frame(row, bg=P["CARD"])
            lbl_frame.pack(side="left", fill="x", expand=True)
            tk.Label(lbl_frame, text=label, bg=P["CARD"], fg=P["SUBTEXT"],
                     font=(self.ui_font_name, scaled(8)), anchor="w").pack(anchor="w")
            tk.Label(lbl_frame, text=val, bg=P["CARD"], fg=P["TEXT"],
                     font=(self.ui_font_name, scaled(10), "bold"), anchor="w").pack(anchor="w")

        # Payment methods card
        pay_outer = tk.Frame(win, bg=P["CARD_BORDER"])
        pay_outer.pack(fill="x", padx=16, pady=4)
        pay_card = tk.Frame(pay_outer, bg=P["CARD"], padx=16, pady=12)
        pay_card.pack(fill="x", padx=1, pady=1)

        tk.Label(pay_card, text="💳  Payment Details", bg=P["CARD"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(8), "bold"), anchor="w").pack(fill="x")
        tk.Label(pay_card, text="01735144475",
                 bg=P["CARD"], fg=P["ACCENT"],
                 font=(self.ui_font_name, scaled(16), "bold"), anchor="w").pack(fill="x", pady=(4, 8))

        tk.Label(pay_card, text="Accepted Payment Methods:", bg=P["CARD"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(8)), anchor="w").pack(fill="x", pady=(0, 6))
        
        tk.Label(pay_card, text="Bkas/Rocket/Nagad", bg=P["CARD"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10)), anchor="w").pack(fill="x", pady=(1, 6))

        methods_frame = tk.Frame(pay_card, bg=P["CARD"])
        methods_frame.pack(fill="x")

        payment_methods = [
             ("bKash", "#e2136e", "#ffffff", "💗"),
             ("Nagad", "#f05a28", "#ffffff", "🧡"),
             ("Rocket", "#8b5cf6", "#ffffff", "💜"),
        ]
        for name, bg_col, fg_col, icon in payment_methods:
            btn_frame = tk.Frame(methods_frame, bg=bg_col)
            btn_frame.pack(side="left", padx=(0, 8), pady=2)
            inner = tk.Frame(btn_frame, bg=bg_col, padx=14, pady=8)
            inner.pack(padx=1, pady=1)
            tk.Label(inner, text=icon, bg=bg_col, fg=fg_col,
                     font=("Segoe UI Emoji", scaled(14))).pack()
            tk.Label(inner, text=name, bg=bg_col, fg=fg_col,
                     font=(self.ui_font_name, scaled(9), "bold")).pack()

        # Buttons
        btn_row = tk.Frame(win, bg=P["BG"])
        btn_row.pack(fill="x", padx=16, pady=12)

        def copy_number():
            self.clipboard_clear()
            self.clipboard_append("01735144475")
            btn_copy_num.config(text="✅ Copied!")
            win.after(2000, lambda: btn_copy_num.config(text="📋 Copy Number"))

        def copy_email():
            self.clipboard_clear()
            self.clipboard_append("asifmiah12@gmail.com")
            btn_copy_email.config(text="✅ Copied!")
            win.after(2000, lambda: btn_copy_email.config(text="✉️ Copy Email"))

        btn_copy_num = tk.Button(btn_row, text="📋 Copy Number", bg=P["ACCENT"], fg="white",
                                  relief="flat", padx=12, pady=8,
                                  font=(self.ui_font_name, scaled(9), "bold"),
                                  cursor="hand2", command=copy_number)
        btn_copy_num.pack(side="left", padx=(0, 6))

        btn_copy_email = tk.Button(btn_row, text="✉️ Copy Email", bg=P["CARD"], fg=P["TEXT"],
                                   relief="flat", padx=12, pady=8,
                                   font=(self.ui_font_name, scaled(9)),
                                   cursor="hand2", command=copy_email)
        btn_copy_email.pack(side="left", padx=(0, 6))

        tk.Button(btn_row, text="Close", bg=P["CARD"], fg=P["SUBTEXT"],
                  relief="flat", padx=12, pady=8,
                  font=(self.ui_font_name, scaled(9)),
                  cursor="hand2", command=win.destroy).pack(side="right")

        tk.Label(win, text="Thank you for your support! 🙏",
                 bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(8))).pack(pady=(0, 8))

    # ══════════════════════════════════════════════════════════════════════════
    # ── EXPORT DIALOGS ────────────────────────────────────────────────────────
    # ══════════════════════════════════════════════════════════════════════════
    def _open_export_dialog(self, export_format):
        fmt_upper = export_format.upper()
        fmt_icon = "📑" if export_format == "pdf" else "📊"
        fmt_color = "#e11d48" if export_format == "pdf" else P["SUCCESS"]

        if export_format == "pdf" and not PDF_AVAILABLE:
            messagebox.showerror("Missing Library", "PDF export requires reportlab.\n\npip install reportlab")
            return
        if export_format == "excel" and not EXCEL_AVAILABLE:
            messagebox.showerror("Missing Library", "Excel export requires openpyxl.\n\npip install openpyxl")
            return

        win = tk.Toplevel(self)
        win.title(f"{fmt_icon} Export as {fmt_upper}")
        win.geometry(f"{scaled(660)}x{scaled(580)}")
        win.configure(bg=P["BG"])
        win.grab_set()

        hdr = tk.Frame(win, bg=P["HEADER_BG"], height=scaled(50))
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"  {fmt_icon}  Export File List as {fmt_upper}",
                 bg=P["HEADER_BG"], fg=fmt_color,
                 font=(self.ui_font_name, scaled(13), "bold")).pack(side="left", padx=12, pady=12)

        body = tk.Frame(win, bg=P["BG"])
        body.pack(fill="both", expand=True, padx=20, pady=12)

        tk.Label(body, text="Export Scope:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(10), "bold")).pack(anchor="w", pady=(0, 4))
        scope_frame = tk.Frame(body, bg=P["CARD"])
        scope_frame.pack(fill="x", pady=(0, 10))
        scope_var = tk.StringVar(value="current")
        for val, lbl in [
            ("current", "📁  Current folder only"),
            ("recursive", "📂  Current folder + all subfolders"),
            ("drive", "💾  Entire drive"),
        ]:
            tk.Radiobutton(scope_frame, text=lbl, variable=scope_var, value=val,
                           bg=P["CARD"], fg=P["TEXT"], selectcolor=P["CARD"],
                           activebackground=P["CARD"], font=(self.ui_font_name, scaled(9)),
                           pady=6).pack(anchor="w", padx=12)

        tk.Label(body, text="File type filter (ext, blank = all):", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(9), "bold")).pack(anchor="w")
        ext_input_var = tk.StringVar()
        tk.Entry(body, textvariable=ext_input_var, width=40,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=(self.ui_font_name, scaled(10)),
                 insertbackground=P["ACCENT"],
                 selectbackground=P["SEL"]).pack(anchor="w", pady=4, ipady=4)

        tk.Label(body, text="Save to folder:", bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(9), "bold")).pack(anchor="w", pady=(10, 2))
        loc_frame = tk.Frame(body, bg=P["BG"])
        loc_frame.pack(fill="x", pady=2)
        export_loc_var = tk.StringVar(value=self.export_dir)
        tk.Entry(loc_frame, textvariable=export_loc_var, width=46,
                 bg=P["INPUT_BG"], fg=P["TEXT"], relief="flat",
                 font=(self.ui_font_name, scaled(9)),
                 insertbackground=P["ACCENT"],
                 selectbackground=P["SEL"]).pack(side="left", ipady=4)
        tk.Button(loc_frame, text="Browse", bg=P["ACCENT"], fg="white",
                  relief="flat", padx=8, pady=4, font=(self.ui_font_name, scaled(9)),
                  cursor="hand2",
                  command=lambda: export_loc_var.set(
                      filedialog.askdirectory(title="Save exports to...", parent=win)
                      or export_loc_var.get())).pack(side="left", padx=6)

        incl_hidden = tk.BooleanVar(value=False)
        tk.Checkbutton(body, text="Include hidden files", variable=incl_hidden,
                       bg=P["BG"], fg=P["SUBTEXT"], selectcolor=P["BG"],
                       activebackground=P["BG"],
                       font=(self.ui_font_name, scaled(9))).pack(anchor="w", pady=4)

        status_var = tk.StringVar(value="Configure options above and click Export.")
        tk.Label(body, textvariable=status_var, bg=P["BG"], fg=P["SUBTEXT"],
                 font=(self.ui_font_name, scaled(8)), anchor="w", wraplength=scaled(580)).pack(fill="x", pady=(8, 2))
        progress = ttk.Progressbar(body, mode="indeterminate",
                                   style="Success.Horizontal.TProgressbar")
        progress.pack(fill="x", pady=4)
        result_label = tk.Label(body, text="", bg=P["BG"], fg=P["SUCCESS"],
                                font=(self.ui_font_name, scaled(9), "bold"), anchor="w")
        result_label.pack(fill="x", pady=2)

        btn_frame = tk.Frame(win, bg=P["BG"])
        btn_frame.pack(fill="x", padx=20, pady=(4, 16))
        export_btn = tk.Button(btn_frame, text=f"{fmt_icon} Export as {fmt_upper}",
                               bg=fmt_color, fg="white", relief="flat",
                               padx=20, pady=10,
                               font=(self.ui_font_name, scaled(11), "bold"), cursor="hand2")
        export_btn.pack(side="left", padx=4)
        tk.Button(btn_frame, text="Cancel", bg=P["CARD"], fg=P["TEXT"],
                  relief="flat", padx=14, pady=10, font=(self.ui_font_name, scaled(9)),
                  cursor="hand2", command=win.destroy).pack(side="left", padx=4)
        open_folder_btn = tk.Button(btn_frame, text="📂 Open Export Folder",
                                    bg=P["CARD"], fg=P["ACCENT"], relief="flat",
                                    padx=12, pady=10, font=(self.ui_font_name, scaled(9)),
                                    cursor="hand2", state="disabled")
        open_folder_btn.pack(side="right", padx=4)

        def open_export_folder():
            path = export_loc_var.get()
            try:
                if sys.platform == "win32": os.startfile(path)
                elif sys.platform == "darwin": subprocess.Popen(["open", path])
                else: subprocess.Popen(["xdg-open", path])
            except: pass

        open_folder_btn.config(command=open_export_folder)
        stop_event = threading.Event()

        def do_export():
            scope = scope_var.get()
            ext_filter_raw = ext_input_var.get().strip()
            save_dir = export_loc_var.get().strip()
            show_hid = incl_hidden.get()
            if not save_dir:
                messagebox.showwarning("No Folder", "Please set a save folder.", parent=win)
                return
            if not ensure_export_dir(save_dir):
                messagebox.showerror("Error", f"Cannot create folder:\n{save_dir}", parent=win)
                return
            cur = self.current_path.get()
            if scope == "drive":
                root_path = cur[:3] if sys.platform == "win32" else "/"
                folder_label = "FullDrive"
            else:
                root_path = cur
                folder_label = os.path.basename(cur) or "Root"
            ext_list = [e.strip().lstrip(".").lower()
                        for e in ext_filter_raw.replace(",", " ").split() if e.strip()]
            out_ext = "pdf" if export_format == "pdf" else "xlsx"
            out_path = make_export_filename(folder_label, out_ext, save_dir)
            export_btn.config(state="disabled")
            stop_event.clear()
            progress.start(12)
            status_var.set(f"Scanning '{root_path}'...")
            result_label.config(text="", fg=P["SUCCESS"])

            def run():
                try:
                    if scope == "current":
                        file_data = []
                        for entry in os.scandir(root_path):
                            if stop_event.is_set(): break
                            if entry.is_file(follow_symlinks=False):
                                if not show_hid and entry.name.startswith("."): continue
                                if ext_list and Path(entry.name).suffix.lower().lstrip(".") not in ext_list:
                                    continue
                                try:
                                    st = entry.stat()
                                    file_data.append({
                                        "name": entry.name, "full_path": entry.path,
                                        "size": human_size(st.st_size), "size_bytes": st.st_size,
                                        "modified": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                                        "type": Path(entry.name).suffix.upper().lstrip(".") + " File" if Path(entry.name).suffix else "File",
                                    })
                                except: pass
                    else:
                        def prog_cb(msg):
                            win.after(0, lambda m=msg: status_var.set(m[:90]))
                        all_data = scan_directory_recursive(root_path, show_hidden=show_hid,
                                                            stop_event=stop_event, progress_cb=prog_cb)
                        file_data = [d for d in all_data if Path(d["name"]).suffix.lower().lstrip(".") in ext_list] if ext_list else all_data

                    if not file_data:
                        win.after(0, lambda: [progress.stop(), status_var.set("No files found."),
                                              export_btn.config(state="normal")])
                        return

                    win.after(0, lambda n=len(file_data): status_var.set(f"Writing {n} rows..."))
                    if export_format == "pdf":
                        _write_pdf(file_data, out_path, folder_label, root_path)
                    else:
                        _write_excel(file_data, out_path, folder_label, root_path)

                    def done():
                        progress.stop()
                        export_btn.config(state="normal")
                        open_folder_btn.config(state="normal")
                        result_label.config(
                            text=f"✅  Saved {len(file_data)} entries → {os.path.basename(out_path)}",
                            fg=P["SUCCESS"])
                        status_var.set(f"Export complete: {out_path}")

                    win.after(0, done)
                except Exception as e:
                    def err(msg=str(e)):
                        progress.stop()
                        export_btn.config(state="normal")
                        result_label.config(text=f"❌  Error: {msg}", fg=P["DANGER"])
                    win.after(0, err)

            threading.Thread(target=run, daemon=True).start()

        export_btn.config(command=do_export)
        win.protocol("WM_DELETE_WINDOW", lambda: [stop_event.set(), win.destroy()])

    def open_export_dialog_pdf(self):
        self._open_export_dialog("pdf")

    def open_export_dialog_excel(self):
        self._open_export_dialog("excel")

    # ── Cleanup ────────────────────────────────────────────────────────────────
    def on_closing(self):
        if hasattr(self, "_sys_search_stop_event"):
            self._sys_search_stop_event.set()
        self._save_prefs()
        self._stop_watcher()
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# ── EXPORT FUNCTIONS ──────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

def _write_excel(file_data, out_path, folder_label, root_path):
    if not EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "File Export"
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"NexFile Pro — File Export: {folder_label}"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0969da", end_color="0969da", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A2:E2")
    meta = ws["A2"]
    meta.value = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
                  f"Root: {root_path}  |  Files: {len(file_data)}")
    meta.font = Font(size=9, color="888888")
    meta.alignment = Alignment(horizontal="center")
    ws.append([])
    headers = ["File Name", "Full Path", "Size", "Last Modified", "Type"]
    header_row = 4
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="0969da", end_color="0969da", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 20
    for i, row in enumerate(file_data):
        r = header_row + 1 + i
        values = [row["name"], row["full_path"], row["size"], row["modified"], row["type"]]
        fill_hex = "f6f8fa" if i % 2 == 0 else "ffffff"
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = Font(size=9)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 2))
    col_widths = [50, 80, 14, 22, 16]
    for col_idx, width in enumerate(col_widths, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:E{header_row + len(file_data)}"
    wb.save(out_path)


def _write_pdf(file_data, out_path, folder_label, root_path):
    if not PDF_AVAILABLE:
        raise RuntimeError("reportlab not installed")
    doc = SimpleDocTemplate(
        out_path, pagesize=landscape(A4),
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.6 * inch, bottomMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("NexTitle", parent=styles["Normal"],
                                 fontSize=14, fontName="Helvetica-Bold",
                                 textColor=colors.HexColor("#0969da"), spaceAfter=4)
    meta_style = ParagraphStyle("NexMeta", parent=styles["Normal"],
                                fontSize=8, fontName="Helvetica",
                                textColor=colors.HexColor("#57606a"), spaceAfter=10)
    cell_style = ParagraphStyle("NexCell", parent=styles["Normal"],
                                fontSize=7.5, fontName="Helvetica", leading=10)
    elements = [
        Paragraph(f"NexFile Pro — File Export: {folder_label}", title_style),
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
                  f"Root: {root_path}  |  Files: {len(file_data)}", meta_style),
        Spacer(1, 0.1 * inch),
    ]
    CHUNK = 500
    headers = ["#", "File Name", "Full Path", "Size", "Modified", "Type"]
    header_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0969da")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d0d7de")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    col_widths_pdf = [0.35 * inch, 2.2 * inch, 3.8 * inch, 0.9 * inch, 1.4 * inch, 1.0 * inch]
    for chunk_start in range(0, len(file_data), CHUNK):
        chunk = file_data[chunk_start:chunk_start + CHUNK]
        table_data = [headers]
        for i, row in enumerate(chunk):
            fpath = row["full_path"]
            if len(fpath) > 70: fpath = "..." + fpath[-67:]
            table_data.append([
                str(chunk_start + i + 1),
                Paragraph(row["name"], cell_style),
                Paragraph(fpath, cell_style),
                row["size"], row["modified"], row["type"],
            ])
        tbl = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
        row_styles = list(header_style)
        for idx in range(1, len(chunk) + 1):
            bg = colors.HexColor("#f6f8fa") if idx % 2 == 1 else colors.HexColor("#ffffff")
            row_styles.append(("BACKGROUND", (0, idx), (-1, idx), bg))
            row_styles.append(("TEXTCOLOR", (0, idx), (-1, idx), colors.HexColor("#1f2328")))
        tbl.setStyle(TableStyle(row_styles))
        elements.append(tbl)
        if chunk_start + CHUNK < len(file_data):
            elements.append(PageBreak())

    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#57606a"))
        canvas.drawString(0.5 * inch, 0.3 * inch,
                          f"NexFile Pro Export  |  {folder_label}  |  Page {doc.page}")
        canvas.drawRightString(landscape(A4)[0] - 0.5 * inch, 0.3 * inch,
                               f"{len(file_data)} files  |  {datetime.now().strftime('%Y-%m-%d')}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)


# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = FileManager()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()